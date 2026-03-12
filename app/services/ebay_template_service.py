import os
import hashlib
from fastapi import UploadFile, HTTPException
from sqlalchemy.orm import Session, selectinload
from datetime import datetime
from typing import Optional, Dict, List, Set, Any
import openpyxl
import re

from app.models.templates import (
    EbayTemplate,
    EbayField,
    TemplateField,
    TemplateFieldAsset,
    TemplateFieldAssetEquipmentType,
)
# IMPORTANT: valid values model name differs in some codebases.
# We'll resolve it safely below.
from app.schemas.templates import EbayTemplateParseSummary
from app.services.storage_policy import (
    ensure_storage_dirs_exist,
    assert_allowed_write_path,
    get_ebay_template_paths,
    rotate_ebay_template_backup,
)


def _normalize_sheet_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name or "").strip().lower())


def _pick_sheet_name(sheet_names: list[str], desired: str) -> Optional[str]:
    desired_norm = _normalize_sheet_name(desired)
    for sheet_name in sheet_names:
        if _normalize_sheet_name(sheet_name) == desired_norm:
            return sheet_name
    for sheet_name in sheet_names:
        if desired_norm and desired_norm in _normalize_sheet_name(sheet_name):
            return sheet_name
    return None


DESCRIPTION_NORM_KEY = "description"
IMAGE_NORM_KEYS = {
    "itemphotourl",
    "itemphotourls",
    "pictureurl",
    "pictureurls",
    "photourl",
    "photourls",
    "imageurl",
    "imageurls",
}
SUPPORTED_TEMPLATE_FIELD_KEYS = {DESCRIPTION_NORM_KEY, *IMAGE_NORM_KEYS}


def _resolve_ebay_valid_value_model():
    """
    Resolve the ORM model used for EbayField.valid_values.
    Different codebases name it differently.
    We try common names in a safe order.
    """
    # Late imports so this file can still load even if a name doesn't exist.
    from app.models import templates as templates_models

    # Most common names (try in order)
    for name in ("EbayFieldValue", "EbayFieldValidValue", "EbayValidValue"):
        if hasattr(templates_models, name):
            return getattr(templates_models, name)

    raise RuntimeError(
        "Could not find a valid-values ORM model. Expected one of: "
        "EbayFieldValue, EbayFieldValidValue, EbayValidValue in app.models.templates"
    )


class EbayTemplateService:
    def __init__(self, db: Session):
        self.db = db
        self.ValidValueModel = _resolve_ebay_valid_value_model()

    async def store_ebay_template_upload(self, file: UploadFile) -> Dict[str, Any]:
        """
        Store an eBay template upload bit-for-bit.

        Steps:
        1. Read bytes & compute SHA256 upload
        2. Rotate existing file to backup
        3. Write new file to canonical path
        4. Re-read back from disk & compute SHA256 persisted
        5. Verify match
        6. Create DB record
        """
        if not file.filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

        await file.seek(0)
        uploaded_bytes = await file.read()
        uploaded_sha256 = hashlib.sha256(uploaded_bytes).hexdigest()
        file_size = len(uploaded_bytes)

        ensure_storage_dirs_exist()
        canonical_path, backup_path = get_ebay_template_paths()

        assert_allowed_write_path(canonical_path)
        assert_allowed_write_path(backup_path)

        current_template = (
            self.db.query(EbayTemplate)
            .order_by(EbayTemplate.uploaded_at.desc(), EbayTemplate.id.desc())
            .first()
        )
        if current_template and current_template.sha256 == uploaded_sha256:
            disk_sha256: Optional[str] = None
            if current_template.file_path and os.path.exists(current_template.file_path):
                with open(current_template.file_path, "rb") as f:
                    disk_sha256 = hashlib.sha256(f.read()).hexdigest()

            if disk_sha256 == uploaded_sha256:
                print(
                    f"[EBAY_UPLOAD] Unchanged template detected: id={current_template.id} "
                    f"sha256={uploaded_sha256} path={current_template.file_path}"
                )
                return {
                    "template": current_template,
                    "template_unchanged": True,
                    "message": "Template unchanged — no re-parse required.",
                }

            print(
                f"[EBAY_UPLOAD] SHA matched DB but disk differs/missing; forcing rewrite: "
                f"id={current_template.id} db_sha256={current_template.sha256} "
                f"disk_sha256={disk_sha256} uploaded_sha256={uploaded_sha256} "
                f"path={current_template.file_path}"
            )

        rotate_ebay_template_backup(canonical_path, backup_path)

        try:
            with open(canonical_path, "wb") as f:
                f.write(uploaded_bytes)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to write file to disk: {e}")

        try:
            with open(canonical_path, "rb") as f:
                persisted_bytes = f.read()
            persisted_sha256 = hashlib.sha256(persisted_bytes).hexdigest()

            if uploaded_sha256 != persisted_sha256:
                if os.path.exists(canonical_path):
                    os.remove(canonical_path)
                raise HTTPException(status_code=500, detail="Persisted file integrity check failed")

        except Exception as e:
            if os.path.exists(canonical_path):
                os.remove(canonical_path)
            raise HTTPException(status_code=500, detail=f"Failed to verify persistence: {e}")

        try:
            # Replace-in-place behavior:
            # If a current eBay template row exists, update it so template_id remains stable
            # across re-uploads. This allows parse-time field preservation to compare against
            # prior settings under the same template_id.
            if current_template:
                # Keep DB metadata aligned to the exact file persisted to disk.
                current_template.original_filename = file.filename
                current_template.file_path = canonical_path
                current_template.file_size = file_size
                current_template.sha256 = uploaded_sha256
                current_template.uploaded_at = datetime.utcnow()
                template_record = current_template
                action = "updated_existing"
            else:
                template_record = EbayTemplate(
                    original_filename=file.filename,
                    file_path=canonical_path,
                    file_size=file_size,
                    sha256=uploaded_sha256,
                    uploaded_at=datetime.utcnow()
                )
                self.db.add(template_record)
                action = "created_new"

            self.db.commit()
            self.db.refresh(template_record)

            print(
                f"[EBAY_UPLOAD] Success: action={action} id={template_record.id} "
                f"sha256={uploaded_sha256} path={template_record.file_path}"
            )
            return {
                "template": template_record,
                "template_unchanged": False,
                "message": None,
            }

        except Exception as e:
            if os.path.exists(canonical_path):
                os.remove(canonical_path)
            print(f"[EBAY_UPLOAD] DB Error: {e}")
            raise HTTPException(status_code=500, detail="Database error saving template record")

    def _cell_to_string(self, value) -> Optional[str]:
        """
        Deterministic normalization of cell values to string for storage.
        """
        if value is None:
            return None

        if isinstance(value, str):
            s = value.strip()
            return s if s else None

        if isinstance(value, int):
            return str(value)

        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value)

        s = str(value).strip()
        return s if s else None

    def _normalize_field_key(self, s: str) -> str:
        """
        Normalize field name for robust matching:
        - Lowercase
        - Remove all non-alphanumeric characters
        """
        if not s:
            return ""
        return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())

    def _is_header_label(self, s: Optional[str]) -> bool:
        if not s:
            return False
        t = self._normalize_field_key(s)
        return t in ("fieldnames", "fieldname", "fields", "field")

    def _resolve_valid_value_fk_attr(self) -> str:
        if hasattr(self.ValidValueModel, "ebay_field_id"):
            return "ebay_field_id"
        if hasattr(self.ValidValueModel, "field_id"):
            return "field_id"
        raise RuntimeError("Valid value model missing ebay_field_id/field_id FK")

    def _is_non_empty_string(self, value: Optional[str]) -> bool:
        return bool(value is not None and str(value).strip())

    def scan_ebay_template(
        self,
        template_id: int,
        header_row_override: Optional[int] = None,
        first_data_row_override: Optional[int] = None,
    ) -> dict:
        """
        Scan-only operation (no DB writes) that detects header/data rows.

        Example response (all row indexes are Excel 1-based):
        {
          "template_sheet_name": "Template",
          "valid_values_sheet_name": "Valid Values",
          "default_values_sheet_name": "Default Values",
          "detected_header_row": 5,
          "detected_first_data_row": 6,
          "header_detection_scores": {
            "base_non_empty": 42,
            "match_known_fields": 39,
            "scanned_rows": 50
          },
          "reasons": ["..."],
          "header_preview": ["Custom Label (SKU)", "Title", "..."],
          "overrides": {
            "header_row_override": null,
            "first_data_row_override": null,
            "override_applied": false
          }
        }
        """
        if header_row_override is not None and header_row_override < 1:
            raise HTTPException(status_code=400, detail="header_row_override must be >= 1")
        if first_data_row_override is not None and first_data_row_override < 1:
            raise HTTPException(status_code=400, detail="first_data_row_override must be >= 1")

        template = self.db.query(EbayTemplate).filter(EbayTemplate.id == template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")
        if not os.path.exists(template.file_path):
            raise HTTPException(status_code=400, detail=f"Template file missing at {template.file_path}")

        try:
            wb = openpyxl.load_workbook(template.file_path, data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to load Excel file: {e}")

        sheet_names = wb.sheetnames
        template_sheet_name = _pick_sheet_name(sheet_names, "Template")
        valid_values_sheet_name = _pick_sheet_name(sheet_names, "Valid Values")
        default_values_sheet_name = _pick_sheet_name(sheet_names, "Default Values")

        if not template_sheet_name:
            raise HTTPException(status_code=400, detail="Could not resolve Template sheet")
        if not valid_values_sheet_name:
            raise HTTPException(status_code=400, detail="Could not resolve Valid Values sheet")
        if not default_values_sheet_name:
            raise HTTPException(status_code=400, detail="Could not resolve Default Values sheet")

        ws_template = wb[template_sheet_name]
        ws_valid = wb[valid_values_sheet_name]
        ws_defaults = wb[default_values_sheet_name]
        max_template_row = ws_template.max_row or 0

        if max_template_row < 1:
            raise HTTPException(status_code=400, detail="Template sheet is empty")

        if header_row_override is not None and header_row_override > max_template_row:
            raise HTTPException(
                status_code=400,
                detail=f"header_row_override must be <= template max_row ({max_template_row})",
            )
        if first_data_row_override is not None and first_data_row_override > max_template_row:
            raise HTTPException(
                status_code=400,
                detail=f"first_data_row_override must be <= template max_row ({max_template_row})",
            )
        if header_row_override is not None and first_data_row_override is not None:
            if first_data_row_override < header_row_override + 1:
                raise HTTPException(
                    status_code=400,
                    detail="first_data_row_override must be >= header_row_override + 1",
                )

        reasons = [
            f"Resolved template sheet: '{template_sheet_name}'",
            f"Resolved valid-values sheet: '{valid_values_sheet_name}'",
            f"Resolved default-values sheet: '{default_values_sheet_name}'",
        ]

        def _extract_known_fields_from_col_a(ws) -> set[str]:
            max_row = ws.max_row or 0
            header_row = None
            fallback_first_non_empty = None
            for r in range(1, max_row + 1):
                cell_val = self._cell_to_string(ws.cell(row=r, column=1).value)
                if cell_val and fallback_first_non_empty is None:
                    fallback_first_non_empty = r
                if cell_val and cell_val.strip().lower() == "field names":
                    header_row = r
                    break
            if header_row is None:
                header_row = fallback_first_non_empty or 1

            out: set[str] = set()
            for r in range(header_row + 1, max_row + 1):
                field_name = self._cell_to_string(ws.cell(row=r, column=1).value)
                if not field_name:
                    continue
                out.add(self._normalize_field_key(field_name))
            return out

        known_fields_norm = _extract_known_fields_from_col_a(ws_valid) | _extract_known_fields_from_col_a(ws_defaults)
        reasons.append(f"Built known-fields set from Valid Values + Default Values: {len(known_fields_norm)} entries")

        max_scan_rows = min(50, ws_template.max_row or 0)
        max_col_template = ws_template.max_column or 0
        best_row = 1
        best_base = -1
        best_match = -1

        for r in range(1, max_scan_rows + 1):
            base_non_empty = 0
            match_known_fields = 0
            for c in range(1, max_col_template + 1):
                cell_val = self._cell_to_string(ws_template.cell(row=r, column=c).value)
                if not cell_val:
                    continue
                base_non_empty += 1
                if self._normalize_field_key(cell_val) in known_fields_norm:
                    match_known_fields += 1
            if (base_non_empty, match_known_fields) > (best_base, best_match):
                best_base = base_non_empty
                best_match = match_known_fields
                best_row = r

        detected_header_row = best_row
        if header_row_override is not None:
            detected_header_row = header_row_override
            reasons.append("Applied header_row_override")
        else:
            reasons.append(
                "Detected header row by maximizing (non-empty cells, known-field matches) with earliest-row tie-break"
            )

        last_non_empty_col = 0
        for c in range(1, max_col_template + 1):
            if self._cell_to_string(ws_template.cell(row=detected_header_row, column=c).value):
                last_non_empty_col = c

        header_preview: list[str] = []
        if last_non_empty_col > 0:
            for c in range(1, last_non_empty_col + 1):
                header_preview.append(self._cell_to_string(ws_template.cell(row=detected_header_row, column=c).value) or "")
        reasons.append(f"Header preview built from row {detected_header_row} through column {last_non_empty_col or 0}")

        if first_data_row_override is not None:
            if header_row_override is None and first_data_row_override < detected_header_row + 1:
                raise HTTPException(
                    status_code=400,
                    detail="first_data_row_override must be >= detected_header_row + 1",
                )
            detected_first_data_row = first_data_row_override
            reasons.append("Applied first_data_row_override")
        else:
            default_first_data_row = detected_header_row + 1
            detected_first_data_row = default_first_data_row
            scan_end_row = min((ws_template.max_row or 0), detected_header_row + 200)
            for r in range(default_first_data_row, scan_end_row + 1):
                has_data = False
                if last_non_empty_col > 0:
                    for c in range(1, last_non_empty_col + 1):
                        if self._cell_to_string(ws_template.cell(row=r, column=c).value):
                            has_data = True
                            break
                if has_data:
                    detected_first_data_row = r
                    break
            reasons.append(
                "Detected first data row from first non-blank row after header (within next 200 rows across header columns)"
            )

        return {
            "template_sheet_name": template_sheet_name,
            "valid_values_sheet_name": valid_values_sheet_name,
            "default_values_sheet_name": default_values_sheet_name,
            "detected_header_row": detected_header_row,
            "detected_first_data_row": detected_first_data_row,
            "header_detection_scores": {
                "base_non_empty": max(0, best_base),
                "match_known_fields": max(0, best_match),
                "scanned_rows": max_scan_rows,
            },
            "reasons": reasons,
            "header_preview": header_preview,
            "overrides": {
                "header_row_override": header_row_override,
                "first_data_row_override": first_data_row_override,
                "override_applied": bool(
                    header_row_override is not None or first_data_row_override is not None
                ),
            },
        }

    def parse_ebay_template(
        self,
        template_id: int,
        header_row_override: Optional[int] = None,
        first_data_row_override: Optional[int] = None,
        reset_to_template_defaults: bool = False,
    ) -> EbayTemplateParseSummary:
        """
        Parse the stored eBay XLSX template and populate DB fields/values.
        Idempotent: Clears existing fields/values for this template before inserting.
        """
        scan_result = self.scan_ebay_template(
            template_id=template_id,
            header_row_override=header_row_override,
            first_data_row_override=first_data_row_override,
        )

        # 1) Load Template Record
        template = self.db.query(EbayTemplate).filter(EbayTemplate.id == template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")

        if not os.path.exists(template.file_path):
            raise HTTPException(status_code=400, detail=f"Template file missing at {template.file_path}")

        # 2) Load Workbook
        try:
            wb = openpyxl.load_workbook(template.file_path, data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to load Excel file: {e}")

        sheet_names = wb.sheetnames
        template_sheet_name = scan_result["template_sheet_name"]
        valid_values_sheet_name = scan_result["valid_values_sheet_name"]
        default_values_sheet_name = scan_result["default_values_sheet_name"]

        # 3) Snapshot Existing Data Before Delete
        existing_fields = (
            self.db.query(EbayField)
            .options(selectinload(EbayField.template_field))
            .filter(EbayField.ebay_template_id == template_id)
            .order_by(EbayField.order_index.asc(), EbayField.id.asc())
            .all()
        )
        existing_field_ids = [f.id for f in existing_fields]
        existing_values_by_field_id: Dict[int, List[str]] = {fid: [] for fid in existing_field_ids}

        if existing_field_ids:
            fk_attr = self._resolve_valid_value_fk_attr()
            fk_column = getattr(self.ValidValueModel, fk_attr)
            existing_values = (
                self.db.query(self.ValidValueModel)
                .filter(fk_column.in_(existing_field_ids))
                .all()
            )
            for v in existing_values:
                fid = getattr(v, fk_attr)
                existing_values_by_field_id.setdefault(fid, []).append(v.value)

        existing_field_snapshot_by_key: Dict[str, dict] = {}
        for field in existing_fields:
            key = self._normalize_field_key(field.field_name)
            if not key or key in existing_field_snapshot_by_key:
                continue
            persistent_field = field.template_field or field
            old_allowed_values = existing_values_by_field_id.get(field.id, [])
            existing_field_snapshot_by_key[key] = {
                "required": persistent_field.required,
                "row_scope": persistent_field.row_scope,
                "selected_value": persistent_field.selected_value,
                "selected_value_source": getattr(persistent_field, "selected_value_source", None),
                "custom_value": persistent_field.custom_value,
                "parent_selected_value": persistent_field.parent_selected_value,
                "parent_selected_value_source": getattr(persistent_field, "parent_selected_value_source", None),
                "parent_custom_value": persistent_field.parent_custom_value,
                "variation_selected_value": persistent_field.variation_selected_value,
                "variation_selected_value_source": getattr(persistent_field, "variation_selected_value_source", None),
                "variation_custom_value": persistent_field.variation_custom_value,
                "parsed_default_value": persistent_field.parsed_default_value,
                "old_allowed_set": set(old_allowed_values),
            }

        canonical_fields_by_key: Dict[str, TemplateField] = {}
        canonical_fields = (
            self.db.query(TemplateField)
            .filter(TemplateField.marketplace == "ebay")
            .all()
        )
        for canonical_field in canonical_fields:
            key = self._normalize_field_key(canonical_field.field_key_norm or canonical_field.field_name)
            if key and key not in canonical_fields_by_key:
                canonical_fields_by_key[key] = canonical_field

        # 4) Parse template structures first
        ws_template = wb[template_sheet_name]
        ws_valid = wb[valid_values_sheet_name]
        ws_defaults = wb[default_values_sheet_name]

        header_row_idx = int(scan_result["detected_header_row"])
        print(f"[EBAY_PARSE] template_id={template_id} selected_template_header_row={header_row_idx} source=scan")

        template_fields_by_key: Dict[str, dict] = {}
        template_fields_in_order: List[dict] = []
        max_col = ws_template.max_column or 0
        for col_idx in range(1, max_col + 1):
            raw_field_name = self._cell_to_string(ws_template.cell(row=header_row_idx, column=col_idx).value)
            if not raw_field_name:
                continue
            key = self._normalize_field_key(raw_field_name)
            if not key:
                continue
            entry = {
                "field_name": raw_field_name,
                "key": key,
                "order_index": col_idx - 1,
            }
            template_fields_in_order.append(entry)
            if key not in template_fields_by_key:
                template_fields_by_key[key] = entry

        # Parse valid values into normalized-key map (dedup while preserving order)
        values_inserted = 0
        values_ignored = 0
        new_allowed_by_key: Dict[str, List[str]] = {}
        seen_allowed_by_key: Dict[str, Set[str]] = {}
        image_pattern_candidates_by_key: Dict[str, List[str]] = {}

        max_row_valid = ws_valid.max_row or 0
        max_col_valid = ws_valid.max_column or 0
        a1_valid = self._cell_to_string(ws_valid.cell(row=1, column=1).value)
        start_row_valid = 2 if self._is_header_label(a1_valid) else 1

        for r in range(start_row_valid, max_row_valid + 1):
            raw_name = self._cell_to_string(ws_valid.cell(row=r, column=1).value)
            if not raw_name or self._is_header_label(raw_name):
                continue

            key = self._normalize_field_key(raw_name)
            if key not in template_fields_by_key:
                for c in range(2, max_col_valid + 1):
                    if self._cell_to_string(ws_valid.cell(row=r, column=c).value):
                        values_ignored += 1
                continue

            if key in IMAGE_NORM_KEYS:
                if key not in image_pattern_candidates_by_key:
                    image_pattern_candidates_by_key[key] = []
                    seen_allowed_by_key[key] = set()
                for c in range(2, max_col_valid + 1):
                    value = self._cell_to_string(ws_valid.cell(row=r, column=c).value)
                    if not value:
                        continue
                    if value in seen_allowed_by_key[key]:
                        continue
                    seen_allowed_by_key[key].add(value)
                    image_pattern_candidates_by_key[key].append(value)
                continue

            if key not in new_allowed_by_key:
                new_allowed_by_key[key] = []
                seen_allowed_by_key[key] = set()

            for c in range(2, max_col_valid + 1):
                value = self._cell_to_string(ws_valid.cell(row=r, column=c).value)
                if not value:
                    continue
                if value in seen_allowed_by_key[key]:
                    continue
                seen_allowed_by_key[key].add(value)
                new_allowed_by_key[key].append(value)
                values_inserted += 1

        # Parse template defaults into normalized-key map (column C)
        defaults_applied = 0
        defaults_ignored = 0
        template_default_by_key: Dict[str, Optional[str]] = {}
        asset_managed_by_key: Dict[str, bool] = {}

        max_row_def = ws_defaults.max_row or 0
        a1_def = self._cell_to_string(ws_defaults.cell(row=1, column=1).value)
        start_row_def = 2 if self._is_header_label(a1_def) else 1

        for r in range(start_row_def, max_row_def + 1):
            raw_name = self._cell_to_string(ws_defaults.cell(row=r, column=1).value)
            if not raw_name or self._is_header_label(raw_name):
                continue

            key = self._normalize_field_key(raw_name)
            default_val = self._cell_to_string(ws_defaults.cell(row=r, column=3).value)
            if default_val is not None and str(default_val).strip().lower() == "asset":
                template_default_by_key[key] = None
                asset_managed_by_key[key] = True
            else:
                template_default_by_key[key] = default_val
                asset_managed_by_key[key] = False

            if key in template_fields_by_key and self._is_non_empty_string(default_val):
                defaults_applied += 1
            elif key not in template_fields_by_key and self._is_non_empty_string(default_val):
                defaults_ignored += 1

        for key, is_asset_managed in asset_managed_by_key.items():
            if is_asset_managed:
                new_allowed_by_key[key] = []

        # 5) Clear Existing Data (delete values first, then fields)
        if existing_field_ids:
            fk_attr = self._resolve_valid_value_fk_attr()
            fk_column = getattr(self.ValidValueModel, fk_attr)
            self.db.query(self.ValidValueModel).filter(
                fk_column.in_(existing_field_ids)
            ).delete(synchronize_session=False)

        self.db.query(EbayField).filter(
            EbayField.ebay_template_id == template_id
        ).delete(synchronize_session=False)
        self.db.flush()

        # 6) Rebuild fields and merge prior settings deterministically
        fields_inserted = 0
        field_map_by_key: Dict[str, EbayField] = {}

        for item in template_fields_in_order:
            key = item["key"]
            is_asset_managed = bool(asset_managed_by_key.get(key, False))
            old_snapshot = existing_field_snapshot_by_key.get(key)
            new_template_default = template_default_by_key.get(key)
            new_allowed_list = new_allowed_by_key.get(key, [])
            new_allowed_set = set(new_allowed_list)
            valid_set = {
                str(v).strip()
                for v in new_allowed_list
                if v is not None and str(v).strip()
            }

            compatible = False
            default_unchanged = False
            if old_snapshot is not None:
                compatible = old_snapshot["old_allowed_set"].issubset(new_allowed_set)
                default_unchanged = old_snapshot.get("parsed_default_value") == new_template_default

            preserve_values = (
                old_snapshot is not None
                and (not reset_to_template_defaults)
                and compatible
                and default_unchanged
            )

            if old_snapshot is not None and not reset_to_template_defaults:
                required_value = old_snapshot.get("required", False)
                row_scope_value = old_snapshot.get("row_scope")
            else:
                required_value = False
                row_scope_value = None

            if preserve_values:
                selected_value = old_snapshot.get("selected_value")
                selected_value_source = old_snapshot.get("selected_value_source")
                custom_value = old_snapshot.get("custom_value")
                parent_selected_value = old_snapshot.get("parent_selected_value")
                parent_selected_value_source = old_snapshot.get("parent_selected_value_source")
                parent_custom_value = old_snapshot.get("parent_custom_value")
                variation_selected_value = old_snapshot.get("variation_selected_value")
                variation_selected_value_source = old_snapshot.get("variation_selected_value_source")
                variation_custom_value = old_snapshot.get("variation_custom_value")
            else:
                selected_value = new_template_default
                selected_value_source = "parsed" if self._is_non_empty_string(new_template_default) else None
                custom_value = None
                parent_selected_value = None
                parent_selected_value_source = None
                parent_custom_value = None
                variation_selected_value = None
                variation_selected_value_source = None
                variation_custom_value = None

            if is_asset_managed:
                new_template_default = None
                # Keep only manual legacy selections for optional fallback UI.
                if selected_value_source != "manual":
                    selected_value = None
                    selected_value_source = None
                if parent_selected_value_source != "manual":
                    parent_selected_value = None
                    parent_selected_value_source = None
                if variation_selected_value_source != "manual":
                    variation_selected_value = None
                    variation_selected_value_source = None

            def _reconcile_selection(selection_label: str, selection_value: Optional[str]) -> Optional[str]:
                if selection_value is None:
                    return None
                trimmed = str(selection_value).strip()
                if not trimmed:
                    return None
                if trimmed.lower() == "any":
                    print(
                        f"[EBAY_PARSE] Clearing stale {item['field_name']} {selection_label} selection: "
                        f"{selection_value!r} not in valid values"
                    )
                    return None
                if trimmed not in valid_set:
                    print(
                        f"[EBAY_PARSE] Clearing stale {item['field_name']} {selection_label} selection: "
                        f"{selection_value!r} not in valid values"
                    )
                    return None
                return selection_value

            if not is_asset_managed:
                selected_value = _reconcile_selection("selected", selected_value)
                parent_selected_value = _reconcile_selection("parent", parent_selected_value)
                variation_selected_value = _reconcile_selection("variation", variation_selected_value)
                if selected_value is None:
                    selected_value_source = None
                if parent_selected_value is None:
                    parent_selected_value_source = None
                if variation_selected_value is None:
                    variation_selected_value_source = None

            canonical_field = canonical_fields_by_key.get(key)
            if canonical_field is None:
                canonical_field = TemplateField(
                    marketplace="ebay",
                    field_name=item["field_name"],
                    field_key_norm=key,
                )
                self.db.add(canonical_field)
                canonical_fields_by_key[key] = canonical_field

            canonical_field.field_name = item["field_name"]
            canonical_field.field_key_norm = key
            canonical_field.order_index = item["order_index"]
            canonical_field.required = required_value
            canonical_field.is_asset_managed = is_asset_managed
            canonical_field.row_scope = row_scope_value
            canonical_field.selected_value = selected_value
            canonical_field.selected_value_source = selected_value_source
            canonical_field.custom_value = custom_value
            canonical_field.parent_selected_value = parent_selected_value
            canonical_field.parent_selected_value_source = parent_selected_value_source
            canonical_field.parent_custom_value = parent_custom_value
            canonical_field.variation_selected_value = variation_selected_value
            canonical_field.variation_selected_value_source = variation_selected_value_source
            canonical_field.variation_custom_value = variation_custom_value
            canonical_field.parsed_default_value = None if is_asset_managed else new_template_default

            # Regression guard: supported keys must always point to a canonical TemplateField.
            if key in SUPPORTED_TEMPLATE_FIELD_KEYS and canonical_field.id is None:
                self.db.flush()

            field = EbayField(
                ebay_template_id=template_id,
                template_field=canonical_field,
                field_name=item["field_name"],
                display_name=item["field_name"],
                required=False,
                order_index=item["order_index"],
                selected_value=None,
                custom_value=None,
                parsed_default_value=None if is_asset_managed else new_template_default,
                parent_selected_value=None,
                parent_custom_value=None,
                variation_selected_value=None,
                variation_custom_value=None,
                row_scope=None,
            )
            self.db.add(field)
            field_map_by_key[key] = field
            fields_inserted += 1

        self.db.flush()

        # 7) For asset-managed image URL fields, hydrate canonical template-owned assets
        # from the template Valid Values row (parent/variation pattern candidates).
        image_asset_types = {"image_parent_pattern", "image_variation_pattern"}
        for key in sorted(k for k in template_fields_by_key.keys() if k in IMAGE_NORM_KEYS):
            if not asset_managed_by_key.get(key, False):
                continue
            canonical_field = canonical_fields_by_key.get(key)
            if canonical_field is None or canonical_field.id is None:
                continue

            candidates = image_pattern_candidates_by_key.get(key, [])
            if not candidates:
                print(
                    f"[EBAY_PARSE] image_patterns field_key={key} template_field_id={canonical_field.id} "
                    "candidates=0 preserve_existing_assets=true"
                )
                continue

            existing_asset_ids = [
                row[0]
                for row in self.db.query(TemplateFieldAsset.id)
                .filter(
                    TemplateFieldAsset.template_field_id == canonical_field.id,
                    TemplateFieldAsset.asset_type.in_(tuple(image_asset_types)),
                    TemplateFieldAsset.source == "template",
                )
                .all()
            ]
            existing_fallback_rows = (
                self.db.query(TemplateFieldAsset)
                .filter(
                    TemplateFieldAsset.template_field_id == canonical_field.id,
                    TemplateFieldAsset.asset_type.in_(tuple(image_asset_types)),
                    TemplateFieldAsset.is_default_fallback.is_(True),
                )
                .all()
            )
            locked_fallback_asset_types = {
                row.asset_type
                for row in existing_fallback_rows
                if (row.source or "user") != "template"
            }
            if existing_asset_ids:
                self.db.query(TemplateFieldAssetEquipmentType).filter(
                    TemplateFieldAssetEquipmentType.asset_id.in_(existing_asset_ids)
                ).delete(synchronize_session=False)
                self.db.query(TemplateFieldAsset).filter(
                    TemplateFieldAsset.id.in_(existing_asset_ids)
                ).delete(synchronize_session=False)
                # Ensure deletes are visible before any replacement inserts for unique fallback keys.
                self.db.flush()

            parent_pattern = candidates[0] if len(candidates) >= 1 else None
            variation_pattern = candidates[1] if len(candidates) >= 2 else (candidates[0] if candidates else None)

            if parent_pattern and "image_parent_pattern" not in locked_fallback_asset_types:
                self.db.add(
                    TemplateFieldAsset(
                        template_field_id=canonical_field.id,
                        asset_type="image_parent_pattern",
                        value=parent_pattern,
                        source="template",
                        is_default_fallback=True,
                    )
                )
            if variation_pattern and "image_variation_pattern" not in locked_fallback_asset_types:
                self.db.add(
                    TemplateFieldAsset(
                        template_field_id=canonical_field.id,
                        asset_type="image_variation_pattern",
                        value=variation_pattern,
                        source="template",
                        is_default_fallback=True,
                    )
                )
            print(
                f"[EBAY_PARSE] image_patterns field_key={key} template_field_id={canonical_field.id} "
                f"candidates={len(candidates)} template_assets_replaced=true"
            )

        # 8) Insert new allowed values exactly as parsed (expanded/replaced by template)
        fk_attr = self._resolve_valid_value_fk_attr()
        for key, field in field_map_by_key.items():
            if asset_managed_by_key.get(key, False):
                continue
            for value in new_allowed_by_key.get(key, []):
                if fk_attr == "ebay_field_id":
                    self.db.add(self.ValidValueModel(ebay_field_id=field.id, value=value))
                else:
                    self.db.add(self.ValidValueModel(field_id=field.id, value=value))

        # Commit
        self.db.commit()

        supported_null_count = (
            self.db.query(EbayField)
            .filter(EbayField.ebay_template_id == template_id)
            .all()
        )
        supported_null_count = sum(
            1
            for row in supported_null_count
            if self._normalize_field_key(row.field_name) in SUPPORTED_TEMPLATE_FIELD_KEYS
            and row.template_field_id is None
        )
        print(
            f"[EBAY_PARSE] template_id={template_id} supported_keys_null_template_field_id={supported_null_count}"
        )

        return EbayTemplateParseSummary(
            template_id=template_id,
            fields_inserted=fields_inserted,
            values_inserted=values_inserted,
            defaults_applied=defaults_applied,
            values_ignored_not_in_template=values_ignored,
            defaults_ignored_not_in_template=defaults_ignored,
            sheet_names=sheet_names
        )
