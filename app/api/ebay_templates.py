from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Body
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session, selectinload
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from typing import Optional, List
from pydantic import BaseModel
import os
import hashlib
from datetime import datetime
from openpyxl import load_workbook

from app.database import get_db
from app.services.ebay_template_service import EbayTemplateService
from app.schemas.templates import (
    EbayTemplateResponse,
    EbayTemplateParseSummary,
    EbayTemplateFieldsResponse,
    EbayFieldResponse,
    EbayFieldUpdateRequest,
    EbayValidValueCreateRequest,
    EbayTemplatePreviewResponse,
    EbayTemplateIntegrityResponse,
    EbayTemplateVerificationResponse,
    EbayFieldEquipmentTypeContentResponse,
    EbayFieldEquipmentTypeContentUpsertRequest,
    EbayFieldEquipmentTypeImagePatternResponse,
    EbayFieldEquipmentTypeImagePatternUpsertRequest,
    TemplateFieldAssetResponse,
    TemplateFieldAssetCreateRequest,
    TemplateFieldAssetUpdateRequest,
)
from app.models.templates import (
    EbayTemplate,
    EbayField,
    EbayFieldValue,
    EbayFieldEquipmentTypeContent,
    EbayFieldEquipmentTypeImagePattern,
    TemplateFieldAsset,
    TemplateFieldAssetEquipmentType,
)
from app.models.core import EquipmentType

router = APIRouter(
    prefix="/ebay-templates",
    tags=["Ebay Templates"]
)

EBAY_FIELD_ROW_SCOPE_VALUES = {"both", "parent_only", "variation_only"}
TEMPLATE_FIELD_ASSET_TYPES = {"description_html", "image_parent_pattern", "image_variation_pattern"}


class EbayTemplateScanRequest(BaseModel):
    header_row_override: Optional[int] = None
    first_data_row_override: Optional[int] = None


class EbayTemplateParseRequest(BaseModel):
    header_row_override: Optional[int] = None
    first_data_row_override: Optional[int] = None
    reset_to_template_defaults: bool = False


class EbayTemplateHeaderDetectionScores(BaseModel):
    base_non_empty: int
    match_known_fields: int
    scanned_rows: int


class EbayTemplateScanOverrides(BaseModel):
    header_row_override: Optional[int] = None
    first_data_row_override: Optional[int] = None
    override_applied: bool


class EbayTemplateScanResponse(BaseModel):
    template_sheet_name: str
    valid_values_sheet_name: str
    default_values_sheet_name: str
    detected_header_row: int
    detected_first_data_row: int
    header_detection_scores: EbayTemplateHeaderDetectionScores
    reasons: List[str]
    header_preview: List[str]
    overrides: EbayTemplateScanOverrides


def _coalesce_ebay_field_row_scope(value: Optional[str]) -> str:
    return value if value in EBAY_FIELD_ROW_SCOPE_VALUES else "both"


def _resolve_template_field_from_ebay_field(field_id: int, db: Session):
    field = (
        db.query(EbayField)
        .options(selectinload(EbayField.template_field))
        .filter(EbayField.id == field_id)
        .first()
    )
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    template_field = field.template_field
    if not template_field:
        raise HTTPException(status_code=400, detail="Field is not linked to canonical template_field")
    return field, template_field


def _build_template_field_asset_response(row: TemplateFieldAsset) -> TemplateFieldAssetResponse:
    equipment_type_ids = sorted(
        [
            int(link.equipment_type_id)
            for link in (row.equipment_type_links or [])
            if link.equipment_type_id is not None
        ]
    )
    return TemplateFieldAssetResponse(
        id=row.id,
        template_field_id=row.template_field_id,
        asset_type=row.asset_type,
        name=(str(getattr(row, "name", "") or "").strip() or f"Untitled (id: {int(row.id)})"),
        value=row.value,
        is_default_fallback=bool(row.is_default_fallback),
        equipment_type_ids=equipment_type_ids,
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


def _validate_asset_type(asset_type: str) -> str:
    normalized = str(asset_type or "").strip().lower()
    if normalized not in TEMPLATE_FIELD_ASSET_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid asset_type. Must be one of {sorted(TEMPLATE_FIELD_ASSET_TYPES)}",
        )
    return normalized


def _is_blank(value: Optional[str]) -> bool:
    return str(value or "").strip() == ""


def _normalize_asset_name(value: Optional[str]) -> Optional[str]:
    text = str(value or "").strip()
    return text if text else None


def _validate_equipment_type_ids(equipment_type_ids: List[int], db: Session) -> List[int]:
    normalized_ids = sorted({int(v) for v in (equipment_type_ids or [])})
    if not normalized_ids:
        return []
    existing = (
        db.query(EquipmentType.id)
        .filter(EquipmentType.id.in_(normalized_ids))
        .all()
    )
    existing_ids = {int(row[0]) for row in existing}
    missing = [eid for eid in normalized_ids if eid not in existing_ids]
    if missing:
        raise HTTPException(status_code=400, detail=f"Equipment type(s) not found: {missing}")
    return normalized_ids


def _validate_no_equipment_type_overlap(
    template_field_id: int,
    asset_type: str,
    equipment_type_ids: List[int],
    db: Session,
    exclude_asset_id: Optional[int] = None,
) -> None:
    if not equipment_type_ids:
        return
    q = (
        db.query(TemplateFieldAssetEquipmentType.equipment_type_id)
        .join(TemplateFieldAsset, TemplateFieldAsset.id == TemplateFieldAssetEquipmentType.asset_id)
        .filter(
            TemplateFieldAsset.template_field_id == template_field_id,
            TemplateFieldAsset.asset_type == asset_type,
            TemplateFieldAssetEquipmentType.equipment_type_id.in_(equipment_type_ids),
        )
    )
    if exclude_asset_id is not None:
        q = q.filter(TemplateFieldAsset.id != exclude_asset_id)
    overlaps = sorted({int(row[0]) for row in q.all()})
    if overlaps:
        raise HTTPException(
            status_code=400,
            detail=(
                "One or more equipment types already assigned to another asset for this field/type: "
                f"{overlaps}"
            ),
        )


def _build_ebay_template_fields_response(template_id: int, db: Session) -> EbayTemplateFieldsResponse:
    """
    Internal helper: Load fields + valid values for a template and map to API response models.
    Uses selectinload to reliably populate one-to-many relationships without join edge-cases.
    """
    # 1) Verify template exists
    template = db.query(EbayTemplate).filter(EbayTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    # 2) Query fields and eagerly load valid_values via SELECT IN (more reliable than joinedload)
    # Order by order_index ASC (nulls last), then id ASC
    fields: List[EbayField] = (
        db.query(EbayField)
        .options(selectinload(EbayField.valid_values), selectinload(EbayField.template_field))
        .filter(EbayField.ebay_template_id == template_id)
        .order_by(func.coalesce(EbayField.order_index, 10**9), EbayField.id)
        .all()
    )

    # 3) Map to response
    response_fields: List[EbayFieldResponse] = []
    for f in fields:
        persistent = f.template_field or f
        # Sort values deterministically by ID ASC
        sorted_values = sorted((f.valid_values or []), key=lambda v: v.id)

        # Map to List[str] as required by schema field 'allowed_values'
        allowed_strs = [v.value for v in sorted_values]
        
        # Map to detailed list with IDs for delete operations
        allowed_detailed = [{"id": v.id, "value": v.value} for v in sorted_values]

        response_fields.append(
            EbayFieldResponse(
                id=f.id,
                ebay_template_id=f.ebay_template_id,
                field_name=f.field_name,
                display_name=f.display_name,
                required=bool(persistent.required),
                is_asset_managed=bool(getattr(persistent, "is_asset_managed", False)),
                order_index=f.order_index,
                selected_value=persistent.selected_value,
                custom_value=persistent.custom_value,
                parsed_default_value=persistent.parsed_default_value,
                parent_selected_value=persistent.parent_selected_value,
                parent_custom_value=persistent.parent_custom_value,
                variation_selected_value=persistent.variation_selected_value,
                variation_custom_value=persistent.variation_custom_value,
                row_scope=_coalesce_ebay_field_row_scope(persistent.row_scope),
                allowed_values=allowed_strs,
                allowed_values_detailed=allowed_detailed
            )
        )

    return EbayTemplateFieldsResponse(
        template_id=template_id,
        fields=response_fields
    )


@router.post("/upload", response_model=EbayTemplateResponse)
async def upload_ebay_template(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Upload and store the canonical eBay XLSX template (bit-for-bit).
    """
    service = EbayTemplateService(db)
    result = await service.store_ebay_template_upload(file)
    template: EbayTemplate = result["template"]
    template_unchanged = bool(result.get("template_unchanged", False))
    message = result.get("message")

    # Auto-parse on new/changed uploads so fields/defaults are available immediately.
    # Keep unchanged upload behavior as-is (no re-parse).
    if not template_unchanged:
        service.parse_ebay_template(template.id)
        message = "Template uploaded and parsed successfully."

    return EbayTemplateResponse(
        id=template.id,
        original_filename=template.original_filename,
        file_size=template.file_size,
        sha256=template.sha256,
        uploaded_at=template.uploaded_at,
        template_unchanged=template_unchanged,
        message=message,
    )


@router.get("/current", response_model=Optional[EbayTemplateResponse])
def get_current_ebay_template(db: Session = Depends(get_db)):
    """
    Get the most recently uploaded eBay template metadata.
    """
    latest = (
        db.query(EbayTemplate)
        .order_by(EbayTemplate.uploaded_at.desc(), EbayTemplate.id.desc())
        .first()
    )

    if not latest:
        return None

    return latest


@router.post("/{template_id}/parse", response_model=EbayTemplateParseSummary)
def parse_ebay_template(
    template_id: int,
    request: EbayTemplateParseRequest = Body(default=EbayTemplateParseRequest()),
    db: Session = Depends(get_db)
):
    """
    Parse the eBay template file and populate metadata in the database.
    Idempotent operation (clears existing fields/values for this template).
    Optional overrides are accepted using Excel 1-based row indexes.
    If omitted, parse uses scan detection.
    """
    service = EbayTemplateService(db)
    return service.parse_ebay_template(
        template_id,
        header_row_override=request.header_row_override,
        first_data_row_override=request.first_data_row_override,
        reset_to_template_defaults=request.reset_to_template_defaults,
    )


@router.post("/{template_id}/scan", response_model=EbayTemplateScanResponse)
def scan_ebay_template(
    template_id: int,
    request: EbayTemplateScanRequest = Body(default=EbayTemplateScanRequest()),
    db: Session = Depends(get_db)
):
    """
    Scan-only detection (no DB writes) for eBay templates.
    All row indexes are Excel 1-based in request and response.
    """
    service = EbayTemplateService(db)
    return service.scan_ebay_template(
        template_id=template_id,
        header_row_override=request.header_row_override,
        first_data_row_override=request.first_data_row_override,
    )


@router.get("/current/fields", response_model=EbayTemplateFieldsResponse)
def get_current_ebay_template_fields(db: Session = Depends(get_db)):
    """
    Get the parsed fields and allowed values for the MOST RECENT template.
    """
    latest = (
        db.query(EbayTemplate)
        .order_by(EbayTemplate.uploaded_at.desc(), EbayTemplate.id.desc())
        .first()
    )

    if not latest:
        raise HTTPException(status_code=404, detail="No eBay template uploaded")

    return _build_ebay_template_fields_response(latest.id, db)


@router.get("/{template_id}/fields", response_model=EbayTemplateFieldsResponse)
def get_ebay_template_fields(template_id: int, db: Session = Depends(get_db)):
    """
    Get the parsed fields and allowed values for a specific template.
    Returns fields ordered by order_index.
    """
    return _build_ebay_template_fields_response(template_id, db)


@router.patch("/fields/{field_id}", response_model=EbayFieldResponse)
def update_ebay_field(
    field_id: int,
    updates: EbayFieldUpdateRequest,
    db: Session = Depends(get_db)
):
    """
    Update eBay field properties (required, selected_value, custom_value).
    
    Validation rules:
    - selected_value "Any" → stored as None
    - selected_value must exist in valid_values OR custom_value must be set
    - empty custom_value → stored as None
    """
    # Load field with valid values for validation
    field = (
        db.query(EbayField)
        .options(selectinload(EbayField.valid_values), selectinload(EbayField.template_field))
        .filter(EbayField.id == field_id)
        .first()
    )
    
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    
    persistent = field.template_field or field

    def _normalize_any_to_none(value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        if text.lower() == "any":
            return None
        return value

    # Update required if provided
    if updates.required is not None:
        persistent.required = updates.required
    
    # Update selected_value only when explicitly provided
    if "selected_value" in updates.model_fields_set:
        normalized_selected = _normalize_any_to_none(updates.selected_value)
        if normalized_selected is None:
            persistent.selected_value = None
            if hasattr(persistent, "selected_value_source"):
                persistent.selected_value_source = None
        else:
            # Get list of valid values
            valid_values_list = [v.value for v in field.valid_values]
            
            # Allow if it's in valid values OR if custom_value is being set
            if normalized_selected in valid_values_list or updates.custom_value is not None:
                persistent.selected_value = normalized_selected
                if hasattr(persistent, "selected_value_source"):
                    persistent.selected_value_source = "manual"
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid selected_value. Must be one of {valid_values_list} or set custom_value"
                )
    
    # Update custom_value if provided
    if updates.custom_value is not None:
        persistent.custom_value = updates.custom_value if updates.custom_value else None

    if "parent_selected_value" in updates.model_fields_set:
        normalized_parent_selected = _normalize_any_to_none(updates.parent_selected_value)
        persistent.parent_selected_value = normalized_parent_selected
        if hasattr(persistent, "parent_selected_value_source"):
            persistent.parent_selected_value_source = "manual" if normalized_parent_selected is not None else None
    if "parent_custom_value" in updates.model_fields_set:
        persistent.parent_custom_value = updates.parent_custom_value
    if "variation_selected_value" in updates.model_fields_set:
        normalized_variation_selected = _normalize_any_to_none(updates.variation_selected_value)
        persistent.variation_selected_value = normalized_variation_selected
        if hasattr(persistent, "variation_selected_value_source"):
            persistent.variation_selected_value_source = "manual" if normalized_variation_selected is not None else None
    if "variation_custom_value" in updates.model_fields_set:
        persistent.variation_custom_value = updates.variation_custom_value

    # Update row_scope only when explicitly provided (including null to clear to DB NULL)
    if "row_scope" in updates.model_fields_set:
        if updates.row_scope is None:
            persistent.row_scope = None
        elif updates.row_scope in EBAY_FIELD_ROW_SCOPE_VALUES:
            persistent.row_scope = updates.row_scope
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid row_scope. Must be one of {sorted(EBAY_FIELD_ROW_SCOPE_VALUES)}"
            )
    
    # Commit changes
    db.commit()
    db.refresh(field)
    if field.template_field:
        db.refresh(field.template_field)
    persistent = field.template_field or field
    
    # Return updated field using same response structure
    sorted_values = sorted((field.valid_values or []), key=lambda v: v.id)
    allowed_strs = [v.value for v in sorted_values]
    allowed_detailed = [{"id": v.id, "value": v.value} for v in sorted_values]
    
    return EbayFieldResponse(
        id=field.id,
        ebay_template_id=field.ebay_template_id,
        field_name=field.field_name,
        display_name=field.display_name,
        required=bool(persistent.required),
        is_asset_managed=bool(getattr(persistent, "is_asset_managed", False)),
        order_index=field.order_index,
        selected_value=persistent.selected_value,
        custom_value=persistent.custom_value,
        parsed_default_value=persistent.parsed_default_value,
        parent_selected_value=persistent.parent_selected_value,
        parent_custom_value=persistent.parent_custom_value,
        variation_selected_value=persistent.variation_selected_value,
        variation_custom_value=persistent.variation_custom_value,
        row_scope=_coalesce_ebay_field_row_scope(persistent.row_scope),
        allowed_values=allowed_strs,
        allowed_values_detailed=allowed_detailed
    )


@router.post("/fields/{field_id}/valid-values", response_model=EbayFieldResponse)
def add_valid_value_to_field(
    field_id: int,
    request: EbayValidValueCreateRequest,
    db: Session = Depends(get_db)
):
    """
    Add a valid value to an eBay field.
    
    Rules:
    - Trims whitespace
    - Rejects empty values
    - De-duplicates (case-sensitive)
    - Returns updated field with all valid values
    """
    # Trim and validate
    value = request.value.strip()
    if not value:
        raise HTTPException(status_code=400, detail="Value cannot be empty")
    
    # Load field with valid values
    field = (
        db.query(EbayField)
        .options(selectinload(EbayField.valid_values), selectinload(EbayField.template_field))
        .filter(EbayField.id == field_id)
        .first()
    )
    
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    
    # Check if value already exists (case-sensitive)
    existing = (
        db.query(EbayFieldValue)
        .filter(
            EbayFieldValue.ebay_field_id == field_id,
            EbayFieldValue.value == value
        )
        .first()
    )
    
    if not existing:
        # Add new value
        new_value = EbayFieldValue(ebay_field_id=field_id, value=value)
        db.add(new_value)
        db.commit()
        db.refresh(field)
        if field.template_field:
            db.refresh(field.template_field)
    persistent = field.template_field or field
    
    # Return updated field
    sorted_values = sorted((field.valid_values or []), key=lambda v: v.id)
    allowed_strs = [v.value for v in sorted_values]
    allowed_detailed = [{"id": v.id, "value": v.value} for v in sorted_values]
    
    return EbayFieldResponse(
        id=field.id,
        ebay_template_id=field.ebay_template_id,
        field_name=field.field_name,
        display_name=field.display_name,
        required=bool(persistent.required),
        is_asset_managed=bool(getattr(persistent, "is_asset_managed", False)),
        order_index=field.order_index,
        selected_value=persistent.selected_value,
        custom_value=persistent.custom_value,
        parsed_default_value=persistent.parsed_default_value,
        parent_selected_value=persistent.parent_selected_value,
        parent_custom_value=persistent.parent_custom_value,
        variation_selected_value=persistent.variation_selected_value,
        variation_custom_value=persistent.variation_custom_value,
        row_scope=_coalesce_ebay_field_row_scope(persistent.row_scope),
        allowed_values=allowed_strs,
        allowed_values_detailed=allowed_detailed
    )


@router.delete("/fields/{field_id}/valid-values/{value_id}", response_model=EbayFieldResponse)
def delete_valid_value_from_field(
    field_id: int,
    value_id: int,
    db: Session = Depends(get_db)
):
    """
    Delete a valid value from an eBay field.
    
    Rules:
    - Ensures value exists and belongs to the field
    - If deleted value equals field.selected_value, clears selected_value
    - Returns updated field
    """
    # Load the value and verify it belongs to this field
    value_obj = (
        db.query(EbayFieldValue)
        .filter(
            EbayFieldValue.id == value_id,
            EbayFieldValue.ebay_field_id == field_id
        )
        .first()
    )
    
    if not value_obj:
        raise HTTPException(
            status_code=404,
            detail="Valid value not found or does not belong to this field"
        )
    
    # Load field
    field = (
        db.query(EbayField)
        .options(selectinload(EbayField.valid_values), selectinload(EbayField.template_field))
        .filter(EbayField.id == field_id)
        .first()
    )
    
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    
    persistent = field.template_field or field

    # If this value is currently selected, clear selected_value
    if persistent.selected_value == value_obj.value:
        persistent.selected_value = None
    
    # Delete the value
    db.delete(value_obj)
    db.commit()
    db.refresh(field)
    if field.template_field:
        db.refresh(field.template_field)
    persistent = field.template_field or field
    
    # Return updated field
    sorted_values = sorted((field.valid_values or []), key=lambda v: v.id)
    allowed_strs = [v.value for v in sorted_values]
    allowed_detailed = [{"id": v.id, "value": v.value} for v in sorted_values]
    
    return EbayFieldResponse(
        id=field.id,
        ebay_template_id=field.ebay_template_id,
        field_name=field.field_name,
        display_name=field.display_name,
        required=bool(persistent.required),
        is_asset_managed=bool(getattr(persistent, "is_asset_managed", False)),
        order_index=field.order_index,
        selected_value=persistent.selected_value,
        custom_value=persistent.custom_value,
        parsed_default_value=persistent.parsed_default_value,
        parent_selected_value=persistent.parent_selected_value,
        parent_custom_value=persistent.parent_custom_value,
        variation_selected_value=persistent.variation_selected_value,
        variation_custom_value=persistent.variation_custom_value,
        row_scope=_coalesce_ebay_field_row_scope(persistent.row_scope),
        allowed_values=allowed_strs,
        allowed_values_detailed=allowed_detailed
    )


@router.get("/fields/{field_id}/assets", response_model=List[TemplateFieldAssetResponse])
def list_template_field_assets(
    field_id: int,
    asset_type: Optional[str] = None,
    db: Session = Depends(get_db),
):
    _, template_field = _resolve_template_field_from_ebay_field(field_id, db)
    q = (
        db.query(TemplateFieldAsset)
        .options(selectinload(TemplateFieldAsset.equipment_type_links))
        .filter(TemplateFieldAsset.template_field_id == template_field.id)
        .order_by(
            TemplateFieldAsset.asset_type.asc(),
            TemplateFieldAsset.is_default_fallback.desc(),
            TemplateFieldAsset.id.asc(),
        )
    )
    if asset_type:
        q = q.filter(TemplateFieldAsset.asset_type == _validate_asset_type(asset_type))
    rows = q.all()
    return [_build_template_field_asset_response(row) for row in rows]


@router.post("/fields/{field_id}/assets", response_model=TemplateFieldAssetResponse)
def create_template_field_asset(
    field_id: int,
    request: TemplateFieldAssetCreateRequest,
    db: Session = Depends(get_db),
):
    _, template_field = _resolve_template_field_from_ebay_field(field_id, db)
    normalized_asset_type = _validate_asset_type(request.asset_type)
    equipment_type_ids = _validate_equipment_type_ids(request.equipment_type_ids, db)
    is_fallback = bool(request.is_default_fallback)
    value = str(request.value or "")
    if normalized_asset_type == "description_html" and is_fallback and _is_blank(value):
        raise HTTPException(status_code=400, detail="Fallback Description HTML cannot be blank")
    if _is_blank(value):
        raise HTTPException(status_code=400, detail="Asset value cannot be blank")
    if is_fallback:
        equipment_type_ids = []

    _validate_no_equipment_type_overlap(
        template_field_id=template_field.id,
        asset_type=normalized_asset_type,
        equipment_type_ids=equipment_type_ids,
        db=db,
    )

    if is_fallback:
        db.query(TemplateFieldAsset).filter(
            TemplateFieldAsset.template_field_id == template_field.id,
            TemplateFieldAsset.asset_type == normalized_asset_type,
        ).update({"is_default_fallback": False}, synchronize_session=False)
        try:
            db.flush()
        except IntegrityError:
            db.rollback()
            raise HTTPException(
                status_code=400,
                detail=(
                    "Only one fallback asset is allowed per field and asset type. "
                    "Unset fallback on the current fallback asset first."
                ),
            )

    row = TemplateFieldAsset(
        template_field_id=template_field.id,
        asset_type=normalized_asset_type,
        name=_normalize_asset_name(request.name),
        value=value,
        source="user",
        is_default_fallback=is_fallback,
    )
    db.add(row)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=(
                "Only one fallback asset is allowed per field and asset type. "
                "Unset fallback on the current fallback asset first."
            ),
        )

    for equipment_type_id in equipment_type_ids:
        db.add(
            TemplateFieldAssetEquipmentType(
                asset_id=row.id,
                equipment_type_id=equipment_type_id,
            )
        )

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=(
                "Only one fallback asset is allowed per field and asset type. "
                "Unset fallback on the current fallback asset first."
            ),
        )
    row = (
        db.query(TemplateFieldAsset)
        .options(selectinload(TemplateFieldAsset.equipment_type_links))
        .filter(TemplateFieldAsset.id == row.id)
        .first()
    )
    return _build_template_field_asset_response(row)


@router.put("/fields/{field_id}/assets/{asset_id}", response_model=TemplateFieldAssetResponse)
def update_template_field_asset(
    field_id: int,
    asset_id: int,
    request: TemplateFieldAssetUpdateRequest,
    db: Session = Depends(get_db),
):
    _, template_field = _resolve_template_field_from_ebay_field(field_id, db)
    row = (
        db.query(TemplateFieldAsset)
        .options(selectinload(TemplateFieldAsset.equipment_type_links))
        .filter(
            TemplateFieldAsset.id == asset_id,
            TemplateFieldAsset.template_field_id == template_field.id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Asset not found")

    value = str(request.value or "")
    if _is_blank(value):
        raise HTTPException(status_code=400, detail="Asset value cannot be blank")
    equipment_type_ids = _validate_equipment_type_ids(request.equipment_type_ids, db)
    is_fallback = bool(request.is_default_fallback)
    if is_fallback:
        equipment_type_ids = []
        if row.asset_type == "description_html" and _is_blank(value):
            raise HTTPException(status_code=400, detail="Fallback Description HTML cannot be blank")

    _validate_no_equipment_type_overlap(
        template_field_id=template_field.id,
        asset_type=row.asset_type,
        equipment_type_ids=equipment_type_ids,
        db=db,
        exclude_asset_id=row.id,
    )

    if is_fallback:
        db.query(TemplateFieldAsset).filter(
            TemplateFieldAsset.template_field_id == template_field.id,
            TemplateFieldAsset.asset_type == row.asset_type,
            TemplateFieldAsset.id != row.id,
        ).update({"is_default_fallback": False}, synchronize_session=False)
        try:
            db.flush()
        except IntegrityError:
            db.rollback()
            raise HTTPException(
                status_code=400,
                detail=(
                    "Only one fallback asset is allowed per field and asset type. "
                    "Unset fallback on the current fallback asset first."
                ),
            )

    row.value = value
    if "name" in request.model_fields_set:
        row.name = _normalize_asset_name(request.name)
    row.is_default_fallback = is_fallback
    if hasattr(row, "source"):
        row.source = "user"
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=(
                "Only one fallback asset is allowed per field and asset type. "
                "Unset fallback on the current fallback asset first."
            ),
        )

    db.query(TemplateFieldAssetEquipmentType).filter(
        TemplateFieldAssetEquipmentType.asset_id == row.id
    ).delete(synchronize_session=False)
    for equipment_type_id in equipment_type_ids:
        db.add(
            TemplateFieldAssetEquipmentType(
                asset_id=row.id,
                equipment_type_id=equipment_type_id,
            )
        )

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=(
                "Only one fallback asset is allowed per field and asset type. "
                "Unset fallback on the current fallback asset first."
            ),
        )
    row = (
        db.query(TemplateFieldAsset)
        .options(selectinload(TemplateFieldAsset.equipment_type_links))
        .filter(TemplateFieldAsset.id == row.id)
        .first()
    )
    return _build_template_field_asset_response(row)


@router.delete("/fields/{field_id}/assets/{asset_id}")
def delete_template_field_asset(
    field_id: int,
    asset_id: int,
    db: Session = Depends(get_db),
):
    _, template_field = _resolve_template_field_from_ebay_field(field_id, db)
    row = (
        db.query(TemplateFieldAsset)
        .filter(
            TemplateFieldAsset.id == asset_id,
            TemplateFieldAsset.template_field_id == template_field.id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Asset not found")
    if bool(row.is_default_fallback):
        fallback_count = (
            db.query(TemplateFieldAsset)
            .filter(
                TemplateFieldAsset.template_field_id == template_field.id,
                TemplateFieldAsset.asset_type == row.asset_type,
                TemplateFieldAsset.is_default_fallback.is_(True),
            )
            .count()
        )
        if fallback_count <= 1:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Cannot delete the only fallback asset for this field/type. "
                    "Create another fallback or switch fallback to a different asset first."
                ),
            )
    db.delete(row)
    db.commit()
    return {"message": "Deleted"}


@router.get("/fields/{field_id}/equipment-type-contents", response_model=List[EbayFieldEquipmentTypeContentResponse])
def list_ebay_field_equipment_type_contents(
    field_id: int,
    db: Session = Depends(get_db),
):
    field = db.query(EbayField).filter(EbayField.id == field_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")

    rows = (
        db.query(EbayFieldEquipmentTypeContent)
        .filter(EbayFieldEquipmentTypeContent.ebay_field_id == field_id)
        .order_by(
            func.coalesce(EbayFieldEquipmentTypeContent.equipment_type_id, -1),
            EbayFieldEquipmentTypeContent.id.asc(),
        )
        .all()
    )
    return rows


@router.put("/fields/{field_id}/equipment-type-contents", response_model=EbayFieldEquipmentTypeContentResponse)
def upsert_ebay_field_equipment_type_content(
    field_id: int,
    request: EbayFieldEquipmentTypeContentUpsertRequest,
    db: Session = Depends(get_db),
):
    field = db.query(EbayField).filter(EbayField.id == field_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")

    is_fallback = bool(request.is_default_fallback) or request.equipment_type_id is None
    target_equipment_type_id = None if is_fallback else request.equipment_type_id

    if target_equipment_type_id is not None:
        equipment_type = db.query(EquipmentType).filter(EquipmentType.id == target_equipment_type_id).first()
        if not equipment_type:
            raise HTTPException(status_code=404, detail="Equipment type not found")

    if is_fallback:
        existing = (
            db.query(EbayFieldEquipmentTypeContent)
            .filter(
                EbayFieldEquipmentTypeContent.ebay_field_id == field_id,
                EbayFieldEquipmentTypeContent.equipment_type_id.is_(None),
            )
            .order_by(EbayFieldEquipmentTypeContent.id.asc())
            .first()
        )
    else:
        existing = (
            db.query(EbayFieldEquipmentTypeContent)
            .filter(
                EbayFieldEquipmentTypeContent.ebay_field_id == field_id,
                EbayFieldEquipmentTypeContent.equipment_type_id == target_equipment_type_id,
            )
            .first()
        )

    if existing:
        existing.html_value = request.html_value
        existing.is_default_fallback = is_fallback
        row = existing
    else:
        row = EbayFieldEquipmentTypeContent(
            ebay_field_id=field_id,
            equipment_type_id=target_equipment_type_id,
            html_value=request.html_value,
            is_default_fallback=is_fallback,
        )
        db.add(row)

    if is_fallback:
        db.flush()
        db.query(EbayFieldEquipmentTypeContent).filter(
            EbayFieldEquipmentTypeContent.ebay_field_id == field_id,
            EbayFieldEquipmentTypeContent.id != row.id,
        ).update({"is_default_fallback": False}, synchronize_session=False)

    db.commit()
    db.refresh(row)
    return row


@router.delete("/fields/{field_id}/equipment-type-contents/{content_id}")
def delete_ebay_field_equipment_type_content(
    field_id: int,
    content_id: int,
    db: Session = Depends(get_db),
):
    row = (
        db.query(EbayFieldEquipmentTypeContent)
        .filter(
            EbayFieldEquipmentTypeContent.id == content_id,
            EbayFieldEquipmentTypeContent.ebay_field_id == field_id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Equipment type content not found")
    db.delete(row)
    db.commit()
    return {"message": "Deleted"}


@router.get("/fields/{field_id}/equipment-type-image-patterns", response_model=List[EbayFieldEquipmentTypeImagePatternResponse])
def list_ebay_field_equipment_type_image_patterns(
    field_id: int,
    db: Session = Depends(get_db),
):
    field = db.query(EbayField).filter(EbayField.id == field_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")

    rows = (
        db.query(EbayFieldEquipmentTypeImagePattern)
        .filter(EbayFieldEquipmentTypeImagePattern.ebay_field_id == field_id)
        .order_by(
            func.coalesce(EbayFieldEquipmentTypeImagePattern.equipment_type_id, -1),
            EbayFieldEquipmentTypeImagePattern.id.asc(),
        )
        .all()
    )
    return rows


@router.put("/fields/{field_id}/equipment-type-image-patterns", response_model=EbayFieldEquipmentTypeImagePatternResponse)
def upsert_ebay_field_equipment_type_image_pattern(
    field_id: int,
    request: EbayFieldEquipmentTypeImagePatternUpsertRequest,
    db: Session = Depends(get_db),
):
    field = db.query(EbayField).filter(EbayField.id == field_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")

    target_equipment_type_id = request.equipment_type_id
    if target_equipment_type_id is not None:
        equipment_type = db.query(EquipmentType).filter(EquipmentType.id == target_equipment_type_id).first()
        if not equipment_type:
            raise HTTPException(status_code=404, detail="Equipment type not found")

    existing = (
        db.query(EbayFieldEquipmentTypeImagePattern)
        .filter(
            EbayFieldEquipmentTypeImagePattern.ebay_field_id == field_id,
            EbayFieldEquipmentTypeImagePattern.equipment_type_id == target_equipment_type_id,
        )
        .first()
    )

    if existing:
        existing.parent_pattern = request.parent_pattern
        existing.variation_pattern = request.variation_pattern
        row = existing
    else:
        row = EbayFieldEquipmentTypeImagePattern(
            ebay_field_id=field_id,
            equipment_type_id=target_equipment_type_id,
            parent_pattern=request.parent_pattern,
            variation_pattern=request.variation_pattern,
        )
        db.add(row)

    db.commit()
    db.refresh(row)
    return row


@router.delete("/fields/{field_id}/equipment-type-image-patterns/{pattern_id}")
def delete_ebay_field_equipment_type_image_pattern(
    field_id: int,
    pattern_id: int,
    db: Session = Depends(get_db),
):
    row = (
        db.query(EbayFieldEquipmentTypeImagePattern)
        .filter(
            EbayFieldEquipmentTypeImagePattern.id == pattern_id,
            EbayFieldEquipmentTypeImagePattern.ebay_field_id == field_id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Image pattern not found")
    db.delete(row)
    db.commit()
    return {"message": "Deleted"}


@router.get("/current/download")
def download_current_ebay_template(mode: str = "inline", db: Session = Depends(get_db)):
    """
    Download the current (latest) eBay template file as-is from disk.
    Returns the bit-for-bit original uploaded file.
    
    Args:
        mode: "inline" (try to display in browser) or "download" (force download)
    """
    # Get current template (newest by uploaded_at, then id desc)
    template = (
        db.query(EbayTemplate)
        .order_by(EbayTemplate.uploaded_at.desc(), EbayTemplate.id.desc())
        .first()
    )
    
    if not template:
        raise HTTPException(status_code=404, detail="No eBay template found")
    
    # Check file exists on disk
    if not os.path.exists(template.file_path):
        raise HTTPException(status_code=400, detail="Template file not found on disk")

    # Lightweight verification instrumentation for current/download consistency.
    with open(template.file_path, "rb") as f:
        disk_bytes = f.read()
    disk_sha256 = hashlib.sha256(disk_bytes).hexdigest()
    print(
        f"[EBAY_DOWNLOAD] id={template.id} db_sha256={template.sha256} "
        f"disk_sha256={disk_sha256} path={template.file_path}"
    )
    
    # Set Content-Disposition based on mode
    disposition = "inline" if mode == "inline" else "attachment"
    
    # Return file
    response = FileResponse(
        path=template.file_path,
        filename=template.original_filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = f'{disposition}; filename="{template.original_filename}"'
    return response


@router.get("/current/preview", response_model=EbayTemplatePreviewResponse)
def preview_current_ebay_template(
    preview_rows: int = 25,
    preview_cols: int = 20,
    db: Session = Depends(get_db)
):
    """
    Preview the current (latest) eBay template as a grid.
    Returns a JSON grid of cell values for the top-left window.
    """
    # Get current template
    template = (
        db.query(EbayTemplate)
        .order_by(EbayTemplate.uploaded_at.desc(), EbayTemplate.id.desc())
        .first()
    )
    
    if not template:
        raise HTTPException(status_code=404, detail="No eBay template found")
    
    # Check file exists on disk
    if not os.path.exists(template.file_path):
        raise HTTPException(status_code=400, detail="Template file not found on disk")
    
    # Load workbook
    try:
        wb = load_workbook(template.file_path, data_only=True)
        
        # Choose sheet: prefer "Template", else first sheet
        if "Template" in wb.sheetnames:
            sheet = wb["Template"]
            sheet_name = "Template"
        else:
            sheet = wb.active
            sheet_name = sheet.title if sheet else "Unknown"
        
        # Get dimensions
        max_row = sheet.max_row or 0
        max_column = sheet.max_column or 0
        
        # Build grid for preview window
        grid = []
        for row_idx in range(1, min(preview_rows + 1, max_row + 1)):
            row_data = []
            for col_idx in range(1, min(preview_cols + 1, max_column + 1)):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                # Convert to string safely
                if value is None:
                    row_data.append("")
                elif isinstance(value, (int, float)):
                    row_data.append(str(value))
                else:
                    row_data.append(str(value).strip())
            grid.append(row_data)
        
        wb.close()
        
        return EbayTemplatePreviewResponse(
            template_id=template.id,
            original_filename=template.original_filename,
            sheet_name=sheet_name,
            max_row=max_row,
            max_column=max_column,
            preview_row_count=len(grid),
            preview_column_count=len(grid[0]) if grid else 0,
            grid=grid
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load template: {str(e)}")


@router.get("/current/integrity", response_model=EbayTemplateIntegrityResponse)
def get_current_ebay_template_integrity(db: Session = Depends(get_db)):
    """
    Get file integrity information for the current (latest) eBay template.
    Returns SHA256 hash, file size, upload timestamp, and filename.
    """
    # Get current template (newest by uploaded_at, then id desc)
    template = (
        db.query(EbayTemplate)
        .order_by(EbayTemplate.uploaded_at.desc(), EbayTemplate.id.desc())
        .first()
    )
    
    if not template:
        raise HTTPException(status_code=404, detail="No eBay template uploaded")
    
    return EbayTemplateIntegrityResponse(
        template_id=template.id,
        original_filename=template.original_filename,
        file_size=template.file_size,
        sha256=template.sha256,
        uploaded_at=template.uploaded_at.isoformat() if template.uploaded_at else None
    )


@router.get("/current/verify", response_model=EbayTemplateVerificationResponse)
def verify_current_ebay_template(db: Session = Depends(get_db)):
    """
    Verify that the current eBay template file on disk matches stored integrity metadata.
    This is a read-only diagnostic endpoint - it does NOT modify the database.
    """
    # Get current template
    template = (
        db.query(EbayTemplate)
        .order_by(EbayTemplate.uploaded_at.desc(), EbayTemplate.id.desc())
        .first()
    )
    
    if not template:
        raise HTTPException(status_code=404, detail="No eBay template uploaded")
    
    verified_at = datetime.utcnow().isoformat()
    stored_sha256 = template.sha256
    stored_file_size = template.file_size
    
    # Check if file exists
    if not os.path.exists(template.file_path):
        return EbayTemplateVerificationResponse(
            template_id=template.id,
            status="missing",
            stored_sha256=stored_sha256,
            stored_file_size=stored_file_size,
            computed_sha256=None,
            computed_file_size=None,
            verified_at=verified_at
        )
    
    # Check if we have a stored hash to compare against
    if not stored_sha256:
        return EbayTemplateVerificationResponse(
            template_id=template.id,
            status="unknown",
            stored_sha256=None,
            stored_file_size=stored_file_size,
            computed_sha256=None,
            computed_file_size=None,
            verified_at=verified_at
        )
    
    # Compute SHA256 and file size from disk
    try:
        sha256_hash = hashlib.sha256()
        computed_file_size = 0
        
        # Stream file in chunks (don't load entire file into memory)
        with open(template.file_path, 'rb') as f:
            while True:
                chunk = f.read(8192)  # 8KB chunks
                if not chunk:
                    break
                sha256_hash.update(chunk)
                computed_file_size += len(chunk)
        
        computed_sha256 = sha256_hash.hexdigest()
        
        # Compare values
        sha256_match = (computed_sha256 == stored_sha256)
        size_match = (computed_file_size == stored_file_size)
        
        if sha256_match and size_match:
            status = "match"
        else:
            status = "mismatch"
        
        return EbayTemplateVerificationResponse(
            template_id=template.id,
            status=status,
            stored_sha256=stored_sha256,
            stored_file_size=stored_file_size,
            computed_sha256=computed_sha256,
            computed_file_size=computed_file_size,
            verified_at=verified_at
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to verify file: {str(e)}")
