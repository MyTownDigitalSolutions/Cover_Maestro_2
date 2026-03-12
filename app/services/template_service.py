import pandas as pd
import json
from io import BytesIO
from sqlalchemy.orm import Session
from fastapi import UploadFile, HTTPException
import os
import hashlib
import shutil
from datetime import datetime
from openpyxl import load_workbook
from app.models.templates import AmazonProductType, ProductTypeKeyword, ProductTypeField, ProductTypeFieldValue
from sqlalchemy import or_
import re
from app.services.storage_policy import (
    ensure_storage_dirs_exist,
    assert_allowed_write_path,
    TEMPLATE_DIR,
)


# ============================================================================
# NORMALIZATION HELPER
# ============================================================================
def normalize_product_type_key(raw: str) -> str:
    """
    Normalize product type key for filename generation.
    Rules:
    - Trim whitespace
    - Replace spaces with underscores
    - Remove characters not in [A-Za-z0-9_()-]
    - Collapse multiple underscores to single underscore
    """
    # Trim
    normalized = raw.strip()

    # Replace spaces with underscores
    normalized = normalized.replace(" ", "_")

    # Remove invalid characters (keep only A-Z, a-z, 0-9, _, (), -)
    normalized = re.sub(r'[^A-Za-z0-9_()-]', '', normalized)

    # Collapse multiple underscores
    normalized = re.sub(r'_+', '_', normalized)

    return normalized


# ============================================================================
# PATH HELPERS
# ============================================================================
def get_template_paths(product_type_key: str) -> tuple[str, str]:
    """
    Get canonical and backup paths for a product type template.
    Returns: (canonical_path, backup_path)
    """
    normalized_key = normalize_product_type_key(product_type_key)
    canonical_filename = f"{normalized_key}(Template).xlsx"
    backup_filename = f"{normalized_key}(Template)_BACKUP.xlsx"

    canonical_path = os.path.join(TEMPLATE_DIR, canonical_filename)
    backup_path = os.path.join(TEMPLATE_DIR, backup_filename)

    return canonical_path, backup_path


def rotate_template_backup(canonical_path: str, backup_path: str):
    """
    Rotate template backup: move existing canonical to backup (overwriting old backup).
    """
    if os.path.exists(canonical_path):
        # Move canonical to backup (overwrite existing backup)
        if os.path.exists(backup_path):
            os.remove(backup_path)
        shutil.move(canonical_path, backup_path)
        print(f"[TEMPLATE_ROTATION] Moved {canonical_path} -> {backup_path}")


def _normalize_sheet_name(name: str) -> str:
    """Normalize sheet names for resilient matching."""
    return re.sub(r'[^a-z0-9]+', '', (name or '').strip().lower())


def _pick_sheet_name(sheet_names: list[str], exact_candidates: list[str], contains_candidates: list[str] | None = None) -> str | None:
    """
    Pick a sheet name by normalized exact match first, then normalized contains match.
    """
    contains_candidates = contains_candidates or []
    normalized_lookup = {s: _normalize_sheet_name(s) for s in sheet_names}
    exact_set = {_normalize_sheet_name(c) for c in exact_candidates}
    contains_set = [_normalize_sheet_name(c) for c in contains_candidates]

    for original, normalized in normalized_lookup.items():
        if normalized in exact_set:
            return original

    for original, normalized in normalized_lookup.items():
        if any(token and token in normalized for token in contains_set):
            return original

    return None


class TemplateService:
    def __init__(self, db: Session):
        self.db = db

    async def import_amazon_template(self, file: UploadFile, product_code: str) -> dict:
        """
        Import Amazon template with this EXACT logic:

        STEP 1: DATA DEFINITIONS sheet - ONLY get:
          - Group names (Column A when Column B is empty)
          - Field names (Column B)
          - Local Label names (Column C)
          - Column D is just descriptions, NOT valid values!

        STEP 2: VALID VALUES sheet - Get selectable options:
          - Column A = Group name (when Column B empty)
          - Column B = "Local Label - [field_hint]" format
          - Column C onwards = The actual valid values users can select

        STEP 3: DEFAULT VALUES sheet - Get defaults and additional values:
          - Column A = Local Label Name
          - Column B = Field Name
          - Column C = Default value to pre-select
          - Column D onwards = Additional values to ADD to valid values

        STEP 4: TEMPLATE sheet - Get field order for export
        """
        # Ensure storage directories exist
        ensure_storage_dirs_exist()

        # Get paths for this product type
        canonical_path, backup_path = get_template_paths(product_code)

        # Validate write paths
        assert_allowed_write_path(canonical_path)
        assert_allowed_write_path(backup_path)

        # Rotate existing canonical to backup
        rotate_template_backup(canonical_path, backup_path)

        # 1. Read into memory ONCE
        await file.seek(0)
        contents = await file.read()
        upload_sha256 = hashlib.sha256(contents).hexdigest()

        # 2. Write to canonical path
        with open(canonical_path, "wb") as out_file:
            out_file.write(contents)

        # 3. Verify
        file_size = os.path.getsize(canonical_path)

        with open(canonical_path, "rb") as f:
            persisted_bytes = f.read()
            persisted_sha256 = hashlib.sha256(persisted_bytes).hexdigest()

        print(f"[TEMPLATE_IMPORT] product_code={product_code} upload_sha256={upload_sha256} persisted_sha256={persisted_sha256} mem_size={len(contents)} disk_size={file_size}")

        if upload_sha256 != persisted_sha256:
            raise HTTPException(status_code=500, detail=f"Persisted file hash mismatch: mem={upload_sha256} disk={persisted_sha256}")

        # [PT_IMPORT] A) & B) Identification & Sheet Discovery
        print(f"[PT_IMPORT] START: product_code={product_code} filename='{file.filename}'")
        sheet_names: list[str] = []
        try:
            wb_temp = load_workbook(BytesIO(contents), read_only=True)
            sheet_names = wb_temp.sheetnames
            print(f"[PT_IMPORT] Sheets found: {sheet_names}")
            
            # [PT_IMPORT] Check for required sheets
            req_sheets = ["Data Definitions", "Valid Values", "Template", "Default Values"]
            for req in req_sheets:
                found_exact = req in sheet_names
                found_loose = any(s.lower() == req.lower() for s in sheet_names)
                used_name = next((s for s in sheet_names if s.lower() == req.lower()), "None")
                print(f"[PT_IMPORT] Sheet '{req}': Found={found_loose} (Exact={found_exact}) -> Using: '{used_name}'")

            wb_temp.close()
        except Exception as e:
            print(f"[PT_IMPORT] Sheet discovery error: {e}")

        resolved_sheets = {
            "data_definitions": _pick_sheet_name(
                sheet_names,
                exact_candidates=["Data Definitions"],
                contains_candidates=["data definitions", "definitions"],
            ),
            "valid_values": _pick_sheet_name(
                sheet_names,
                exact_candidates=["Valid Values"],
                contains_candidates=["valid values"],
            ),
            "template": _pick_sheet_name(
                sheet_names,
                exact_candidates=["Template"],
                contains_candidates=["template"],
            ),
            "default_values": _pick_sheet_name(
                sheet_names,
                exact_candidates=["Default Values"],
                contains_candidates=["default values", "defaults"],
            ),
        }

        missing = [k for k, v in resolved_sheets.items() if v is None]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Invalid Amazon Product Type template: missing required sheets "
                    f"{missing}. Found sheets ({len(sheet_names)}): {sheet_names}"
                ),
            )

        # 4. Use BytesIO for processing
        try:
            excel_file = BytesIO(contents)
            print(f"[TEMPLATE_IMPORT] BytesIO created successfully, size={len(contents)}")
        except Exception as e:
            import traceback
            print(f"[TEMPLATE_IMPORT] ERROR creating BytesIO: {e}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Failed to create BytesIO: {str(e)}")

        existing = self.db.query(AmazonProductType).filter(
            AmazonProductType.code == product_code
        ).first()

        existing_field_settings = {}
        if existing:
            existing_fields = self.db.query(ProductTypeField).filter(
                ProductTypeField.product_type_id == existing.id
            ).all()
            for field in existing_fields:
                existing_field_settings[field.field_name] = {
                    'required': field.required,
                    'selected_value': field.selected_value,
                    'custom_value': field.custom_value
                }

            self.db.query(ProductTypeFieldValue).filter(
                ProductTypeFieldValue.product_type_field_id.in_(
                    self.db.query(ProductTypeField.id).filter(
                        ProductTypeField.product_type_id == existing.id
                    )
                )
            ).delete(synchronize_session=False)
            self.db.query(ProductTypeField).filter(
                ProductTypeField.product_type_id == existing.id
            ).delete(synchronize_session=False)
            self.db.query(ProductTypeKeyword).filter(
                ProductTypeKeyword.product_type_id == existing.id
            ).delete(synchronize_session=False)
            self.db.commit()
            product_type = existing
        else:
            product_type = AmazonProductType(
                code=product_code,
                name=product_code.replace("_", " ").title()
            )
            self.db.add(product_type)
            self.db.commit()
            self.db.refresh(product_type)

        # Update metadata (use canonical path)
        product_type.original_filename = file.filename
        product_type.file_path = canonical_path
        product_type.file_size = file_size
        product_type.upload_date = datetime.utcnow()
        self.db.commit()

        fields_imported = 0
        keywords_imported = 0
        valid_values_imported = 0
        vv_mapping_misses = 0  # [PT_IMPORT] Track valid value mapping failures

        field_definitions = {}
        local_label_to_field = {}

        print("=" * 60)
        print("STEP 1: Parsing DATA DEFINITIONS sheet")
        print("=" * 60)

        dd_df = None
        try:
            # ✅ CRITICAL: rewind BEFORE every read
            excel_file.seek(0)
            dd_df = pd.read_excel(excel_file, sheet_name=resolved_sheets["data_definitions"], header=None)
            excel_file.seek(0)

            # Minimal diagnostics
            print(f"[TEMPLATE_IMPORT] Data Definitions shape={dd_df.shape}")
            try:
                a0 = dd_df.iat[0, 0] if dd_df.shape[0] > 0 and dd_df.shape[1] > 0 else None
                b0 = dd_df.iat[0, 1] if dd_df.shape[0] > 0 and dd_df.shape[1] > 1 else None
                print(f"[TEMPLATE_IMPORT] Data Definitions top-left A1={a0} B1={b0}")
            except Exception:
                pass

        except ValueError as e:
            if "Worksheet named" in str(e) or "not found" in str(e).lower():
                raise HTTPException(status_code=400, detail="Invalid Amazon Product Type template: missing 'Data Definitions' sheet")
            import traceback
            print(f"[TEMPLATE_IMPORT] ERROR reading Data Definitions sheet: {e}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Failed to read Data Definitions sheet: {str(e)}")
        except Exception as e:
            import traceback
            print(f"[TEMPLATE_IMPORT] ERROR reading Data Definitions sheet: {e}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Failed to read Data Definitions sheet: {str(e)}")

        # ✅ Parse AFTER successful read
        try:
            current_group = None

            for row_idx in range(2, len(dd_df)):
                row = dd_df.iloc[row_idx]

                col_a = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
                col_b = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else None
                col_c = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else None

                if col_a and not col_b:
                    current_group = col_a
                    # print(f"  GROUP: {current_group}")
                    continue

                if col_b:
                    field_name = col_b
                    local_label = col_c

                    field_definitions[field_name] = {
                        "group_name": current_group,
                        "local_label": local_label,
                    }

                    if local_label:
                        local_label_to_field[local_label] = field_name

            print(f"  TOTAL: {len(field_definitions)} field definitions")
            
            # [PT_IMPORT] Data Defs Details
            print(f"[PT_IMPORT] Data Definitions loaded. Mapping size: {len(local_label_to_field)}")
            print(f"[PT_IMPORT] Sample Mappings: {list(local_label_to_field.items())[:5]}")
        except Exception as e:
            print(f"  ERROR parsing Data Definitions: {e}")

        valid_values_by_field = {}
        vv_distinct_labels = set()

        print("=" * 60)
        print("STEP 2: Parsing VALID VALUES sheet")
        print("=" * 60)

        try:
            excel_file.seek(0)
            vv_df = pd.read_excel(excel_file, sheet_name=resolved_sheets["valid_values"], header=None)
            excel_file.seek(0)

            print(f"[PT_IMPORT] Valid Values shape={vv_df.shape}")

            current_group = None
            for row_idx in range(1, len(vv_df)):
                row = vv_df.iloc[row_idx]
                col_a_group = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
                col_b_local_hint = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else None

                if col_a_group and not col_b_local_hint:
                    current_group = col_a_group
                    continue

                if not col_b_local_hint:
                    continue

                # Extract local label from "Local Label - [field_hint]"
                local_label = col_b_local_hint.split(" - ")[0].strip() if " - " in col_b_local_hint else col_b_local_hint
                vv_distinct_labels.add(local_label)
                
                field_name = local_label_to_field.get(local_label)

                if not field_name:
                    # Try fuzzy match on local label to field
                    for ll, fn in local_label_to_field.items():
                        if ll and local_label and (ll.lower() == local_label.lower()):
                            field_name = fn
                            break

                if not field_name:
                    vv_mapping_misses += 1
                    print(f"[PT_IMPORT][VV_MISS] local_label='{local_label}' (no match in local_label_to_field)")
                    continue

                values = [str(v).strip() for v in row.iloc[2:] if pd.notna(v)]
                if not values:
                    continue

                if field_name not in valid_values_by_field:
                    valid_values_by_field[field_name] = []
                for v in values:
                    if v not in valid_values_by_field[field_name]:
                        valid_values_by_field[field_name].append(v)

                valid_values_imported += len(values)

            print(f"  TOTAL: {len(valid_values_by_field)} fields with valid values")
            print(f"[PT_IMPORT] Valid Values Summary: Misses={vv_mapping_misses}, Distinct Labels={len(vv_distinct_labels)}, Fields Populated={len(valid_values_by_field)}")
        except Exception as e:
            print(f"  Valid Values sheet: {e}")

        template_field_order = {}

        print("=" * 60)
        print("STEP 3: Parsing TEMPLATE sheet for field order")
        print("=" * 60)

        try:
            excel_file.seek(0)
            template_df = pd.read_excel(excel_file, sheet_name=resolved_sheets["template"], header=None)
            excel_file.seek(0)

            print(f"[TEMPLATE_IMPORT] Template shape={template_df.shape}")
        except ValueError as e:
            if "Worksheet named" in str(e) or "not found" in str(e).lower():
                raise HTTPException(status_code=400, detail="Invalid Amazon Product Type template: missing 'Template' sheet")
            import traceback
            print(f"[TEMPLATE_IMPORT] ERROR reading Template sheet: {e}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Failed to read Template sheet: {str(e)}")
        except Exception as e:
            import traceback
            print(f"[TEMPLATE_IMPORT] ERROR reading Template sheet: {e}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Failed to read Template sheet: {str(e)}")

        try:
            header_rows = []
            for i in range(min(6, len(template_df))):
                row_data = []
                for val in template_df.iloc[i]:
                    if pd.isna(val):
                        row_data.append(None)
                    else:
                        row_data.append(str(val))
                header_rows.append(row_data)

            product_type.header_rows = header_rows

            row5_field_names = template_df.iloc[4].tolist() if len(template_df) > 4 else []
            row4_display_names = template_df.iloc[3].tolist() if len(template_df) > 3 else []
            row3_groups = template_df.iloc[2].tolist() if len(template_df) > 2 else []

            current_group = None
            for idx, field_name in enumerate(row5_field_names):
                if pd.isna(field_name):
                    continue

                field_name_str = str(field_name).strip()

                group_from_template = (
                    str(row3_groups[idx]).strip()
                    if idx < len(row3_groups) and pd.notna(row3_groups[idx])
                    else None
                )
                if group_from_template:
                    current_group = group_from_template

                display_name = (
                    str(row4_display_names[idx]).strip()
                    if idx < len(row4_display_names) and pd.notna(row4_display_names[idx])
                    else None
                )

                template_field_order[field_name_str] = {
                    "order_index": idx,
                    "display_name": display_name,
                    "group_from_template": current_group,
                }

            self.db.commit()
            print(f"  TOTAL: {len(template_field_order)} fields in template")

        except Exception as e:
            print(f"  ERROR: {e}")

        all_known_fields = set(field_definitions.keys()) | set(template_field_order.keys())

        default_values_by_field = {}
        other_values_by_field = {}

        print("=" * 60)
        print("STEP 4: Parsing DEFAULT VALUES sheet")
        print("=" * 60)

        try:
            # ✅ CRITICAL: rewind BEFORE read
            excel_file.seek(0)
            dv_df = pd.read_excel(excel_file, sheet_name=resolved_sheets["default_values"], header=None)
            excel_file.seek(0)

            print(f"[TEMPLATE_IMPORT] Default Values shape={dv_df.shape}")

            for row_idx in range(1, len(dv_df)):
                row = dv_df.iloc[row_idx]

                col_a_local_label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
                col_b_field_name = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else None
                col_c_default = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else None

                if not col_a_local_label and not col_b_field_name:
                    continue

                matched_field = None

                if col_b_field_name and col_b_field_name in all_known_fields:
                    matched_field = col_b_field_name
                elif col_a_local_label and col_a_local_label in local_label_to_field:
                    matched_field = local_label_to_field[col_a_local_label]

                if not matched_field and col_b_field_name:
                    for fn in all_known_fields:
                        if col_b_field_name in fn or fn in col_b_field_name:
                            matched_field = fn
                            break

                if matched_field:
                    if col_c_default:
                        default_values_by_field[matched_field] = col_c_default

                    other_values = [str(v).strip() for v in row.iloc[3:] if pd.notna(v)]
                    if other_values:
                        if matched_field not in other_values_by_field:
                            other_values_by_field[matched_field] = []
                        other_values_by_field[matched_field].extend(other_values)

            print(f"  TOTAL: {len(default_values_by_field)} defaults, {len(other_values_by_field)} with other values")

        except Exception as e:
            print(f"  Default Values sheet: {e}")

        print("=" * 60)
        print("STEP 5: Creating database records")
        print("=" * 60)

        print("[PT_IMPORT] About to insert valid values...")
        max_field_id_inserted = 0
        total_vv_inserted_count = 0

        field_name_to_db = {}

        # [PT_IMPORT] Helper for fan-out
        def _group_key(fn: str) -> str:
            return fn.split("#")[0]

        for field_name, template_info in template_field_order.items():
            dd_info = field_definitions.get(field_name, {})

            group_name = dd_info.get("group_name") or template_info.get("group_from_template")
            local_label = dd_info.get("local_label")
            display_name = template_info.get("display_name") or local_label
            default_value = default_values_by_field.get(field_name)

            prev_settings = existing_field_settings.get(field_name, {})

            # [PT_IMPORT] Fan-out prep
            direct_vv = valid_values_by_field.get(field_name, [])
            fanout_values = []
            
            if not direct_vv:
                gk = _group_key(field_name)
                # Check for other fields sharing this group key
                src_keys_count = 0
                seen_fanout = set()
                
                for k, vals in valid_values_by_field.items():
                    if _group_key(k) == gk:
                        src_keys_count += 1
                        for v in vals:
                            if v not in seen_fanout:
                                seen_fanout.add(v)
                                fanout_values.append(v)
                
                if fanout_values:
                     print(f"[PT_IMPORT][VV_FANOUT] field_name='{field_name}' used_group='{gk}' src_keys={src_keys_count} values={len(fanout_values)}")

            # Determine whether this field should behave like a selectable dropdown.
            # Deterministic rule: if the template provides ANY valid/other values for this field,
            # we treat it as selectable and store defaults in selected_value (NOT custom_value).
            has_selectable_values = (
                bool(direct_vv) or bool(fanout_values) or field_name in other_values_by_field
            )

            # Build the final list of values for this field (used for reconciliation below)
            all_values = []
            
            if direct_vv:
                all_values.extend(direct_vv)
            elif fanout_values:
                all_values.extend(fanout_values)
                
            if field_name in other_values_by_field:
                for ov in other_values_by_field[field_name]:
                    if ov not in all_values:
                        all_values.append(ov)
            if default_value and default_value not in all_values:
                all_values.insert(0, default_value)

            # Apply template defaults on re-upload.
            # Critical fix: if a previous selected_value no longer exists in the new valid values,
            # we MUST clear it and choose the template default (or first available value) so the
            # main page doesn't show a stale selection.
            # [PT_IMPORT] Application of Defaults
            if has_selectable_values:
                # Logic: Default Sheet > None (Authoritative Reset)
                
                if default_value:
                     selected_value = default_value
                     # Check strict validity for logging warning
                     is_strictly_valid = (default_value in direct_vv) or (default_value in fanout_values)
                     if not is_strictly_valid and field_name in other_values_by_field:
                         is_strictly_valid = default_value in other_values_by_field[field_name]
                         
                     if is_strictly_valid:
                         print(f"[PT_IMPORT][DEFAULT_SET] field='{field_name}' value='{default_value}' source='Default Values'")
                     else:
                         print(f"[PT_IMPORT][DEFAULT_WARN] field='{field_name}' default='{default_value}' note='Default not in valid values list'")
                else:
                     print(f"[PT_IMPORT][DEFAULT_CLEARED] field='{field_name}' reason='Missing/blank in Default Values sheet'")
                     selected_value = None

                custom_value = None
            else:
                # Free-text field: store defaults in custom_value
                if default_value:
                    print(f"[PT_IMPORT][DEFAULT_SET] field='{field_name}' value='{default_value}' source='Default Values'")
                    custom_value = default_value
                else:
                    print(f"[PT_IMPORT][DEFAULT_CLEARED] field='{field_name}' reason='Missing/blank in Default Values sheet'")
                    custom_value = None
                
                selected_value = None

            field = ProductTypeField(
                product_type_id=product_type.id,
                field_name=field_name,
                display_name=display_name,
                attribute_group=group_name,
                order_index=template_info["order_index"],
                required=prev_settings.get('required', False),
                selected_value=selected_value,
                custom_value=custom_value
            )
            self.db.add(field)
            self.db.flush()
            field_name_to_db[field_name] = field
            fields_imported += 1
            if field.id > max_field_id_inserted:
                max_field_id_inserted = field.id
            
            # [PT_IMPORT]
            if len(all_values) == 0:
                 has_default = len(default_values_by_field.get(field_name, "")) > 0
                 print(f"[PT_IMPORT][NO_VALUES] field_name='{field_name}' -> all_values empty (valid=0 other=0 default_present={has_default})")
            elif len(all_values) > 0:
                 # Minimal log for success to verify process
                 pass

            # all_values already built above
            for value in all_values:
                field_value = ProductTypeFieldValue(
                    product_type_field_id=field.id,
                    value=value
                )
                self.db.add(field_value)
                total_vv_inserted_count += 1

        self.db.commit()

        print(f"  Created {fields_imported} fields")
        print(f"[PT_IMPORT] Inserted {total_vv_inserted_count} product_type_field_values rows. Max ID: {max_field_id_inserted}")
        print(f"[PT_IMPORT] Final Summary: fields_created={fields_imported}, values_inserted={total_vv_inserted_count}, vv_fields_count={len(valid_values_by_field)}, vv_miss_count={vv_mapping_misses}")
        print("=" * 60)
        print(f"DONE: {fields_imported} fields, {keywords_imported} keywords, {valid_values_imported} valid values")
        print("=" * 60)

        # ADDITIVE: Run non-destructive field indexer to find sparse fields skipped by pandas
        additive_fields = self._build_field_index(canonical_path, product_type)
        fields_imported += additive_fields

        return {
            "product_code": product_code,
            "fields_imported": fields_imported,
            "keywords_imported": keywords_imported,
            "valid_values_imported": valid_values_imported,
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
        }

    def _build_field_index(self, file_path: str, product_type: AmazonProductType) -> int:
        """
        Additive field indexer using openpyxl.
        Scans 'Template' sheet for columns that might have been skipped by pandas due to empty header rows.
        Only adds NEW fields. Does NOT modify existing ones.
        """
        added_count = 0
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            if "Template" not in wb.sheetnames:
                return 0

            ws = wb["Template"]

            # Row 5 (1-based) is Field Names in standardized Amazon templates
            # Row 3 is Group
            # Row 4 is Display Name

            # Helper to safely get value
            def get_val(r, c):
                v = ws.cell(row=r, column=c).value
                return str(v).strip() if v is not None else None

            max_col = ws.max_column

            # Get existing fields to avoid duplicates
            existing_fields = {
                f.field_name for f in self.db.query(ProductTypeField).filter(
                    ProductTypeField.product_type_id == product_type.id
                ).all()
            }

            # Scan all columns up to max_col
            print(f"[FIELD_INDEX] product_code={product_type.code} max_col={max_col} existing_count={len(existing_fields)}")

            for col in range(1, max_col + 1):
                field_name = get_val(5, col)

                if field_name and field_name not in existing_fields:
                    # Found a field skipped by pandas!
                    group_name = get_val(3, col)
                    display_name = get_val(4, col) or field_name

                    # Add it
                    new_field = ProductTypeField(
                        product_type_id=product_type.id,
                        field_name=field_name,
                        display_name=display_name,
                        attribute_group=group_name,
                        order_index=col - 1, # approximation
                        required=False
                    )
                    self.db.add(new_field)
                    existing_fields.add(field_name) # IMMEDIATE guard against duplicates in same pass
                    added_count += 1
                    print(f"  [Additive Index] Found sparse field: {field_name}")

            if added_count > 0:
                self.db.commit()
                print(f"[FIELD_INDEX] Successfully added {added_count} new fields")
            else:
                print("[FIELD_INDEX] No new fields found")

            wb.close()

        except Exception as e:
            print(f"  [Additive Index] Error: {e}")

        return added_count

    def get_header_rows(self, product_code: str) -> list:
        product_type = self.db.query(AmazonProductType).filter(
            AmazonProductType.code == product_code
        ).first()

        if product_type and product_type.header_rows:
            return product_type.header_rows
        return []


