import csv
import os
import re
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
TEMPLATE_SHEET_NAME = "Template"
VALID_VALUES_SHEET_NAME = "Valid Values"
DEFAULT_VALUES_SHEET_NAME = "Default Values"
FIELD_HEADER_ALIASES = {"field_name", "field_names", "field", "field name", "name"}


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_sheet_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name or "").strip().lower())


def _normalize_header_cell(value: object) -> str:
    return re.sub(r"[^a-z0-9?]+", "", _clean_cell(value).lower())


NORMALIZED_FIELD_HEADER_ALIASES = {
    _normalize_header_cell(alias) for alias in FIELD_HEADER_ALIASES
}


def _parse_required_flag(value: object) -> Optional[bool]:
    normalized = _clean_cell(value).strip().upper()
    if normalized == "YES":
        return True
    if normalized == "NO":
        return False
    return None


def _find_sheet_name_exact(sheet_names: List[str], required_name: str) -> Optional[str]:
    required_norm = _normalize_sheet_name(required_name)
    for sheet_name in sheet_names or []:
        if _normalize_sheet_name(sheet_name) == required_norm:
            return sheet_name
    return None


def _require_sheet_name(sheet_names: List[str], required_name: str) -> str:
    found = _find_sheet_name_exact(sheet_names, required_name)
    if found:
        return found
    raise ValueError(
        f"Invalid Reverb workbook template: missing required sheet '{required_name}'. "
        f"Found sheets: {list(sheet_names or [])}"
    )


def is_reverb_workbook_template(file_path: str, original_filename: Optional[str] = None) -> bool:
    name = (original_filename or file_path or "").strip().lower()
    ext = os.path.splitext(name)[1]
    return ext in WORKBOOK_EXTENSIONS


def _read_template_headers_from_worksheet(ws) -> List[str]:
    for row in ws.iter_rows(values_only=True):
        values = [_clean_cell(v) for v in row]
        if not any(values):
            continue
        headers = [cell for cell in values if cell]
        if not headers:
            continue
        return headers
    raise ValueError("Invalid Reverb workbook template: Template sheet is empty or missing header row.")


def _parse_valid_values_sheet(ws) -> Tuple[Dict[str, List[str]], Dict[str, bool]]:
    valid_values_by_field: Dict[str, List[str]] = {}
    required_by_field: Dict[str, bool] = {}
    rows = list(ws.iter_rows(values_only=True))
    has_required_column = False
    header_checked = False

    for row in rows:
        values = [_clean_cell(v) for v in row]
        if not any(values):
            continue
        field_name = values[0]
        if not field_name:
            continue
        first_header = _normalize_header_cell(values[0])
        second_header = _normalize_header_cell(values[1] if len(values) > 1 else "")
        if not header_checked:
            header_checked = True
            if first_header in NORMALIZED_FIELD_HEADER_ALIASES:
                has_required_column = second_header == "required?"
                continue

        values_start_idx = 1
        explicit_required = None
        if has_required_column:
            explicit_required = _parse_required_flag(values[1] if len(values) > 1 else "")
            values_start_idx = 2

        if explicit_required is not None:
            required_by_field[field_name] = explicit_required

        parsed_values: List[str] = []
        for candidate in values[values_start_idx:]:
            if candidate and candidate not in parsed_values:
                parsed_values.append(candidate)
        if not parsed_values:
            continue
        existing = valid_values_by_field.setdefault(field_name, [])
        for candidate in parsed_values:
            if candidate not in existing:
                existing.append(candidate)
    return valid_values_by_field, required_by_field


def _parse_default_values_sheet(ws) -> Dict[str, Dict[str, str]]:
    rows = [[_clean_cell(v) for v in row] for row in ws.iter_rows(values_only=True)]
    rows = [row for row in rows if any(row)]
    if not rows:
        return {}

    header_row = [c.strip().lower() for c in rows[0]]

    def _find_index(aliases: List[str]) -> Optional[int]:
        for alias in aliases:
            norm = alias.strip().lower()
            if norm in header_row:
                return header_row.index(norm)
        return None

    field_idx = _find_index(["field_name", "field", "field name", "name"])
    selected_idx = _find_index(["selected_value", "selected", "default", "default_value", "value"])
    custom_idx = _find_index(["custom_value", "custom", "custom default", "custom_default"])

    if field_idx is None:
        raise ValueError(
            "Invalid Reverb workbook template: Default Values sheet requires a 'field_name' column."
        )
    if selected_idx is None and custom_idx is None:
        raise ValueError(
            "Invalid Reverb workbook template: Default Values sheet requires "
            "'selected_value' and/or 'custom_value' columns."
        )

    defaults_by_field: Dict[str, Dict[str, str]] = {}
    for row in rows[1:]:
        field_name = row[field_idx] if field_idx < len(row) else ""
        if not field_name:
            continue
        selected_value = row[selected_idx] if selected_idx is not None and selected_idx < len(row) else ""
        custom_value = row[custom_idx] if custom_idx is not None and custom_idx < len(row) else ""

        payload: Dict[str, str] = {}
        if selected_value:
            payload["selected_value"] = selected_value
        if custom_value:
            payload["custom_value"] = custom_value
        if payload:
            defaults_by_field[field_name] = payload

    return defaults_by_field


def parse_reverb_workbook(file_path: str) -> Dict[str, object]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        template_sheet = _require_sheet_name(wb.sheetnames, TEMPLATE_SHEET_NAME)
        valid_values_sheet = _require_sheet_name(wb.sheetnames, VALID_VALUES_SHEET_NAME)
        default_values_sheet = _require_sheet_name(wb.sheetnames, DEFAULT_VALUES_SHEET_NAME)

        ws_template = wb[template_sheet]
        ws_valid = wb[valid_values_sheet]
        ws_default = wb[default_values_sheet]

        headers = _read_template_headers_from_worksheet(ws_template)
        valid_values_by_field, required_by_field = _parse_valid_values_sheet(ws_valid)
        defaults_by_field = _parse_default_values_sheet(ws_default)

        return {
            "headers": headers,
            "template_sheet_name": ws_template.title,
            "valid_values_sheet_name": ws_valid.title,
            "default_values_sheet_name": ws_default.title,
            "valid_values_by_field": valid_values_by_field,
            "required_by_field": required_by_field,
            "defaults_by_field": defaults_by_field,
        }
    finally:
        wb.close()


def parse_reverb_csv(file_path: str) -> Dict[str, object]:
    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        header_row = next(reader, [])
        value_row = next(reader, [])

    headers = [_clean_cell(v) for v in header_row if _clean_cell(v)]
    if not headers:
        raise ValueError("Invalid Reverb CSV template: missing header row.")

    defaults_by_field: Dict[str, Dict[str, str]] = {}
    if value_row:
        for idx, header in enumerate(headers):
            if idx >= len(value_row):
                continue
            value = _clean_cell(value_row[idx])
            if value:
                defaults_by_field[header] = {"custom_value": value}

    return {
        "headers": headers,
        "template_sheet_name": "csv",
        "valid_values_sheet_name": None,
        "default_values_sheet_name": "csv_row_2" if defaults_by_field else None,
        "valid_values_by_field": {},
        "required_by_field": {},
        "defaults_by_field": defaults_by_field,
    }


def load_reverb_runtime_template(file_path: str, original_filename: Optional[str] = None) -> Dict[str, object]:
    if is_reverb_workbook_template(file_path=file_path, original_filename=original_filename):
        return parse_reverb_workbook(file_path)
    return parse_reverb_csv(file_path)


def load_reverb_template_headers(file_path: str, original_filename: Optional[str] = None) -> Tuple[List[str], str]:
    parsed = load_reverb_runtime_template(file_path=file_path, original_filename=original_filename)
    headers = list(parsed.get("headers", []) or [])
    sheet_name = str(parsed.get("template_sheet_name") or "csv")
    return headers, sheet_name


def read_reverb_template_preview(
    file_path: str,
    original_filename: Optional[str] = None,
    preview_rows: int = 25,
    preview_cols: int = 20,
) -> Dict[str, object]:
    if is_reverb_workbook_template(file_path=file_path, original_filename=original_filename):
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            template_sheet = _require_sheet_name(wb.sheetnames, TEMPLATE_SHEET_NAME)
            ws = wb[template_sheet]

            grid: List[List[str]] = []
            max_col_seen = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= preview_rows:
                    break
                values = [_clean_cell(v) for v in row]
                if len(values) > max_col_seen:
                    max_col_seen = len(values)
                grid.append(values[:preview_cols])

            return {
                "sheet_name": ws.title,
                "max_row": int(ws.max_row or len(grid)),
                "max_column": int(ws.max_column or max_col_seen),
                "grid": grid,
            }
        finally:
            wb.close()

    grid = []
    max_col = 0
    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i >= preview_rows:
                break
            clean_row = [_clean_cell(v) for v in row]
            if len(clean_row) > max_col:
                max_col = len(clean_row)
            grid.append(clean_row[:preview_cols])

    return {
        "sheet_name": "csv",
        "max_row": len(grid),
        "max_column": max_col,
        "grid": grid,
    }
