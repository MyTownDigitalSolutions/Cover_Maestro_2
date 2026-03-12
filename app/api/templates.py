from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session, selectinload
from typing import List
from app.database import get_db
from app.models.core import EquipmentType
from app.models.templates import AmazonProductType, ProductTypeField, EquipmentTypeProductType, ProductTypeFieldValue
from app.schemas.templates import (
    AmazonProductTypeResponse, ProductTypeFieldResponse, TemplateImportResponse,
    EquipmentTypeProductTypeLinkCreate, EquipmentTypeProductTypeLinkResponse,
    ProductTypeFieldUpdate, ProductTypeFieldValueCreate, ProductTypeFieldValueResponse,
    AmazonProductTypeTemplatePreviewResponse, ProductTypeExportConfigUpdate
)
import os
from fastapi.responses import FileResponse
from app.services.template_service import TemplateService

router = APIRouter(prefix="/templates", tags=["templates"])


# PATCH Endpoint to update export configuration

@router.patch("/{product_code}/export-config", response_model=AmazonProductTypeResponse)
def update_product_type_export_config(
    product_code: str, 
    config: ProductTypeExportConfigUpdate, 
    db: Session = Depends(get_db)
):
    product_type = db.query(AmazonProductType).filter(AmazonProductType.code == product_code).first()
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found")
        
    update_data = config.model_dump(exclude_unset=True)
    
    if "export_sheet_name_override" in update_data:
        product_type.export_sheet_name_override = update_data["export_sheet_name_override"]
        
    if "export_start_row_override" in update_data:
        product_type.export_start_row_override = update_data["export_start_row_override"]
        
    db.commit()
    db.refresh(product_type)
    return product_type

@router.post("/import", response_model=TemplateImportResponse)
async def import_template(
    file: UploadFile = File(...),
    product_code: str = Form(...),
    db: Session = Depends(get_db)
):
    service = TemplateService(db)
    result = await service.import_amazon_template(file, product_code)
    return result

@router.get("", response_model=List[AmazonProductTypeResponse])
def list_product_types(db: Session = Depends(get_db)):
    return db.query(AmazonProductType).options(
        selectinload(AmazonProductType.fields).selectinload(ProductTypeField.valid_values),
        selectinload(AmazonProductType.keywords)
    ).all()

@router.post("/equipment-type-links", response_model=EquipmentTypeProductTypeLinkResponse)
def link_equipment_type_to_product_type(
    link: EquipmentTypeProductTypeLinkCreate,
    db: Session = Depends(get_db)
):
    equipment_type = db.query(EquipmentType).filter(EquipmentType.id == link.equipment_type_id).first()
    if not equipment_type:
        raise HTTPException(status_code=404, detail="Equipment type not found")
    
    product_type = db.query(AmazonProductType).filter(AmazonProductType.id == link.product_type_id).first()
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found")
    
    existing = db.query(EquipmentTypeProductType).filter(
        EquipmentTypeProductType.equipment_type_id == link.equipment_type_id
    ).first()
    if existing:
        if existing.product_type_id == link.product_type_id:
            raise HTTPException(status_code=400, detail="Link already exists")
        else:
            raise HTTPException(status_code=400, detail="An Amazon template is already assigned to this equipment type. Only one template is allowed.")
    
    new_link = EquipmentTypeProductType(
        equipment_type_id=link.equipment_type_id,
        product_type_id=link.product_type_id
    )
    db.add(new_link)
    db.commit()
    db.refresh(new_link)
    return new_link

@router.get("/equipment-type-links", response_model=List[EquipmentTypeProductTypeLinkResponse])
def list_equipment_type_links(db: Session = Depends(get_db)):
    # Step 1: Filter at DB level
    links = db.query(EquipmentTypeProductType).filter(
        EquipmentTypeProductType.equipment_type_id.isnot(None)
    ).all()
    
    # Step 2: Defensive fallback (extra safety)
    valid_links = [l for l in links if l.equipment_type_id is not None]
    return valid_links

@router.get("/equipment-type/{equipment_type_id}/product-type", response_model=AmazonProductTypeResponse | None)
def get_product_type_for_equipment_type(equipment_type_id: int, db: Session = Depends(get_db)):
    link = db.query(EquipmentTypeProductType).filter(
        EquipmentTypeProductType.equipment_type_id == equipment_type_id
    ).first()
    if not link:
        return None
    return db.query(AmazonProductType).filter(AmazonProductType.id == link.product_type_id).first()

@router.delete("/equipment-type-links/{link_id}")
def delete_equipment_type_link(link_id: int, db: Session = Depends(get_db)):
    link = db.query(EquipmentTypeProductType).filter(EquipmentTypeProductType.id == link_id).first()
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    db.delete(link)
    db.commit()
    return {"message": "Link deleted"}

@router.get("/product-types/{product_code}/download")
def download_product_type_template(product_code: str, db: Session = Depends(get_db)):
    product_type = db.query(AmazonProductType).filter(AmazonProductType.code == product_code).first()
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found")

    # These columns are introduced by Option A (Phase 1)
    if not getattr(product_type, "file_path", None):
        raise HTTPException(status_code=404, detail="No template file stored for this product type")

    if not os.path.exists(product_type.file_path):
        raise HTTPException(status_code=404, detail="File missing on disk")

    filename = getattr(product_type, "original_filename", None) or f"{product_code}.xlsx"
    return FileResponse(product_type.file_path, filename=filename)


@router.get("/product-types/{product_code}/preview", response_model=AmazonProductTypeTemplatePreviewResponse)
def preview_product_type_template(product_code: str, db: Session = Depends(get_db)):
    product_type = db.query(AmazonProductType).filter(AmazonProductType.code == product_code).first()
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found")

    if not getattr(product_type, "file_path", None):
        raise HTTPException(status_code=404, detail="No template file stored for this product type")

    if not os.path.exists(product_type.file_path):
        raise HTTPException(status_code=404, detail="File missing on disk")

    MAX_PREVIEW_ROWS = 50
    MAX_PREVIEW_COLS = 50

    try:
        from openpyxl import load_workbook

        wb = load_workbook(product_type.file_path, read_only=True, data_only=True)
        
        if "Template" in wb.sheetnames:
            ws = wb["Template"]
        else:
            # Preview is allowed to fallback to first sheet for non-standard files
            ws = wb.worksheets[0]

        sheet_name = ws.title
        print(f"[PT_PREVIEW] product_code={product_code} sheet={sheet_name}")

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        row_limit = min(max_row, MAX_PREVIEW_ROWS)
        col_limit = min(max_col, MAX_PREVIEW_COLS)

        grid: List[List[str]] = []
        for r in range(1, row_limit + 1):
            row_values: List[str] = []
            for c in range(1, col_limit + 1):
                v = ws.cell(row=r, column=c).value
                row_values.append("" if v is None else str(v))
            grid.append(row_values)

        wb.close()

        filename = getattr(product_type, "original_filename", None) or f"{product_code}.xlsx"

        return AmazonProductTypeTemplatePreviewResponse(
            product_code=product_code,
            original_filename=filename,
            sheet_name=sheet_name,
            max_row=max_row,
            max_column=max_col,
            preview_row_count=row_limit,
            preview_column_count=col_limit,
            grid=grid,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to preview product type template: {str(e)}")

@router.get("/{product_code}", response_model=AmazonProductTypeResponse)
def get_product_type(product_code: str, db: Session = Depends(get_db)):
    product_type = db.query(AmazonProductType).filter(
        AmazonProductType.code == product_code
    ).first()
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found")
    return product_type

@router.get("/{product_code}/fields", response_model=List[ProductTypeFieldResponse])
def get_product_type_fields(product_code: str, db: Session = Depends(get_db)):
    product_type = db.query(AmazonProductType).filter(
        AmazonProductType.code == product_code
    ).first()
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found")
    return db.query(ProductTypeField).filter(
        ProductTypeField.product_type_id == product_type.id
    ).order_by(ProductTypeField.order_index).all()

@router.get("/{product_code}/header-rows")
def get_header_rows(product_code: str, db: Session = Depends(get_db)):
    product_type = db.query(AmazonProductType).filter(
        AmazonProductType.code == product_code
    ).first()
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found")
    return {"header_rows": product_type.header_rows or []}

@router.delete("/{product_code}")
def delete_product_type(product_code: str, db: Session = Depends(get_db)):
    product_type = db.query(AmazonProductType).filter(
        AmazonProductType.code == product_code
    ).first()
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found")
    db.delete(product_type)
    db.commit()
    return {"message": "Product type deleted"}

@router.get("/fields/{field_id}", response_model=ProductTypeFieldResponse)
def get_field(field_id: int, db: Session = Depends(get_db)):
    field = db.query(ProductTypeField).filter(ProductTypeField.id == field_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    return field

@router.patch("/fields/{field_id}", response_model=ProductTypeFieldResponse)
def update_field(field_id: int, update: ProductTypeFieldUpdate, db: Session = Depends(get_db)):
    field = db.query(ProductTypeField).filter(ProductTypeField.id == field_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    
    if update.required is not None:
        field.required = update.required
    if update.selected_value is not None:
        field.selected_value = update.selected_value if update.selected_value != "" else None
    
    db.commit()
    db.refresh(field)
    return field

@router.post("/fields/{field_id}/values", response_model=ProductTypeFieldValueResponse)
def add_field_value(field_id: int, value: ProductTypeFieldValueCreate, db: Session = Depends(get_db)):
    field = db.query(ProductTypeField).filter(ProductTypeField.id == field_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    
    existing = db.query(ProductTypeFieldValue).filter(
        ProductTypeFieldValue.product_type_field_id == field_id,
        ProductTypeFieldValue.value == value.value
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Value already exists")
    
    new_value = ProductTypeFieldValue(
        product_type_field_id=field_id,
        value=value.value
    )
    db.add(new_value)
    db.commit()
    db.refresh(new_value)
    
    # UX Improvement: If this is the only value, auto-select it
    count = db.query(ProductTypeFieldValue).filter(
        ProductTypeFieldValue.product_type_field_id == field_id
    ).count()
    
    if count == 1:
        field.selected_value = new_value.value
        db.add(field)
        db.commit()
        db.refresh(field)
        
    return new_value

@router.delete("/fields/{field_id}/values/{value_id}")
def delete_field_value(field_id: int, value_id: int, db: Session = Depends(get_db)):
    value = db.query(ProductTypeFieldValue).filter(
        ProductTypeFieldValue.id == value_id,
        ProductTypeFieldValue.product_type_field_id == field_id
    ).first()
    if not value:
        raise HTTPException(status_code=404, detail="Value not found")
        
    deleted_val_str = value.value
    db.delete(value)
    db.commit()
    
    # UX Improvement: Handle default selection
    field = db.query(ProductTypeField).filter(ProductTypeField.id == field_id).first()
    if field:
        # If deleted value was the selected default, clear it
        if field.selected_value == deleted_val_str:
            field.selected_value = None
            db.add(field) # Stage for check below
            
        remaining_values = db.query(ProductTypeFieldValue).filter(
            ProductTypeFieldValue.product_type_field_id == field_id
        ).all()
        
        # If exactly one remains, auto-select it
        if len(remaining_values) == 1:
            field.selected_value = remaining_values[0].value
            db.add(field)
            
        db.commit()
        
    return {"message": "Value deleted"}
