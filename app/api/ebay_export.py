"""
app/api/ebay_export.py

eBay CSV export (computed variations)

Current behavior:
    - Exports eBay parent/child rows from current configuration tables
    - Computes child variations in-memory (does not use ModelVariationSKU as source of truth)
    - Uses current parsed eBay template columns from EbayField
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
from datetime import datetime
from itertools import product
from typing import Dict, List, Optional, Set, Union, Tuple

import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session, selectinload

from app.database import get_db
from app.models.core import (
    Model,
    Series,
    Manufacturer,
    MarketplaceListing,
    MaterialRoleConfig,
    DesignOption,
    Color,
    MaterialColor,
    MaterialColourSurcharge,
    Material,
    EquipmentTypeDesignOption,
    MaterialRoleAssignment,
    ModelPricingSnapshot,
    ExportSetting,
    EbayStoreCategory,
    EbayStoreCategoryNode,
    EbayVariationPresetAsset,
    EquipmentType,
)
from app.models.templates import (
    EbayTemplate,
    EbayField,
    TemplateField,
    TemplateFieldAsset,
    TemplateFieldAssetEquipmentType,
    EbayFieldEquipmentTypeContent,
    EbayFieldEquipmentTypeImagePattern,
)
from app.services.template_asset_resolver import resolve_ebay_field_assets
from app.services.material_color_resolver import resolve_material_color_context
from app.services.shared_template_tokens import (
    is_price_placeholder_token,
    resolve_price_placeholders_in_value,
)


router = APIRouter(prefix="/ebay-export", tags=["eBay Export"])
logger = logging.getLogger(__name__)
STORE_CATEGORY_NORM_KEYS = {"storecategory"}
DESCRIPTION_NORM_KEY = "description"
IMAGE_NORM_KEYS = {
    "itemphotourl",
    "itemphotourls",
    "pictureurl",
    "pictureurls",
    "photourl",
    "photourls",
    "imageurl",
    "imageurls",
}
IMAGE_URL_HEADER_NAME_KEYS = {
    "item photo url",
    "item photo url 1",
    "picture url",
    "image url",
}
EBAY_DESCRIPTION_SUPPORTED_PLACEHOLDERS = (
    "[Manufacturer_Name]",
    "[Series_Name]",
    "[Model_Name]",
    "[Equipment_Type]",
    "[Equipment_Type_Name]",
    "[MANUFACTURER_NAME]",
    "[SERIES_NAME]",
    "[MODEL_NAME]",
    "[EQUIPMENT_TYPE]",
    "[EQUIPMENT_TYPE_NAME]",
    "[EQUIPMENTTYPE_NAME]",
    "[Width]",
    "[Depth]",
    "[Height]",
    "[SKU]",
    "[BASE_SKU]",
    "[Base_Sku]",
    "[base_sku]",
    "[COLOR_ABBR]",
    "[COLOR_SKU]",
    "[INDEX]",
    "[IMAGE_INDEX]",
)
EBAY_DESCRIPTION_MIN_REQUIRED_TOKEN_GROUPS = (
    ("[Manufacturer_Name]", "[MANUFACTURER_NAME]"),
    ("[Series_Name]", "[SERIES_NAME]"),
    ("[Model_Name]", "[MODEL_NAME]"),
    ("[Equipment_Type_Name]", "[Equipment_Type]", "[EQUIPMENT_TYPE]", "[EQUIPMENT_TYPE_NAME]", "[EQUIPMENTTYPE_NAME]"),
)
EBAY_DESCRIPTION_TOKEN_PATTERN = re.compile(r"\[[^\]]+\]")
DESCRIPTION_URL_PATTERN = re.compile(r"https?://[^\s\"'<>]+", flags=re.IGNORECASE)
EBAY_DESCRIPTION_SELECTION_MODE_GLOBAL_PRIMARY = "GLOBAL_PRIMARY"
EBAY_DESCRIPTION_SELECTION_MODE_EQUIPMENT_TYPE_PRIMARY = "EQUIPMENT_TYPE_PRIMARY"
EBAY_DESCRIPTION_SELECTION_MODE_DEFAULT = EBAY_DESCRIPTION_SELECTION_MODE_GLOBAL_PRIMARY
EBAY_DESCRIPTION_SELECTION_MODES = {
    EBAY_DESCRIPTION_SELECTION_MODE_GLOBAL_PRIMARY,
    EBAY_DESCRIPTION_SELECTION_MODE_EQUIPMENT_TYPE_PRIMARY,
}


class EbayExportRequest(BaseModel):
    model_ids: List[int]
    # Currently ignored: eBay export is always computed from configuration tables.
    use_generated_variation_skus: bool = True
    use_variation_presets: bool = False

    # Selection-driven fields (eBay only)
    export_mode: Union[str, None] = "data_driven"  # "data_driven" or "selection_driven"

    # Selection criteria (optional filters)
    role_keys: List[str] = []
    color_surcharge_ids: List[int] = []
    design_option_ids: List[int] = []
    # FIX for Phase 4 Chunk 6-B: explicitly allow str/bool/None for with_padding
    with_padding: Union[str, bool, None] = "both"  # "both", "with_padding", "no_padding", True, False


def _normalize_key(s: str) -> str:
    # Normalize: strip, lower, handle non-breaking spaces, AND underscores->spaces
    return (s or "").replace("\u00A0", " ").replace("_", " ").strip().lower()


def _normalize_for_url(name: str) -> str:
    if not name:
        return ""
    return re.sub(r"[^a-zA-Z0-9]", "", name)


def _build_ebay_filename_tokens(export_models: List[Model], db: Session) -> Tuple[str, str, str]:
    """
    Build stable filename tokens for eBay export artifacts.
    Returns: (manufacturer_token, series_token, date_token)
    """
    if not export_models:
        return "Unknown", "Unknown", datetime.now().strftime("%Y-%m-%d")

    first_model = export_models[0]
    first_series = db.query(Series).filter(Series.id == first_model.series_id).first()
    first_manufacturer = (
        db.query(Manufacturer).filter(Manufacturer.id == first_series.manufacturer_id).first()
        if first_series
        else None
    )

    mfr_name = _normalize_for_url(first_manufacturer.name) if first_manufacturer else "Unknown"
    series_ids = {m.series_id for m in export_models}
    if len(series_ids) > 1:
        series_name = "Multiple_Series"
    else:
        series_name = _normalize_for_url(first_series.name) if first_series else "Unknown"

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    return mfr_name, series_name, date_str


def _normalize_image_token(value: str) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def _normalize_color_name_for_match(value: Optional[str]) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _normalize_color_sku_for_match(value: Optional[str]) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).lower()


def _color_identity_key(internal_name: Optional[str], friendly_name: Optional[str], sku_abbrev: Optional[str]) -> Tuple[str, str, str]:
    return (
        _normalize_color_name_for_match(internal_name),
        _normalize_color_name_for_match(friendly_name),
        _normalize_color_sku_for_match(sku_abbrev),
    )


def _validate_color_label_consistency_for_children(
    export_children: List[Dict[str, object]],
    parent_color_values: Set[str],
) -> None:
    child_color_values = {
        str(item.get("attrs", {}).get("Color", "") or "").strip()
        for item in export_children
        if str(item.get("attrs", {}).get("Color", "") or "").strip()
    }
    picture_color_values = {
        str(item.get("color_label", "") or "").strip()
        for item in export_children
        if str(item.get("color_label", "") or "").strip()
    }
    parent_color_values_clean = {str(v or "").strip() for v in parent_color_values if str(v or "").strip()}

    if child_color_values != picture_color_values:
        raise HTTPException(
            status_code=400,
            detail=(
                "Color label mismatch for eBay export: child VariationSpecifics labels "
                f"{sorted(child_color_values)} differ from picture color labels {sorted(picture_color_values)}."
            ),
        )
    if parent_color_values_clean != child_color_values:
        raise HTTPException(
            status_code=400,
            detail=(
                "Color label mismatch for eBay export: parent VariationSpecificsSet color values "
                f"{sorted(parent_color_values_clean)} differ from child VariationSpecifics color values "
                f"{sorted(child_color_values)}."
            ),
        )


def _build_image_token_context(token_context: Dict[str, str]) -> Dict[str, str]:
    image_token_keys = {
        "Manufacturer_Name",
        "Series_Name",
        "Model_Name",
        "Equipment_Type",
        "Equipment_Type_Name",
        "EquipmentType_Name",
        "MANUFACTURER_NAME",
        "SERIES_NAME",
        "MODEL_NAME",
        "EQUIPMENT_TYPE",
        "EQUIPMENT_TYPE_NAME",
        "EQUIPMENTTYPE_NAME",
    }
    out = dict(token_context or {})
    for key in image_token_keys:
        if key in out:
            out[key] = _normalize_image_token(out.get(key, ""))
    return out


def _normalize_image_filename_in_url(value: str) -> str:
    text = str(value or "")
    if not text:
        return text

    split_idx = len(text)
    q_idx = text.find("?")
    h_idx = text.find("#")
    if q_idx >= 0:
        split_idx = min(split_idx, q_idx)
    if h_idx >= 0:
        split_idx = min(split_idx, h_idx)

    core = text[:split_idx]
    suffix = text[split_idx:] if split_idx < len(text) else ""
    last_sep = max(core.rfind("/"), core.rfind("\\"))
    if last_sep < 0:
        return _normalize_image_token(core) + suffix
    return core[: last_sep + 1] + _normalize_image_token(core[last_sep + 1 :]) + suffix


def _extract_image_render_config(raw_pattern: str) -> tuple[str, str, str]:
    text = str(raw_pattern or "")
    value_separator = "|"
    variation_prefix_pattern = ""

    while True:
        sep_match = re.match(r"^\[\[SEP:(.)\]\](.*)$", text, flags=re.S)
        if sep_match:
            value_separator = sep_match.group(1) or "|"
            text = sep_match.group(2) or ""
            continue
        prefix_match = re.match(r"^\[\[VPFX:(.*?)\]\](.*)$", text, flags=re.S)
        if prefix_match:
            variation_prefix_pattern = prefix_match.group(1) or ""
            text = prefix_match.group(2) or ""
            continue
        break

    return value_separator, variation_prefix_pattern, text


def _resolve_allowed_padding_set(raw_with_padding: Union[str, bool, None]) -> Set[bool]:
    allowed_padding_set: Set[bool] = {True, False}
    if isinstance(raw_with_padding, bool):
        return {raw_with_padding}
    if isinstance(raw_with_padding, str):
        s = raw_with_padding.lower().strip()
        if s == "both":
            return {True, False}
        if s in ("with_padding", "with", "true"):
            return {True}
        if s in ("no_padding", "without", "false"):
            return {False}
        return {True, False}
    if raw_with_padding is None:
        return {True, False}
    return allowed_padding_set


def _load_ebay_variation_presets(db: Session) -> List[EbayVariationPresetAsset]:
    return (
        db.query(EbayVariationPresetAsset)
        .filter(EbayVariationPresetAsset.marketplace == "EBAY")
        .all()
    )


def _select_preset_for_equipment_type(
    presets: List[EbayVariationPresetAsset],
    equipment_type_id: Optional[int],
    equipment_type_name: str,
    model_ids: List[int],
) -> EbayVariationPresetAsset:
    if equipment_type_id is None:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "variation_preset_missing_for_equipment_type",
                "message": "No eBay variation preset assigned for equipment type",
                "equipment_type_id": None,
                "equipment_type_name": equipment_type_name,
                "model_ids": model_ids,
            },
        )
    matches: List[EbayVariationPresetAsset] = []
    for preset in presets:
        preset_equipment_type_ids = list(getattr(preset, "equipment_type_ids", None) or [])
        if int(equipment_type_id) in [int(x) for x in preset_equipment_type_ids]:
            matches.append(preset)
    if len(matches) == 1:
        return matches[0]
    if len(matches) == 0:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "variation_preset_missing_for_equipment_type",
                "message": "No eBay variation preset assigned for equipment type",
                "equipment_type_id": int(equipment_type_id),
                "equipment_type_name": equipment_type_name,
                "model_ids": model_ids,
            },
        )
    raise HTTPException(
        status_code=400,
        detail={
            "code": "variation_preset_ambiguous_for_equipment_type",
            "message": "Multiple eBay variation presets assigned for equipment type",
            "equipment_type_id": int(equipment_type_id),
            "equipment_type_name": equipment_type_name,
            "preset_ids": sorted([int(m.id) for m in matches]),
            "model_ids": model_ids,
        },
    )


def _map_preset_with_padding_to_export_value(with_padding: str) -> str:
    value = str(with_padding or "").strip().lower()
    if value == "non_padded":
        return "no_padding"
    if value == "padded":
        return "with_padding"
    if value == "both":
        return "both"
    raise ValueError("Invalid preset with_padding value")


def _load_current_template_columns(db: Session) -> List[TemplateField]:
    """Load canonical eBay template fields ordered deterministically."""
    fields: List[TemplateField] = (
        db.query(TemplateField)
        .filter(TemplateField.marketplace == "ebay")
        .order_by(func.coalesce(TemplateField.order_index, 10**9), TemplateField.id)
        .all()
    )
    if not fields:
        raise HTTPException(
            status_code=400,
            detail="Current eBay template has no canonical fields. Parse the template first.",
        )

    return fields


def _template_default_value(field: TemplateField) -> str:
    """Return default value for a template column (custom_value wins, then selected_value)."""
    if field.custom_value is not None and str(field.custom_value).strip() != "":
        return str(field.custom_value)
    if field.selected_value is not None and str(field.selected_value).strip() != "":
        return str(field.selected_value)
    return ""


def _resolve_parent_value(field: TemplateField) -> str:
    if getattr(field, "parent_custom_value", None) is not None and str(getattr(field, "parent_custom_value", "")).strip() != "":
        return str(getattr(field, "parent_custom_value"))
    if getattr(field, "parent_selected_value", None) is not None and str(getattr(field, "parent_selected_value", "")).strip() != "":
        return str(getattr(field, "parent_selected_value"))
    if field.custom_value is not None and str(field.custom_value).strip() != "":
        return str(field.custom_value)
    if field.selected_value is not None and str(field.selected_value).strip() != "":
        return str(field.selected_value)
    if getattr(field, "parsed_default_value", None) is not None and str(getattr(field, "parsed_default_value", "")).strip() != "":
        return str(getattr(field, "parsed_default_value"))
    return ""


def _resolve_variation_value(field: TemplateField) -> str:
    if getattr(field, "variation_custom_value", None) is not None and str(getattr(field, "variation_custom_value", "")).strip() != "":
        return str(getattr(field, "variation_custom_value"))
    if getattr(field, "variation_selected_value", None) is not None and str(getattr(field, "variation_selected_value", "")).strip() != "":
        return str(getattr(field, "variation_selected_value"))
    if field.custom_value is not None and str(field.custom_value).strip() != "":
        return str(field.custom_value)
    if field.selected_value is not None and str(field.selected_value).strip() != "":
        return str(field.selected_value)
    if getattr(field, "parsed_default_value", None) is not None and str(getattr(field, "parsed_default_value", "")).strip() != "":
        return str(getattr(field, "parsed_default_value"))
    return ""


def _build_ebay_token_context(model: Model) -> Dict[str, str]:
    series = getattr(model, "series", None)
    manufacturer = getattr(series, "manufacturer", None) if series is not None else None
    equipment_type = getattr(model, "equipment_type", None)
    model_sku = str(getattr(model, "sku_override", "") or getattr(model, "parent_sku", "") or "")
    model_base_sku = str(getattr(model, "parent_sku", "") or "")
    width_val = getattr(model, "width", None)
    depth_val = getattr(model, "depth", None)
    height_val = getattr(model, "height", None)
    width_str = "" if width_val is None else str(width_val)
    depth_str = "" if depth_val is None else str(depth_val)
    height_str = "" if height_val is None else str(height_val)
    model_name = str(getattr(model, "name", "") or "")
    series_name = str(getattr(series, "name", "") or "")
    manufacturer_name = str(getattr(manufacturer, "name", "") or "")
    equipment_type_name = str(getattr(equipment_type, "name", "") or "")

    return {
        "Model_Name": model_name,
        "Series_Name": series_name,
        "Manufacturer_Name": manufacturer_name,
        "Equipment_Type": equipment_type_name,
        "Equipment_Type_Name": equipment_type_name,
        "EquipmentType_Name": equipment_type_name,
        "MODEL_NAME": model_name,
        "SERIES_NAME": series_name,
        "MANUFACTURER_NAME": manufacturer_name,
        "EQUIPMENT_TYPE": equipment_type_name,
        "EQUIPMENT_TYPE_NAME": equipment_type_name,
        "EQUIPMENTTYPE_NAME": equipment_type_name,
        "Width": width_str,
        "Depth": depth_str,
        "Height": height_str,
        "Model_Width": width_str,
        "Model_Depth": depth_str,
        "Model_Height": height_str,
        "SKU": model_sku,
        "Parent_SKU": model_sku,
        "BASE_SKU": model_base_sku,
        "Base_Sku": model_base_sku,
        "base_sku": model_base_sku,
    }


def _render_ebay_template_value(
    value: str,
    token_context: Dict[str, str],
    *,
    db: Optional[Session] = None,
    model_id: Optional[int] = None,
    field_name: str = "ebay_template_value",
) -> str:
    rendered = str(value or "")
    for token_name, token_value in token_context.items():
        rendered = rendered.replace(f"[{token_name}]", token_value)
    if db is not None and model_id is not None:
        rendered = resolve_price_placeholders_in_value(
            value=rendered,
            db=db,
            model_id=int(model_id),
            field_name=field_name,
        )
    return rendered


def _extract_bracket_tokens(value: str) -> List[str]:
    return EBAY_DESCRIPTION_TOKEN_PATTERN.findall(str(value or ""))


def _validate_description_html_tokens_for_model(
    model: Model,
    description_html: str,
    token_context: Dict[str, str],
    field_name: str = "Description",
) -> None:
    tokens = _extract_bracket_tokens(description_html)
    token_set = set(tokens)
    supported_set = set(EBAY_DESCRIPTION_SUPPORTED_PLACEHOLDERS)

    missing_minimum_tokens: List[str] = []
    for token_group in EBAY_DESCRIPTION_MIN_REQUIRED_TOKEN_GROUPS:
        if not any(token in token_set for token in token_group):
            missing_minimum_tokens.append(token_group[0])
    unknown_tokens = sorted(
        token
        for token in token_set
        if token not in supported_set and not is_price_placeholder_token(token)
    )

    # Color/index description placeholders are optional and may only be available for child rows.
    non_empty_enforced_tokens = [
        token for token in EBAY_DESCRIPTION_SUPPORTED_PLACEHOLDERS
        if token not in {"[COLOR_ABBR]", "[COLOR_SKU]", "[INDEX]", "[IMAGE_INDEX]"}
    ]
    missing_or_empty_token_values: List[str] = []
    for token in non_empty_enforced_tokens:
        if token not in token_set:
            continue
        token_name = token[1:-1]
        token_value = token_context.get(token_name, "")
        if str(token_value or "").strip() == "":
            missing_or_empty_token_values.append(token)

    if missing_minimum_tokens or unknown_tokens or missing_or_empty_token_values:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "description_token_validation_failed",
                "message": "Description HTML token validation failed",
                "model_id": int(getattr(model, "id", 0) or 0),
                "model_name": str(getattr(model, "name", "") or ""),
                "field_name": field_name,
                "missing_minimum_tokens": missing_minimum_tokens,
                "unknown_tokens": unknown_tokens,
                "missing_or_empty_token_values": missing_or_empty_token_values,
            },
        )


def _render_description_html_value(
    description_html: str,
    token_context: Dict[str, str],
    color_abbr: Optional[str] = None,
    image_index: int = 1,
    db: Optional[Session] = None,
    model_id: Optional[int] = None,
    field_name: str = "Description",
) -> str:
    image_token_context = _build_image_token_context(token_context)
    normalized_color = _normalize_image_token(color_abbr or "")
    try:
        normalized_index = int(image_index)
    except (TypeError, ValueError):
        normalized_index = 1
    normalized_index = max(1, min(12, normalized_index))
    index_text = f"{normalized_index:03d}"

    url_token_context = dict(image_token_context)
    url_token_context.update(
        {
            "COLOR_ABBR": normalized_color,
            "COLOR_SKU": normalized_color,
            "INDEX": index_text,
            "IMAGE_INDEX": index_text,
        }
    )

    raw_html = str(description_html or "")

    def _replace_url(match: re.Match) -> str:
        raw_url = str(match.group(0) or "")
        rendered_url = _render_ebay_template_value(
            raw_url,
            url_token_context,
            db=db,
            model_id=model_id,
            field_name=field_name,
        )
        return _normalize_image_filename_in_url(rendered_url)

    rendered_html = DESCRIPTION_URL_PATTERN.sub(_replace_url, raw_html)
    text_token_context = dict(token_context or {})
    if "Manufacturer_Name" in text_token_context:
        text_token_context["MANUFACTURER_NAME"] = text_token_context.get("Manufacturer_Name", "")
    if "Series_Name" in text_token_context:
        text_token_context["SERIES_NAME"] = text_token_context.get("Series_Name", "")
    if "Model_Name" in text_token_context:
        text_token_context["MODEL_NAME"] = text_token_context.get("Model_Name", "")
    equipment_type_text = (
        text_token_context.get("Equipment_Type_Name", "")
        or text_token_context.get("EquipmentType_Name", "")
    )
    if equipment_type_text:
        text_token_context["EQUIPMENT_TYPE"] = equipment_type_text
        text_token_context["EQUIPMENT_TYPE_NAME"] = equipment_type_text
        text_token_context["EQUIPMENTTYPE_NAME"] = equipment_type_text
    rendered_html = _render_ebay_template_value(
        rendered_html,
        text_token_context,
        db=db,
        model_id=model_id,
        field_name=field_name,
    )
    return rendered_html


def _normalize_description_selection_mode(value: Optional[str]) -> str:
    text = str(value or "").strip().upper()
    if text in EBAY_DESCRIPTION_SELECTION_MODES:
        return text
    return EBAY_DESCRIPTION_SELECTION_MODE_DEFAULT


def _is_empty_value(v) -> bool:
    return v is None or str(v).strip() == ""


def _compact_key(s: str) -> str:
    return "".join(str(s or "").split()).lower()


def _normalize_header_name_key(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _header_key_matches(header_key: str, target: str) -> bool:
    norm = _normalize_key(header_key)
    target_norm = _normalize_key(target)
    compact_norm = norm.replace(" ", "")
    compact_target = target_norm.replace(" ", "")
    return (
        norm == target_norm
        or norm.startswith(f"{target_norm}(")
        or compact_norm == compact_target
        or compact_norm.startswith(f"{compact_target}(")
    )


def _normalize_sheet_name(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s or "").strip().lower())


def _pick_template_sheet_name(sheet_names: List[str]) -> Optional[str]:
    target = _normalize_sheet_name("Template")
    for sheet_name in sheet_names:
        if _normalize_sheet_name(sheet_name) == target:
            return sheet_name
    return sheet_names[0] if sheet_names else None


def _detect_header_row_index(ws) -> int:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row < 1 or max_col < 1:
        return 1

    required_norms = {"*action", "custom label (sku)"}
    best_row = 1
    best_score = -1
    scan_limit = min(max_row, 80)
    for r in range(1, scan_limit + 1):
        non_empty = 0
        row_tokens: Set[str] = set()
        for c in range(1, max_col + 1):
            raw = ws.cell(row=r, column=c).value
            text = str(raw).strip() if raw is not None else ""
            if not text:
                continue
            non_empty += 1
            row_tokens.add(_normalize_key(text))

        if non_empty < 3:
            continue
        required_hits = sum(1 for token in required_norms if token in row_tokens)
        score = (required_hits * 1000) + non_empty
        if score > best_score:
            best_score = score
            best_row = r

    return best_row


def _coalesce_row_scope(value: Optional[str]) -> str:
    if value in ("parent_only", "variation_only"):
        return value
    return "both"


def _row_scope_allows(row_type: str, row_scope_value: Optional[str]) -> bool:
    scope = _coalesce_row_scope(row_scope_value)
    if row_type == "parent":
        return scope in ("both", "parent_only")
    return scope in ("both", "variation_only")


def _build_node_binding_map(node: EbayStoreCategoryNode) -> Dict[str, Set[int]]:
    by_type: Dict[str, Set[int]] = {}
    for binding in (getattr(node, "bindings", None) or []):
        binding_type = str(getattr(binding, "binding_type", "") or "").strip().lower()
        if not binding_type:
            continue
        binding_id = getattr(binding, "binding_id", None)
        if binding_id is None:
            continue
        by_type.setdefault(binding_type, set()).add(int(binding_id))
    return by_type


def _node_matches_model(
    node: EbayStoreCategoryNode,
    binding_map: Dict[str, Set[int]],
    model: Model,
    manufacturer_id: Optional[int],
) -> bool:
    binding_type = str(getattr(node, "binding_type", "") or "").strip().lower()
    if binding_type in ("none", "custom"):
        return True
    if binding_type == "equipment_type":
        return model.equipment_type_id is not None and model.equipment_type_id in binding_map.get("equipment_type", set())
    if binding_type == "manufacturer":
        return manufacturer_id is not None and manufacturer_id in binding_map.get("manufacturer", set())
    if binding_type == "series":
        return model.series_id is not None and model.series_id in binding_map.get("series", set())
    if binding_type == "model":
        return model.id is not None and model.id in binding_map.get("model", set())
    return False


def _resolve_store_category_number_from_nodes(
    model: Model,
    nodes_by_level: Dict[str, List[EbayStoreCategoryNode]],
    node_binding_maps: Dict[int, Dict[str, Set[int]]],
) -> Optional[int]:
    series_obj = getattr(model, "series", None)
    manufacturer_id = getattr(series_obj, "manufacturer_id", None)
    for level in ("third", "second", "top"):
        matches = [
            node
            for node in nodes_by_level.get(level, [])
            if _node_matches_model(node, node_binding_maps.get(node.id, {}), model, manufacturer_id)
        ]
        if matches:
            chosen = sorted(matches, key=lambda n: int(getattr(n, "id", 0) or 0))[0]
            return getattr(chosen, "store_category_number", None)
    return None


def _resolve_store_category_number_from_legacy_map(
    db: Session,
    model: Model,
    store_category_default_level: str,
) -> Optional[int]:
    store_category_query = (
        db.query(EbayStoreCategory)
        .filter(
            EbayStoreCategory.system == "ebay",
            EbayStoreCategory.is_enabled == True,  # noqa: E712
            EbayStoreCategory.equipment_type_id == model.equipment_type_id,
        )
    )
    series_obj = getattr(model, "series", None)
    if store_category_default_level == "series":
        store_category_row = (
            store_category_query
            .filter(
                EbayStoreCategory.level == "series",
                EbayStoreCategory.series_id == model.series_id,
            )
            .first()
        )
    elif store_category_default_level == "manufacturer":
        manufacturer_id = getattr(series_obj, "manufacturer_id", None)
        store_category_row = (
            store_category_query
            .filter(
                EbayStoreCategory.level == "manufacturer",
                EbayStoreCategory.manufacturer_id == manufacturer_id,
            )
            .first()
        ) if manufacturer_id is not None else None
    else:
        store_category_row = (
            store_category_query
            .filter(EbayStoreCategory.level == "equipment_type")
            .first()
        )
    return getattr(store_category_row, "store_category_number", None) if store_category_row else None


def _pick_description_html_for_equipment_type(
    content_map: Dict[Optional[int], str],
    equipment_type_id: Optional[int],
) -> Optional[str]:
    if equipment_type_id in content_map and str(content_map[equipment_type_id]).strip() != "":
        return content_map[equipment_type_id]
    fallback = content_map.get(None)
    if fallback is not None and str(fallback).strip() != "":
        return fallback
    return None


def _canonical_description_source_map_for_template_field(
    db: Session,
    template_field_id: int,
    equipment_type_id: Optional[int],
) -> Dict[Optional[int], str]:
    out: Dict[Optional[int], str] = {}
    fallback_row = (
        db.query(TemplateFieldAsset)
        .filter(
            TemplateFieldAsset.template_field_id == template_field_id,
            TemplateFieldAsset.asset_type == "description_html",
            TemplateFieldAsset.is_default_fallback.is_(True),
        )
        .order_by(TemplateFieldAsset.id.desc())
        .first()
    )
    if fallback_row is not None and str(getattr(fallback_row, "value", "") or "").strip() != "":
        out[None] = str(getattr(fallback_row, "value", "") or "")

    if equipment_type_id is not None:
        exact_row = (
            db.query(TemplateFieldAsset)
            .join(
                TemplateFieldAssetEquipmentType,
                TemplateFieldAssetEquipmentType.asset_id == TemplateFieldAsset.id,
            )
            .filter(
                TemplateFieldAsset.template_field_id == template_field_id,
                TemplateFieldAsset.asset_type == "description_html",
                TemplateFieldAssetEquipmentType.equipment_type_id == equipment_type_id,
            )
            .order_by(TemplateFieldAsset.id.desc())
            .first()
        )
        if exact_row is not None and str(getattr(exact_row, "value", "") or "").strip() != "":
            out[equipment_type_id] = str(getattr(exact_row, "value", "") or "")

    return out


def _resolve_required_description_html_for_model(
    content_map: Dict[Optional[int], str],
    model: Model,
    selection_mode: str,
    field_name: str = "Description",
) -> str:
    equipment_type_id = getattr(model, "equipment_type_id", None)
    equipment_type_name = str(getattr(getattr(model, "equipment_type", None), "name", "") or "")
    normalized_mode = _normalize_description_selection_mode(selection_mode)

    specific = content_map.get(equipment_type_id) if equipment_type_id is not None else None
    global_value = content_map.get(None)
    specific_text = str(specific or "")
    global_text = str(global_value or "")

    if normalized_mode == EBAY_DESCRIPTION_SELECTION_MODE_GLOBAL_PRIMARY:
        if global_text.strip():
            return global_text
        if specific_text.strip():
            return specific_text
    else:
        if specific_text.strip():
            return specific_text
        if global_text.strip():
            return global_text

    raise HTTPException(
        status_code=400,
        detail={
            "code": "description_asset_missing_for_equipment_type",
            "message": "Description HTML asset missing for equipment type",
            "model_id": model.id,
            "model_name": model.name,
            "equipment_type_id": equipment_type_id,
            "equipment_type_name": equipment_type_name,
            "field_name": field_name,
        },
    )

def _slice_to_version_prefix(base_sku: str) -> Optional[str]:
    """
    Take the model’s parent SKU, trim it to the version marker only (...V1, ...V2),
    dropping any trailing zeros or other characters after the numeric version.
    """
    if not base_sku:
        return None

    matches = list(re.finditer(r"V(\d+)", base_sku, flags=re.IGNORECASE))
    if not matches:
        return base_sku + "-"  # Fallback if no version marker, assume we append

    last_match = matches[-1]

    digits = last_match.group(1)
    # User rule: V10000000 -> V1. Strip trailing zeros.
    stripped = digits.rstrip("0")
    if not stripped:
        stripped = "0"

    # The prefix ends right before the 'V'
    prefix_body = base_sku[: last_match.start()]

    # Reconstruct version marker
    return f"{prefix_body}V{stripped}"


def _build_design_suffix(design_ids: List[int], all_opts: Dict[int, DesignOption]) -> str:
    """Build sort-safe design suffix: keys mapped to abbrevs, sorted alphanumeric, joined."""
    found = []
    for did in design_ids:
        opt = all_opts.get(did)
        if opt and opt.sku_abbreviation:
            found.append(opt.sku_abbreviation.strip().upper())

    found.sort()
    return "".join(found)


def _get_role_rank_from_abbrev(abbrev: str) -> int:
    """Strict explicit ranking: C=0, CG=1, L=2, LG=3."""
    mapping = {"C": 0, "CG": 1, "L": 2, "LG": 3}
    return mapping.get((abbrev or "").upper().strip(), 999)


def _get_color_sort_tuple_from_code(code: str) -> Tuple[int, str]:
    s = (code or "").strip().upper()
    if s == "PBK":
        return (0, s)
    return (1, s)


@router.post("/export")
def export_ebay_csv(request: EbayExportRequest, db: Session = Depends(get_db)):
    """Export selected models to an eBay CSV using the *current* parsed eBay template."""
    export_mode_norm = (request.export_mode or "data_driven").strip().lower()
    if export_mode_norm not in ("selection_driven", "data_driven"):
        if "selection" in export_mode_norm:
            export_mode_norm = "selection_driven"
        else:
            export_mode_norm = "data_driven"
    if request.use_variation_presets:
        export_mode_norm = "selection_driven"

    wp_raw = request.with_padding
    allowed_padding_set: Set[bool] = _resolve_allowed_padding_set(wp_raw)

    logger.debug(
        "ebay export params mode=%s with_padding=%s raw=%s",
        export_mode_norm,
        allowed_padding_set,
        wp_raw,
    )

    if not request.model_ids:
        raise HTTPException(status_code=400, detail="No model_ids provided")

    models: List[Model] = db.query(Model).filter(Model.id.in_(request.model_ids)).all()
    if len(models) != len(set(request.model_ids)):
        found = {m.id for m in models}
        missing = sorted(set(request.model_ids) - found)
        raise HTTPException(status_code=404, detail=f"Models not found: {missing}")

    model_by_id: Dict[int, Model] = {m.id: m for m in models}
    model_effective_variation_inputs: Dict[int, Dict[str, Union[List[str], List[int], str, Set[bool]]]] = {}
    if request.use_variation_presets:
        presets = _load_ebay_variation_presets(db)
        equipment_type_name_by_id: Dict[int, str] = {
            int(row.id): str(row.name or "")
            for row in db.query(EquipmentType).all()
        }
        model_ids_by_equipment_type: Dict[Optional[int], List[int]] = {}
        for model in models:
            et_id = getattr(model, "equipment_type_id", None)
            model_ids_by_equipment_type.setdefault(et_id, []).append(int(model.id))
        for equipment_type_id, grouped_model_ids in model_ids_by_equipment_type.items():
            grouped_model_ids_sorted = sorted(grouped_model_ids)
            equipment_type_name = equipment_type_name_by_id.get(
                int(equipment_type_id) if equipment_type_id is not None else -1,
                "",
            )
            selected_preset = _select_preset_for_equipment_type(
                presets=presets,
                equipment_type_id=equipment_type_id,
                equipment_type_name=equipment_type_name,
                model_ids=grouped_model_ids_sorted,
            )
            payload = dict(getattr(selected_preset, "payload", None) or {})
            required_payload_keys = {"role_keys", "color_surcharge_ids", "design_option_ids", "with_padding"}
            missing_payload_keys = sorted([k for k in required_payload_keys if k not in payload])
            if missing_payload_keys:
                raise HTTPException(
                    status_code=400,
                    detail={
                        "code": "variation_preset_invalid_payload",
                        "message": "eBay variation preset payload is invalid",
                        "preset_id": int(selected_preset.id),
                        "equipment_type_id": int(equipment_type_id) if equipment_type_id is not None else None,
                        "equipment_type_name": equipment_type_name,
                        "missing_payload_keys": missing_payload_keys,
                        "model_ids": grouped_model_ids_sorted,
                    },
                )
            role_keys = list(payload.get("role_keys") or [])
            color_surcharge_ids = [int(v) for v in list(payload.get("color_surcharge_ids") or [])]
            design_option_ids = [int(v) for v in list(payload.get("design_option_ids") or [])]
            try:
                mapped_with_padding = _map_preset_with_padding_to_export_value(str(payload.get("with_padding") or ""))
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail={
                        "code": "variation_preset_invalid_payload",
                        "message": "eBay variation preset payload is invalid",
                        "preset_id": int(selected_preset.id),
                        "equipment_type_id": int(equipment_type_id) if equipment_type_id is not None else None,
                        "equipment_type_name": equipment_type_name,
                        "invalid_with_padding": payload.get("with_padding"),
                        "model_ids": grouped_model_ids_sorted,
                    },
                )
            preset_allowed_padding_set = _resolve_allowed_padding_set(mapped_with_padding)
            for grouped_model_id in grouped_model_ids_sorted:
                model_effective_variation_inputs[int(grouped_model_id)] = {
                    "role_keys": role_keys,
                    "color_surcharge_ids": color_surcharge_ids,
                    "design_option_ids": design_option_ids,
                    "with_padding_raw": mapped_with_padding,
                    "allowed_padding_set": preset_allowed_padding_set,
                }

    fields = _load_current_template_columns(db)
    fallback_headers = [f.field_name for f in fields]
    headers = list(fallback_headers)
    header_norms = [_normalize_key(h) for h in headers]

    preamble_rows: List[List[str]] = []
    latest_template = (
        db.query(EbayTemplate)
        .order_by(EbayTemplate.uploaded_at.desc(), EbayTemplate.id.desc())
        .first()
    )
    if latest_template and latest_template.file_path and os.path.exists(latest_template.file_path):
        try:
            wb = openpyxl.load_workbook(latest_template.file_path, read_only=True, data_only=True)
            template_sheet_name = _pick_template_sheet_name(wb.sheetnames)
            if template_sheet_name:
                ws = wb[template_sheet_name]
                header_row_index = _detect_header_row_index(ws)
                header_width = len(headers) if headers else (ws.max_column or 1)

                # Canonical parsed TemplateField order is the primary export header source.
                # Only fall back to workbook header detection when canonical headers are unavailable.
                if not fallback_headers:
                    max_col = ws.max_column or 1
                    header_row_cells = [
                        str(ws.cell(row=header_row_index, column=col_idx).value or "")
                        for col_idx in range(1, max_col + 1)
                    ]
                    while header_row_cells and not str(header_row_cells[-1] or "").strip():
                        header_row_cells.pop()
                    if header_row_cells:
                        headers = header_row_cells
                        header_norms = [_normalize_key(h) for h in headers]
                        header_width = len(headers)

                for pre_row_idx in range(1, header_row_index):
                    row_values = [
                        str(ws.cell(row=pre_row_idx, column=col_idx).value or "")
                        for col_idx in range(1, header_width + 1)
                    ]
                    preamble_rows.append(row_values)
            wb.close()
        except Exception as exc:
            logger.warning("Falling back to DB headers; failed to read eBay template XLSX: %s", exc)
    image_url_header_norm_keys = {
        _normalize_key(f.field_name)
        for f in fields
        if _normalize_header_name_key(f.field_name) in IMAGE_URL_HEADER_NAME_KEYS
    }
    if not image_url_header_norm_keys:
        image_url_header_norm_keys = {_normalize_key("Item photo URL")}
    parent_defaults_by_norm = {_normalize_key(f.field_name): _resolve_parent_value(f) for f in fields}
    variation_defaults_by_norm = {_normalize_key(f.field_name): _resolve_variation_value(f) for f in fields}
    required_by_norm = {_normalize_key(f.field_name): bool(f.required) for f in fields}
    parent_allowed_keys = {
        _normalize_key(f.field_name)
        for f in fields
        if _row_scope_allows("parent", getattr(f, "row_scope", None))
    }
    variation_allowed_keys = {
        _normalize_key(f.field_name)
        for f in fields
        if _row_scope_allows("variation", getattr(f, "row_scope", None))
    }
    export_settings = db.query(ExportSetting).first()
    if export_settings is None:
        store_category_default_level = "series"
        settings_parent_image_pattern = ""
        settings_variation_image_pattern = ""
        description_selection_mode = EBAY_DESCRIPTION_SELECTION_MODE_DEFAULT
    else:
        store_category_default_level = (
            (getattr(export_settings, "ebay_store_category_default_level", None) or "series").strip().lower()
        )
        settings_parent_image_pattern = str(
            getattr(export_settings, "ebay_parent_image_pattern", "") or ""
        ).strip()
        settings_variation_image_pattern = str(
            getattr(export_settings, "ebay_variation_image_pattern", "") or ""
        ).strip()
        description_selection_mode = _normalize_description_selection_mode(
            getattr(export_settings, "ebay_description_selection_mode", None)
        )
    if store_category_default_level not in ("series", "manufacturer", "equipment_type"):
        store_category_default_level = "series"
    enabled_store_category_nodes: List[EbayStoreCategoryNode] = (
        db.query(EbayStoreCategoryNode)
        .options(selectinload(EbayStoreCategoryNode.bindings))
        .filter(
            EbayStoreCategoryNode.system == "ebay",
            EbayStoreCategoryNode.is_enabled == True,  # noqa: E712
        )
        .all()
    )
    nodes_by_level: Dict[str, List[EbayStoreCategoryNode]] = {"top": [], "second": [], "third": []}
    node_binding_maps: Dict[int, Dict[str, Set[int]]] = {}
    for node in enabled_store_category_nodes:
        level = str(getattr(node, "level", "") or "").strip().lower()
        if level in nodes_by_level:
            nodes_by_level[level].append(node)
        node_binding_maps[node.id] = _build_node_binding_map(node)
    for level_key in nodes_by_level:
        nodes_by_level[level_key].sort(key=lambda n: int(getattr(n, "id", 0) or 0))

    ebay_listing_by_model_id: Dict[int, str] = {}
    listing_rows = (
        db.query(MarketplaceListing.model_id, MarketplaceListing.external_id)
        .filter(
            MarketplaceListing.model_id.in_(list(model_by_id.keys())),
            func.lower(MarketplaceListing.marketplace) == "ebay",
        )
        .all()
    )
    for model_id, external_id in listing_rows:
        if model_id is None or external_id is None:
            continue
        text = str(external_id).strip()
        if text and int(model_id) not in ebay_listing_by_model_id:
            ebay_listing_by_model_id[int(model_id)] = text

    role_configs = {_normalize_key(rc.role): rc for rc in db.query(MaterialRoleConfig).all()}
    all_design_options = {d.id: d for d in db.query(DesignOption).all()}
    variant_key_by_role_abbrev = {
        "C": "choice_no_padding",
        "CG": "choice_padded",
        "L": "premium_no_padding",
        "LG": "premium_padded",
    }

    output = io.StringIO(newline="")
    writer = csv.writer(output)
    for preamble_row in preamble_rows:
        writer.writerow(preamble_row)
    writer.writerow(headers)
    export_warnings: List[str] = []

    KEY_ACTION = "*action"
    KEY_ITEM_ID = "item id"
    KEY_SKU = "custom label (sku)"
    KEY_TITLE = "title"
    KEY_REL = "relationship"
    KEY_REL_DETAILS = "relationship details"
    KEY_PRICE = "start price"
    KEY_QTY = "quantity"
    action_header_keys = {h for h in header_norms if _header_key_matches(h, KEY_ACTION)}
    item_id_header_keys = {h for h in header_norms if _header_key_matches(h, KEY_ITEM_ID)}

    # Relationship must never inherit global defaults; only row-specific parent/variation values apply.
    for field in fields:
        if _compact_key(getattr(field, "field_name", "")) != KEY_REL:
            continue
        rel_norm_key = _normalize_key(field.field_name)

        parent_value = ""
        parent_custom = getattr(field, "parent_custom_value", None)
        if parent_custom is not None and str(parent_custom).strip() != "":
            parent_value = str(parent_custom)
        else:
            parent_selected = getattr(field, "parent_selected_value", None)
            if parent_selected is not None and str(parent_selected).strip() != "":
                parent_value = str(parent_selected)

        variation_value = ""
        variation_custom = getattr(field, "variation_custom_value", None)
        if variation_custom is not None and str(variation_custom).strip() != "":
            variation_value = str(variation_custom)
        else:
            variation_selected = getattr(field, "variation_selected_value", None)
            if variation_selected is not None and str(variation_selected).strip() != "":
                variation_value = str(variation_selected)

        parent_defaults_by_norm[rel_norm_key] = parent_value
        variation_defaults_by_norm[rel_norm_key] = variation_value

    def _build_image_urls_from_pattern(
        raw_pattern: str,
        color_sku: Optional[str] = None,
        color_friendly_name: Optional[str] = None,
    ) -> str:
        value_separator, variation_prefix_pattern, pattern = _extract_image_render_config(raw_pattern)
        normalized_color = _normalize_image_token(color_sku or "")
        color_name = str(color_friendly_name or "").strip()
        variation_prefix = variation_prefix_pattern
        variation_prefix = variation_prefix.replace("[COLOR_FRIENDLY_NAME]", color_name)
        variation_prefix = variation_prefix.replace("[Color Friendly Name]", color_name)
        variation_prefix = variation_prefix.replace("[COLOR FRIENDLY NAME]", color_name)
        variation_prefix = variation_prefix.replace("[COLOR_ABBR]", normalized_color)
        variation_prefix = variation_prefix.replace("[COLOR_SKU]", normalized_color)
        pattern = pattern.replace("[COLOR_ABBR]", normalized_color)
        pattern = pattern.replace("[COLOR_SKU]", normalized_color)
        if "[INDEX]" in pattern or "[IMAGE_INDEX]" in pattern:
            out_urls: List[str] = []
            for idx in range(1, 13):
                idx_text = f"{idx:03d}"
                rendered_url = pattern.replace("[IMAGE_INDEX]", idx_text).replace("[INDEX]", idx_text)
                out_urls.append(_normalize_image_filename_in_url(rendered_url))
            return f"{variation_prefix}{value_separator.join(out_urls)}"
        return f"{variation_prefix}{_normalize_image_filename_in_url(pattern)}"

    def _render_settings_image_pattern(
        pattern: str,
        token_context: Dict[str, str],
        color_abbr: Optional[str] = None,
        image_index: int = 1,
        model_id: Optional[int] = None,
        field_name: str = "Image URL",
    ) -> str:
        rendered = _render_ebay_template_value(
            str(pattern or ""),
            _build_image_token_context(token_context),
            db=db,
            model_id=model_id,
            field_name=field_name,
        )
        normalized_color_abbr = _normalize_image_token(color_abbr or "")
        rendered = rendered.replace("[COLOR_ABBR]", normalized_color_abbr)
        rendered = rendered.replace("[COLOR_SKU]", normalized_color_abbr)
        try:
            normalized_index = int(image_index)
        except (TypeError, ValueError):
            normalized_index = 1
        normalized_index = max(1, min(12, normalized_index))
        index_text = f"{normalized_index:03d}"
        rendered = rendered.replace("[INDEX]", index_text)
        rendered = rendered.replace("[IMAGE_INDEX]", index_text)
        return _normalize_image_filename_in_url(rendered)

    def _legacy_description_source_map_for_template_field(template_field_id: int) -> Dict[Optional[int], str]:
        row_ids = [
            int(r[0])
            for r in (
                db.query(EbayField.id)
                .filter(EbayField.template_field_id == template_field_id)
                .all()
            )
        ]
        if not row_ids:
            return {}
        out: Dict[Optional[int], str] = {}
        legacy_rows: List[EbayFieldEquipmentTypeContent] = (
            db.query(EbayFieldEquipmentTypeContent)
            .filter(EbayFieldEquipmentTypeContent.ebay_field_id.in_(row_ids))
            .order_by(
                EbayFieldEquipmentTypeContent.is_default_fallback.desc(),
                func.coalesce(EbayFieldEquipmentTypeContent.equipment_type_id, -1).asc(),
                EbayFieldEquipmentTypeContent.id.asc(),
            )
            .all()
        )
        for row in legacy_rows:
            key = int(row.equipment_type_id) if row.equipment_type_id is not None else None
            if key not in out and str(row.html_value or "").strip():
                out[key] = str(row.html_value or "")
        return out

    def _legacy_image_source_map_for_template_field(template_field_id: int) -> Dict[Optional[int], Dict[str, str]]:
        row_ids = [
            int(r[0])
            for r in (
                db.query(EbayField.id)
                .filter(EbayField.template_field_id == template_field_id)
                .all()
            )
        ]
        if not row_ids:
            return {}
        out: Dict[Optional[int], Dict[str, str]] = {}
        legacy_rows: List[EbayFieldEquipmentTypeImagePattern] = (
            db.query(EbayFieldEquipmentTypeImagePattern)
            .filter(EbayFieldEquipmentTypeImagePattern.ebay_field_id.in_(row_ids))
            .order_by(
                func.coalesce(EbayFieldEquipmentTypeImagePattern.equipment_type_id, -1).asc(),
                EbayFieldEquipmentTypeImagePattern.id.asc(),
            )
            .all()
        )
        for row in legacy_rows:
            key = int(row.equipment_type_id) if row.equipment_type_id is not None else None
            if key not in out:
                out[key] = {
                    "parent_pattern": str(row.parent_pattern or ""),
                    "variation_pattern": str(row.variation_pattern or ""),
                }
        return out

    def readable_role(role_value: Optional[str]) -> str:
        raw = (role_value or "").strip()
        if not raw:
            return "Unknown Fabric"
        return raw.replace("_", " ").title()

    def fabric_label(role_config: MaterialRoleConfig, padded: bool) -> str:
        base = (
            (role_config.display_name or "").strip()
            or readable_role(role_config.role)
        )
        if not padded:
            return base
        padded_name = ((getattr(role_config, "display_name_with_padding", "") or "").strip())
        return padded_name or base

    def _sort_design_options_for_suffix(options: List[DesignOption]) -> List[DesignOption]:
        return sorted(
            options,
            key=lambda o: o.id,
        )

    def _build_option_suffix_from_opts(options: List[DesignOption]) -> str:
        return "".join(
            (opt.sku_abbreviation or "").strip().upper()
            for opt in _sort_design_options_for_suffix(options)
            if (opt.sku_abbreviation or "").strip()
        )

    def _active_material_by_role() -> Dict[str, Material]:
        now = datetime.utcnow()
        assignments: List[MaterialRoleAssignment] = (
            db.query(MaterialRoleAssignment)
            .filter(
                (MaterialRoleAssignment.end_date.is_(None)) | (MaterialRoleAssignment.end_date > now)
            )
            .order_by(
                MaterialRoleAssignment.role.asc(),
                MaterialRoleAssignment.effective_date.desc(),
                MaterialRoleAssignment.id.desc(),
            )
            .all()
        )
        out: Dict[str, Material] = {}
        for a in assignments:
            role_key = _normalize_key(a.role or "")
            if not role_key or role_key in out:
                continue
            if a.material is not None:
                out[role_key] = a.material
        return out

    def _color_sort_key_from_entry(entry: Dict[str, object]) -> Tuple[int, str]:
        return _get_color_sort_tuple_from_code(str(entry.get("color_code", "") or "").strip().upper())

    def _get_enabled_colors_for_material(material_id: int) -> List[Dict[str, object]]:
        # Canonical-first: use material_colors -> colors when present for this material.
        canonical_links: List[MaterialColor] = (
            db.query(MaterialColor)
            .join(Color, Color.id == MaterialColor.color_id)
            .options(selectinload(MaterialColor.color))
            .filter(
                MaterialColor.material_id == material_id,
                MaterialColor.ebay_variation_enabled.is_(True),
            )
            .all()
        )

        if canonical_links:
            legacy_rows_for_mapping: List[MaterialColourSurcharge] = (
                db.query(MaterialColourSurcharge)
                .filter(MaterialColourSurcharge.material_id == material_id)
                .all()
            )
            legacy_ids_by_identity: Dict[Tuple[str, str, str], List[int]] = {}
            for legacy_row in legacy_rows_for_mapping:
                legacy_key = _color_identity_key(
                    legacy_row.colour,
                    legacy_row.color_friendly_name or legacy_row.colour,
                    legacy_row.sku_abbreviation,
                )
                legacy_ids_by_identity.setdefault(legacy_key, []).append(int(legacy_row.id))

            canonical_entries: List[Dict[str, object]] = []
            for link in canonical_links:
                color_row = link.color
                if color_row is None:
                    continue
                color_code = str(color_row.sku_abbrev or "").strip().upper()
                if not color_code:
                    continue
                color_label = str(color_row.friendly_name or color_row.internal_name or "Unknown Color")
                key = _color_identity_key(
                    color_row.internal_name,
                    color_row.friendly_name,
                    color_row.sku_abbrev,
                )
                selector_ids = {int(link.id)}
                for legacy_id in legacy_ids_by_identity.get(key, []):
                    selector_ids.add(int(legacy_id))
                canonical_entries.append(
                    {
                        "id": int(link.id),
                        "selector_ids": selector_ids,
                        "source": "canonical",
                        "color_code": color_code,
                        "color_label": color_label,
                        "surcharge": float(link.surcharge or 0.0),
                        "sort_order": link.sort_order,
                    }
                )

            canonical_entries.sort(
                key=lambda entry: (
                    entry.get("sort_order") is None,
                    int(entry.get("sort_order") or 0),
                    _color_sort_key_from_entry(entry),
                )
            )
            return canonical_entries

        # Transition fallback: resolve legacy rows with canonical-first resolver when available.
        legacy_rows: List[MaterialColourSurcharge] = (
            db.query(MaterialColourSurcharge)
            .filter(
                MaterialColourSurcharge.material_id == material_id,
                MaterialColourSurcharge.ebay_variation_enabled.is_(True),
            )
            .all()
        )
        legacy_entries: List[Dict[str, object]] = []
        for legacy_row in legacy_rows:
            resolved = resolve_material_color_context(db, legacy_surcharge_id=int(legacy_row.id))
            if resolved is not None:
                color_code = str(resolved.sku_abbrev or "").strip().upper()
                color_label = str(resolved.friendly_name or resolved.internal_name or "Unknown Color")
                surcharge = float(resolved.surcharge or 0.0)
            else:
                color_code = str(legacy_row.sku_abbreviation or "").strip().upper()
                color_label = str(legacy_row.color_friendly_name or legacy_row.colour or "Unknown Color")
                surcharge = float(legacy_row.surcharge or 0.0)

            if not color_code:
                continue
            legacy_entries.append(
                {
                    "id": int(legacy_row.id),
                    "selector_ids": {int(legacy_row.id)},
                    "source": "legacy",
                    "color_code": color_code,
                    "color_label": color_label,
                    "surcharge": surcharge,
                    "sort_order": None,
                }
            )

        legacy_entries.sort(key=_color_sort_key_from_entry)
        return legacy_entries

    def _money_to_cents(value: Optional[float]) -> int:
        if value is None:
            return 0
        try:
            # MaterialColourSurcharge.surcharge is stored as dollars (Float in core model).
            return int(round(float(value) * 100))
        except (TypeError, ValueError):
            return 0

    def _color_surcharge_cents(color: Dict[str, object]) -> int:
        color_code = str(color.get("color_code", "") or "").strip().upper()
        if color_code == "PBK":
            return 0
        return _money_to_cents(float(color.get("surcharge", 0.0) or 0.0))

    def _design_options_price_cents(options: List[DesignOption]) -> int:
        total = 0
        for opt in options:
            total += int(getattr(opt, "price_cents", 0) or 0)
        return total

    def _get_base_price_cents_for_role_abbrev(model_id: int, role_abbrev: str) -> Optional[int]:
        variant_key = variant_key_by_role_abbrev.get((role_abbrev or "").strip().upper())
        if not variant_key:
            return None

        snapshots: List[ModelPricingSnapshot] = (
            db.query(ModelPricingSnapshot)
            .filter(
                ModelPricingSnapshot.model_id == model_id,
                ModelPricingSnapshot.variant_key == variant_key,
                func.lower(ModelPricingSnapshot.marketplace).in_(["ebay", "default"]),
            )
            .all()
        )
        if not snapshots:
            return None

        def _snap_priority(s: ModelPricingSnapshot) -> Tuple[int, int]:
            mp = (s.marketplace or "").strip().lower()
            if mp == "ebay":
                return (0, -int(s.id or 0))
            if mp == "default":
                return (1, -int(s.id or 0))
            return (9, -int(s.id or 0))

        chosen = sorted(snapshots, key=_snap_priority)[0]
        return int(chosen.retail_price_cents) if chosen.retail_price_cents is not None else None

    def get_child_attrs(
        role_config: MaterialRoleConfig,
        with_padding: bool,
        color: Dict[str, object],
        enabled_design_options: List[DesignOption],
        selected_design_option_ids: Set[int],
    ) -> Dict[str, str]:
        vals: Dict[str, str] = {}
        vals["Fabric"] = fabric_label(role_config, with_padding)
        vals["Color"] = str(color.get("color_label", "") or "Unknown Color")
        for opt in enabled_design_options:
            vals[opt.name] = "Yes" if opt.id in selected_design_option_ids else "No"
        return vals

    active_material_by_role = _active_material_by_role()

    for mid in sorted(model_by_id.keys()):
        model = model_by_id[mid]
        model_effective_inputs = model_effective_variation_inputs.get(mid, {})
        effective_role_keys: List[str] = list(model_effective_inputs.get("role_keys", request.role_keys) or [])
        effective_color_surcharge_ids: List[int] = list(
            model_effective_inputs.get("color_surcharge_ids", request.color_surcharge_ids) or []
        )
        effective_design_option_ids: List[int] = list(
            model_effective_inputs.get("design_option_ids", request.design_option_ids) or []
        )
        effective_with_padding_raw = model_effective_inputs.get("with_padding_raw", wp_raw)
        effective_allowed_padding_set: Set[bool] = set(
            model_effective_inputs.get("allowed_padding_set", allowed_padding_set) or set()
        )
        effective_export_mode_norm = "selection_driven" if request.use_variation_presets else export_mode_norm

        token_context = _build_ebay_token_context(model)
        image_token_context = _build_image_token_context(token_context)
        equipment_type_id = getattr(model, "equipment_type_id", None)
        model_parent_defaults_by_norm = {
            key: _render_ebay_template_value(
                value,
                token_context,
                db=db,
                model_id=model.id,
                field_name=key,
            )
            for key, value in parent_defaults_by_norm.items()
        }
        model_variation_defaults_by_norm = {
            key: _render_ebay_template_value(
                value,
                token_context,
                db=db,
                model_id=model.id,
                field_name=key,
            )
            for key, value in variation_defaults_by_norm.items()
        }
        store_category_auto_value = ""
        node_store_category_number = _resolve_store_category_number_from_nodes(
            model=model,
            nodes_by_level=nodes_by_level,
            node_binding_maps=node_binding_maps,
        )
        if node_store_category_number is not None:
            store_category_auto_value = str(node_store_category_number)
        else:
            legacy_store_category_number = _resolve_store_category_number_from_legacy_map(
                db=db,
                model=model,
                store_category_default_level=store_category_default_level,
            )
            if legacy_store_category_number is not None:
                store_category_auto_value = str(legacy_store_category_number)

        for key in list(model_parent_defaults_by_norm.keys()):
            if _compact_key(key) in STORE_CATEGORY_NORM_KEYS and _is_empty_value(model_parent_defaults_by_norm.get(key)):
                model_parent_defaults_by_norm[key] = store_category_auto_value
        for key in list(model_variation_defaults_by_norm.keys()):
            if _compact_key(key) in STORE_CATEGORY_NORM_KEYS and _is_empty_value(model_variation_defaults_by_norm.get(key)):
                model_variation_defaults_by_norm[key] = store_category_auto_value

        description_template_by_norm: Dict[str, str] = {}
        variation_image_pattern_by_norm: Dict[str, str] = {}
        for field in fields:
            if _compact_key(field.field_name) != DESCRIPTION_NORM_KEY:
                continue
            canonical_description_source_map = _canonical_description_source_map_for_template_field(
                db=db,
                template_field_id=int(field.id),
                equipment_type_id=equipment_type_id,
            )
            legacy_description_source_map = _legacy_description_source_map_for_template_field(int(field.id))
            description_source_map: Dict[Optional[int], str] = {}
            if str(canonical_description_source_map.get(None, "") or "").strip():
                description_source_map[None] = str(canonical_description_source_map.get(None, "") or "")
            elif str(legacy_description_source_map.get(None, "") or "").strip():
                description_source_map[None] = str(legacy_description_source_map.get(None, "") or "")
            if equipment_type_id is not None:
                if str(canonical_description_source_map.get(equipment_type_id, "") or "").strip():
                    description_source_map[equipment_type_id] = str(
                        canonical_description_source_map.get(equipment_type_id, "") or ""
                    )
                elif str(legacy_description_source_map.get(equipment_type_id, "") or "").strip():
                    description_source_map[equipment_type_id] = str(
                        legacy_description_source_map.get(equipment_type_id, "") or ""
                    )
            description_html = _resolve_required_description_html_for_model(
                description_source_map,
                model,
                selection_mode=description_selection_mode,
                field_name=field.field_name,
            )
            _validate_description_html_tokens_for_model(
                model=model,
                description_html=description_html,
                token_context=token_context,
                field_name=field.field_name,
            )
            description_template_by_norm[_normalize_key(field.field_name)] = description_html
            rendered_description = _render_description_html_value(
                description_html=description_html,
                token_context=token_context,
                image_index=1,
                db=db,
                model_id=model.id,
                field_name=field.field_name,
            )
            description_norm_key = _normalize_key(field.field_name)
            model_parent_defaults_by_norm[description_norm_key] = rendered_description
            model_variation_defaults_by_norm[description_norm_key] = rendered_description

        for field in fields:
            if _compact_key(field.field_name) not in IMAGE_NORM_KEYS:
                continue
            assets = resolve_ebay_field_assets(
                db=db,
                template_field_id=int(field.id),
                equipment_type_id=equipment_type_id,
            )
            has_parent_fallback = (
                db.query(TemplateFieldAsset.id)
                .filter(
                    TemplateFieldAsset.template_field_id == int(field.id),
                    TemplateFieldAsset.asset_type == "image_parent_pattern",
                    TemplateFieldAsset.is_default_fallback.is_(True),
                )
                .first()
                is not None
            )
            has_parent_exact = (
                equipment_type_id is not None and
                db.query(TemplateFieldAsset.id)
                .join(
                    TemplateFieldAssetEquipmentType,
                    TemplateFieldAssetEquipmentType.asset_id == TemplateFieldAsset.id,
                )
                .filter(
                    TemplateFieldAsset.template_field_id == int(field.id),
                    TemplateFieldAsset.asset_type == "image_parent_pattern",
                    TemplateFieldAssetEquipmentType.equipment_type_id == equipment_type_id,
                )
                .first()
                is not None
            )
            has_variation_fallback = (
                db.query(TemplateFieldAsset.id)
                .filter(
                    TemplateFieldAsset.template_field_id == int(field.id),
                    TemplateFieldAsset.asset_type == "image_variation_pattern",
                    TemplateFieldAsset.is_default_fallback.is_(True),
                )
                .first()
                is not None
            )
            has_variation_exact = (
                equipment_type_id is not None and
                db.query(TemplateFieldAsset.id)
                .join(
                    TemplateFieldAssetEquipmentType,
                    TemplateFieldAssetEquipmentType.asset_id == TemplateFieldAsset.id,
                )
                .filter(
                    TemplateFieldAsset.template_field_id == int(field.id),
                    TemplateFieldAsset.asset_type == "image_variation_pattern",
                    TemplateFieldAssetEquipmentType.equipment_type_id == equipment_type_id,
                )
                .first()
                is not None
            )
            has_parent_asset = bool(has_parent_fallback or has_parent_exact)
            has_variation_asset = bool(has_variation_fallback or has_variation_exact)

            parent_pattern = str(assets["parent_pattern"] or "") if has_parent_asset else ""
            variation_pattern = str(assets["variation_pattern"] or "") if has_variation_asset else ""
            rendered_parent_pattern = _render_ebay_template_value(
                parent_pattern,
                image_token_context,
                db=db,
                model_id=model.id,
                field_name=field.field_name,
            ) if parent_pattern else ""
            rendered_variation_pattern = _render_ebay_template_value(
                variation_pattern,
                image_token_context,
                db=db,
                model_id=model.id,
                field_name=field.field_name,
            ) if variation_pattern else ""
            image_norm_key = _normalize_key(field.field_name)
            model_parent_defaults_by_norm[image_norm_key] = (
                _build_image_urls_from_pattern(rendered_parent_pattern) if rendered_parent_pattern else ""
            )
            model_variation_defaults_by_norm[image_norm_key] = ""
            variation_image_pattern_by_norm[image_norm_key] = rendered_variation_pattern

        parent_sku = model.sku_override or model.parent_sku or f"MOD-{model.id}"
        version_prefix = _slice_to_version_prefix(parent_sku)

        export_children = []

        enabled_role_configs: List[MaterialRoleConfig] = (
            db.query(MaterialRoleConfig)
            .filter(MaterialRoleConfig.ebay_variation_enabled.is_(True))
            .all()
        )
        enabled_role_configs.sort(
            key=lambda rc: min(
                _get_role_rank_from_abbrev((rc.sku_abbrev_no_padding or "").strip().upper()),
                _get_role_rank_from_abbrev((rc.sku_abbrev_with_padding or "").strip().upper()),
            )
        )

        _all_enabled_design_options: List[DesignOption] = (
            db.query(DesignOption)
            .join(EquipmentTypeDesignOption, EquipmentTypeDesignOption.design_option_id == DesignOption.id)
            .filter(
                EquipmentTypeDesignOption.equipment_type_id == model.equipment_type_id,
                DesignOption.ebay_variation_enabled.is_(True),
            )
            .distinct(DesignOption.id)
            .all()
        )
        # BUG FIX 1: Filter strictly by selected design_option_ids when provided.
        # If design_option_ids is empty → no design option axis is used.
        selected_design_option_id_set: Set[int] = set(effective_design_option_ids)
        if selected_design_option_id_set:
            enabled_design_options = [
                o for o in _all_enabled_design_options
                if o.id in selected_design_option_id_set
            ]
        else:
            enabled_design_options = []
        enabled_design_options.sort(key=lambda o: o.id)
        option_axes = [opt.name for opt in enabled_design_options]
        option_yes_no = list(product([False, True], repeat=len(enabled_design_options))) if enabled_design_options else [()]

        target_colors = set(effective_color_surcharge_ids) if (effective_export_mode_norm == "selection_driven") else set()
        target_design_opts = set(effective_design_option_ids) if (effective_export_mode_norm == "selection_driven") else set()
        target_role_keys = (
            {_normalize_key(rk) for rk in effective_role_keys if _normalize_key(rk)}
            if (effective_export_mode_norm == "selection_driven" and effective_role_keys)
            else set()
        )
        per_role_diag_segments: List[str] = []
        diag_roles_used = 0
        diag_fabric_variants_considered = 0
        diag_colors_before_filter = 0
        diag_colors_after_filter = 0
        diag_option_flag_combos = 0
        diag_candidate_combos = 0

        for rc in enabled_role_configs:
            role_key = _normalize_key(rc.role or "")
            if not role_key:
                continue
            if target_role_keys and role_key not in target_role_keys:
                continue

            material = active_material_by_role.get(role_key)
            if not material:
                per_role_diag_segments.append(
                    f"{role_key}:ha=false,mid=null,cb=0,ca=0,ov=0,ids=[]"
                )
                logger.warning(
                    "Model %s: No active material assignment for role '%s'; skipping role.",
                    model.id,
                    role_key,
                )
                continue
            diag_roles_used += 1

            colors = _get_enabled_colors_for_material(material.id)
            role_color_ids_sorted = sorted([
                int(c.get("id", 0))
                for c in colors
                if int(c.get("id", 0)) > 0
            ])
            role_overlap_count = (
                sum(
                    1
                    for c in colors
                    if any(int(cid) in target_colors for cid in list(c.get("selector_ids", set()) or set()))
                )
                if target_colors
                else 0
            )
            role_colors_after_count = role_overlap_count if target_colors else len(role_color_ids_sorted)
            per_role_diag_segments.append(
                f"{role_key}:ha=true,mid={material.id},cb={len(role_color_ids_sorted)},ca={role_colors_after_count},"
                f"ov={role_overlap_count},ids={role_color_ids_sorted[:10]}"
            )
            if not colors:
                logger.warning(
                    "Model %s: Material '%s' has no eBay-enabled colors; skipping role '%s'.",
                    model.id,
                    material.name,
                    role_key,
                )
                continue
            diag_colors_before_filter += len(colors)

            fabric_variants = []
            no_pad_abbrev = (rc.sku_abbrev_no_padding or "").strip().upper()
            with_pad_abbrev = (rc.sku_abbrev_with_padding or "").strip().upper()
            if False in effective_allowed_padding_set and no_pad_abbrev:
                fabric_variants.append((False, no_pad_abbrev))
            if True in effective_allowed_padding_set and with_pad_abbrev:
                fabric_variants.append((True, with_pad_abbrev))
            diag_fabric_variants_considered += len(fabric_variants)

            for with_padding, role_abbrev in fabric_variants:
                for color in colors:
                    selector_ids = {int(cid) for cid in list(color.get("selector_ids", set()) or set())}
                    if target_colors and not (selector_ids & target_colors):
                        continue
                    diag_colors_after_filter += 1
                    color_abbrev = str(color.get("color_code", "") or "").strip().upper()
                    if not color_abbrev:
                        continue

                    for flags in option_yes_no:
                        diag_option_flag_combos += 1
                        selected_opts = [opt for opt, is_yes in zip(enabled_design_options, flags) if is_yes]
                        selected_ids = {opt.id for opt in selected_opts}
                        if target_design_opts and not selected_ids.issubset(target_design_opts):
                            continue
                        diag_candidate_combos += 1

                        design_suffix = _build_option_suffix_from_opts(selected_opts)
                        final_sku = f"{version_prefix}{role_abbrev}{color_abbrev}{design_suffix}"
                        attrs = get_child_attrs(rc, with_padding, color, enabled_design_options, selected_ids)
                        base_price_cents = _get_base_price_cents_for_role_abbrev(model.id, role_abbrev)

                        if base_price_cents is None:
                            msg = (
                                f"Missing base pricing snapshot for model {model.id} fabric variant {role_abbrev}."
                            )
                            if effective_export_mode_norm == "selection_driven":
                                raise HTTPException(status_code=400, detail=msg)
                            logger.warning(msg)
                            retail_price_cents = None
                        else:
                            retail_price_cents = (
                                base_price_cents
                                + _color_surcharge_cents(color)
                                + _design_options_price_cents(selected_opts)
                            )

                        sort_key = (
                            _get_role_rank_from_abbrev(role_abbrev),
                            1 if with_padding else 0,
                            _get_color_sort_tuple_from_code(color_abbrev),
                            len(design_suffix),
                            design_suffix,
                        )

                        export_children.append(
                            {
                                "sku": final_sku,
                                "attrs": attrs,
                                "sort_key": sort_key,
                                "color_code": color_abbrev,
                                "color_label": attrs["Color"],
                                "color_selector_ids": sorted(selector_ids),
                                "role_key": role_key,
                                "with_padding": with_padding,
                                "retail_price_cents": retail_price_cents,
                            }
                        )

        export_children.sort(key=lambda x: x["sort_key"])
        if not export_children:
            if effective_export_mode_norm == "selection_driven":
                export_warnings.append(
                    "Selection-driven export produced zero variations for model "
                    f"{model.id}; exported parent-only. diag: "
                    f"model_id={model.id} "
                    f"mode={effective_export_mode_norm} "
                    f"with_padding={effective_with_padding_raw} "
                    f"roles_sel={len(effective_role_keys or [])} "
                    f"colors_sel={len(effective_color_surcharge_ids or [])} "
                    f"opts_sel={len(effective_design_option_ids or [])} "
                    f"roles_enabled={len(enabled_role_configs)} "
                    f"roles_used={diag_roles_used} "
                    f"fabric_variants={diag_fabric_variants_considered} "
                    f"colors_before={diag_colors_before_filter} "
                    f"colors_after={diag_colors_after_filter} "
                    f"option_flag_combos={diag_option_flag_combos} "
                    f"candidate_combos={diag_candidate_combos} "
                    f"target_colors_count={len(target_colors)} "
                    f"target_color_ids={sorted(list(target_colors))[:10]} "
                    f"per_role=[{';'.join(per_role_diag_segments)}]"
                )
            logger.warning(
                "Model %s produced zero child variations under current filters.",
                model.id,
            )

        valid_axes = ["Fabric", "Color"] + option_axes

        axis_values_map = {axis: set() for axis in valid_axes}
        color_code_by_label: Dict[str, str] = {}
        for item in export_children:
            for k, v in item["attrs"].items():
                if k in axis_values_map:
                    axis_values_map[k].add(v)
            color_code_by_label[item["color_label"]] = item["color_code"]

        _validate_color_label_consistency_for_children(
            export_children=export_children,
            parent_color_values={str(v) for v in axis_values_map.get("Color", set())},
        )

        try:
            row_by_norm = dict(model_parent_defaults_by_norm)
            explicit_parent_action_override = any(
                str(getattr(field, attr, "") or "").strip()
                for field in fields
                if _header_key_matches(getattr(field, "field_name", ""), KEY_ACTION)
                for attr in ("parent_custom_value", "parent_selected_value")
            )
            current_action = ""
            for action_key in action_header_keys:
                value = str(row_by_norm.get(action_key, "") or "").strip()
                if value:
                    current_action = value
                    break
            ebay_item_id = str(ebay_listing_by_model_id.get(model.id, "") or "").strip()
            if ebay_item_id:
                if not explicit_parent_action_override:
                    for action_key in action_header_keys:
                        row_by_norm[action_key] = "Revise"
            elif not current_action and not explicit_parent_action_override:
                for action_key in action_header_keys:
                    row_by_norm[action_key] = "Add"
            if ebay_item_id:
                for item_id_key in item_id_header_keys:
                    row_by_norm[item_id_key] = ebay_item_id
            row_by_norm[KEY_SKU] = parent_sku

            rel_segments = []
            for axis in valid_axes:
                raw_vals = list(axis_values_map[axis])

                if axis == "Color":
                    raw_vals.sort(key=lambda v: _get_color_sort_tuple_from_code(color_code_by_label.get(v, "")))
                elif axis == "Fabric":
                    def _fabric_sort_key(s: str):
                        rc2 = None
                        is_pad = False
                        for candidate in role_configs.values():
                            if fabric_label(candidate, False) == s:
                                rc2 = candidate
                                is_pad = False
                                break
                            if fabric_label(candidate, True) == s:
                                rc2 = candidate
                                is_pad = True
                                break
                        rank = 999
                        if rc2:
                            ab = rc2.sku_abbrev_with_padding if is_pad else rc2.sku_abbrev_no_padding
                            if ab:
                                rank = _get_role_rank_from_abbrev(ab)
                        return (rank, 1 if is_pad else 0, s)

                    raw_vals.sort(key=_fabric_sort_key)
                elif axis in option_axes:
                    # BUG FIX 2: Parent row must always show both options as "No;Yes"
                    # regardless of what child combinations exist.
                    rel_segments.append(f"{axis}=No;Yes")
                    continue
                else:
                    raw_vals.sort()

                if raw_vals:
                    rel_segments.append(f"{axis}={';'.join(raw_vals)}")

            row_by_norm[KEY_REL_DETAILS] = "|".join(rel_segments)
            for description_norm_key, description_html_template in description_template_by_norm.items():
                row_by_norm[description_norm_key] = _render_description_html_value(
                    description_html=description_html_template,
                    token_context=token_context,
                    image_index=1,
                    db=db,
                    model_id=model.id,
                    field_name=description_norm_key,
                )
            if settings_parent_image_pattern:
                rendered_parent_image = _render_settings_image_pattern(
                    settings_parent_image_pattern,
                    token_context=token_context,
                    model_id=model.id,
                    field_name="settings_parent_image_pattern",
                )
                for image_header_key in image_url_header_norm_keys:
                    row_by_norm[image_header_key] = rendered_parent_image

            row_out = [
                row_by_norm.get(header_key, "")
                if header_key in parent_allowed_keys
                else ""
                for header_key in header_norms
            ]
            writer.writerow(row_out)

        except Exception as e:
            logger.error(f"Error writing parent {parent_sku}: {e}")
            raise

        for item in export_children:
            try:
                sku = item["sku"]
                attrs = item["attrs"]

                row_by_norm = dict(model_variation_defaults_by_norm)
                row_by_norm[KEY_ACTION] = "Add"
                row_by_norm[KEY_SKU] = sku

                rel_segments = []
                for axis in valid_axes:
                    val = attrs.get(axis)
                    if val:
                        rel_segments.append(f"{axis}={val}")
                row_by_norm[KEY_REL_DETAILS] = "|".join(rel_segments)
                for description_norm_key, description_html_template in description_template_by_norm.items():
                    row_by_norm[description_norm_key] = _render_description_html_value(
                        description_html=description_html_template,
                        token_context=token_context,
                        color_abbr=item.get("color_code"),
                        image_index=1,
                        db=db,
                        model_id=model.id,
                        field_name=description_norm_key,
                    )

                if item.get("retail_price_cents") is not None:
                    row_by_norm[KEY_PRICE] = f"{item['retail_price_cents'] / 100:.2f}"
                if not row_by_norm.get(KEY_QTY):
                    row_by_norm[KEY_QTY] = "1"
                for image_norm_key, image_variation_pattern in variation_image_pattern_by_norm.items():
                    row_by_norm[image_norm_key] = _build_image_urls_from_pattern(
                        image_variation_pattern,
                        color_sku=item.get("color_code"),
                        color_friendly_name=item.get("color_label"),
                    )
                if settings_variation_image_pattern:
                    rendered_variation_image = _render_settings_image_pattern(
                        settings_variation_image_pattern,
                        token_context=token_context,
                        color_abbr=item.get("color_code"),
                        model_id=model.id,
                        field_name="settings_variation_image_pattern",
                    )
                    for image_header_key in image_url_header_norm_keys:
                        row_by_norm[image_header_key] = rendered_variation_image

                row_out = [
                    row_by_norm.get(header_key, "")
                    if header_key in variation_allowed_keys
                    else ""
                    for header_key in header_norms
                ]
                writer.writerow(row_out)

            except Exception as e:
                logger.error(f"Error writing child {sku}: {e}")
                raise

    csv_bytes = output.getvalue().encode("utf-8-sig")
    output.close()

    mfr_name, series_name, date_str = _build_ebay_filename_tokens(models, db)
    filename = f"Ebay_{mfr_name}_{series_name}_{date_str}.csv"
    response_headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    if export_warnings:
        response_headers["X-Export-Warnings"] = "|".join(export_warnings)
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8",
        headers=response_headers,
    )

