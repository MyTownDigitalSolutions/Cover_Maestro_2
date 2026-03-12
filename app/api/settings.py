from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.orm import Session
from sqlalchemy import desc
from typing import List
from datetime import datetime, timedelta, timezone
import traceback
import shutil
import os
from fastapi import File, UploadFile
import hashlib


from app.database import get_db
from app.models.enums import Carrier
from app.models.core import (
    MaterialRoleAssignment, Material,
    ShippingRateCard, ShippingRateTier, ShippingZoneRate, MarketplaceShippingProfile,
    LaborSetting, MarketplaceFeeRate, VariantProfitSetting, ShippingZone, ShippingDefaultSetting,
    ExportSetting, EquipmentType, EbayVariationPresetAsset
)
from app.models.templates import AmazonCustomizationTemplate, EquipmentTypeCustomizationTemplate, ReverbTemplate
from app.services.storage_policy import (
    ensure_storage_dirs_exist,
    assert_allowed_write_path,
    get_customization_template_paths,
    rotate_customization_template_backup,
)
from app.schemas.core import (
    MaterialRoleAssignmentCreate, MaterialRoleAssignmentResponse,
    ShippingRateCardCreate, ShippingRateCardResponse, ShippingRateCardUpdate,
    ShippingRateTierCreate, ShippingRateTierResponse, ShippingRateTierUpdate, TierCreateRequest,
    ShippingZoneRateCreate, ShippingZoneRateResponse,  # existing
    MarketplaceShippingProfileCreate, MarketplaceShippingProfileResponse, MarketplaceShippingProfileUpdate,
    LaborSettingCreate, LaborSettingResponse,
    MarketplaceFeeRateCreate, MarketplaceFeeRateResponse,
    VariantProfitSettingCreate, VariantProfitSettingResponse,
    ShippingZoneResponse,
    ShippingZoneRateNormalizedResponse, ShippingZoneRateUpsertRequest,
    ShippingDefaultSettingCreate, ShippingDefaultSettingResponse,
    EbayVariationPresetCreate, EbayVariationPresetUpdate, EbayVariationPresetResponse,
    AmazonCustomizationTemplateAssignmentRequest, EquipmentTypeResponse,
    AmazonCustomizationTemplatePreviewResponse,
    EquipmentTypeCustomizationTemplateAssignRequest,
    EquipmentTypeCustomizationTemplateItem,
    EquipmentTypeCustomizationTemplatesResponse,
    EquipmentTypeCustomizationTemplateSetDefaultRequest,
    ReverbTemplateAssignmentRequest
)

router = APIRouter(prefix="/settings", tags=["settings"])


def _normalize_to_utc_naive(dt: datetime | None) -> datetime:
    if dt is None:
        return datetime.utcnow()
    if dt.tzinfo is not None:
        return dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def _coerce_today_midnight_to_now(dt: datetime) -> datetime:
    now = datetime.utcnow()
    if (
        dt.hour == 0
        and dt.minute == 0
        and dt.second == 0
        and dt.microsecond == 0
        and dt.date() == now.date()
    ):
        return now
    return dt


# ------------------------------------------------------------------
# 1. Material Roles
# ------------------------------------------------------------------

@router.get("/material-roles", response_model=List[MaterialRoleAssignmentResponse])
def get_material_role_assignments(include_history: bool = False, db: Session = Depends(get_db)):
    query = db.query(MaterialRoleAssignment)
    if not include_history:
        # Return only currently active
        now = datetime.utcnow()
        query = query.filter(
            MaterialRoleAssignment.effective_date <= now,
            (MaterialRoleAssignment.end_date == None) | (MaterialRoleAssignment.end_date > now)
        )
    return query.order_by(desc(MaterialRoleAssignment.effective_date)).all()


@router.post("/material-roles/assign", response_model=MaterialRoleAssignmentResponse)
def assign_material_role(data: MaterialRoleAssignmentCreate, db: Session = Depends(get_db)):
    # Validate material
    material = db.query(Material).filter(Material.id == data.material_id).first()
    if not material:
        raise HTTPException(status_code=400, detail="Invalid material ID")

    now = _coerce_today_midnight_to_now(_normalize_to_utc_naive(data.effective_date))

    # Prevent backdating
    if now < datetime.utcnow() - timedelta(minutes=1):  # Allow small leeway for network time
        raise HTTPException(status_code=400, detail="Effective date cannot be in the past.")

    # Close existing active assignment for this role
    existing = db.query(MaterialRoleAssignment).filter(
        MaterialRoleAssignment.role == data.role,
        MaterialRoleAssignment.effective_date <= now,
        (MaterialRoleAssignment.end_date == None) | (MaterialRoleAssignment.end_date > now)
    ).first()

    if existing:
        if existing.effective_date >= now:
            raise HTTPException(status_code=400, detail="Cannot supersede an assignment with same or future effective date.")
        existing.end_date = now

    new_assignment = MaterialRoleAssignment(
        role=data.role,
        material_id=data.material_id,
        effective_date=now
    )
    db.add(new_assignment)
    db.commit()
    db.refresh(new_assignment)
    return new_assignment


# ------------------------------------------------------------------
# 2. Shipping Configuration
# ------------------------------------------------------------------

@router.get("/shipping/zones", response_model=List[ShippingZoneResponse])
def get_shipping_zones(db: Session = Depends(get_db)):
    return db.query(ShippingZone).order_by(ShippingZone.sort_order).all()


# --------------------------
# Rate Cards
# --------------------------

@router.get("/shipping/rate-cards", response_model=List[ShippingRateCardResponse])
def list_rate_cards(include_inactive: bool = False, db: Session = Depends(get_db)):
    """
    Hardening:
      - Avoid 500 if DB/model mismatch around `active`
      - Return a meaningful error detail + print stack trace
    """
    try:
        query = db.query(ShippingRateCard)

        # If the model doesn't have `active` (schema mismatch), don't filter by it.
        if hasattr(ShippingRateCard, "active"):
            if not include_inactive:
                query = query.filter(ShippingRateCard.active == True)  # noqa: E712

        return query.order_by(ShippingRateCard.name, ShippingRateCard.id).all()

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"rate-cards query failed: {e}")


@router.post("/shipping/rate-cards", response_model=ShippingRateCardResponse)
def create_rate_card(data: ShippingRateCardCreate, db: Session = Depends(get_db)):
    # Validate uniqueness of name
    existing = db.query(ShippingRateCard).filter(ShippingRateCard.name == data.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Rate card with this name already exists")

    # Build a safe dict from schema (frontend now sends only {name})
    payload = data.dict()

    # If the DB model has a carrier column, default it to USPS.
    # (This prevents a crash if the column exists & is non-nullable.)
    if hasattr(ShippingRateCard, "carrier") and "carrier" not in payload:
        payload["carrier"] = Carrier.USPS

    # If model has active and schema doesn't include it, default it True
    if hasattr(ShippingRateCard, "active") and "active" not in payload:
        payload["active"] = True

    card = ShippingRateCard(**payload)
    db.add(card)
    db.commit()
    db.refresh(card)
    return card


@router.put("/shipping/rate-cards/{card_id}", response_model=ShippingRateCardResponse)
def update_rate_card(card_id: int, data: ShippingRateCardUpdate, db: Session = Depends(get_db)):
    card = db.query(ShippingRateCard).filter(ShippingRateCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Rate card not found")

    if data.name is not None:
        # Check uniqueness if name changing
        if data.name != card.name:
            existing = db.query(ShippingRateCard).filter(ShippingRateCard.name == data.name).first()
            if existing:
                raise HTTPException(status_code=400, detail="Rate card with this name already exists")
        card.name = data.name

    if data.active is not None and hasattr(card, "active"):
        card.active = data.active

    db.commit()
    db.refresh(card)
    return card


@router.delete("/shipping/rate-cards/{card_id}")
def delete_rate_card(card_id: int, db: Session = Depends(get_db)):
    card = db.query(ShippingRateCard).filter(ShippingRateCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Rate card not found")

    # Soft delete (only if active exists)
    if hasattr(card, "active"):
        card.active = False
        db.commit()
        return {"message": "Rate card archived (soft deleted)"}

    # If the model doesn't support soft-delete, fail loudly
    raise HTTPException(status_code=500, detail="Rate card model does not support archiving (missing active column)")


# --------------------------
# Tiers
# --------------------------

@router.get("/shipping/rate-cards/{card_id}/tiers", response_model=List[ShippingRateTierResponse])
def list_tiers(card_id: int, include_inactive: bool = False, db: Session = Depends(get_db)):
    query = db.query(ShippingRateTier).filter(ShippingRateTier.rate_card_id == card_id)
    if hasattr(ShippingRateTier, "active") and not include_inactive:
        query = query.filter(ShippingRateTier.active == True)  # noqa: E712
    return query.order_by(ShippingRateTier.max_oz, ShippingRateTier.id).all()


@router.post("/shipping/rate-cards/{card_id}/tiers", response_model=ShippingRateTierResponse)
def create_tier_under_card(card_id: int, data: TierCreateRequest, db: Session = Depends(get_db)):
    # Verify card exists
    card = db.query(ShippingRateCard).filter(ShippingRateCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Rate card not found")

    tier = ShippingRateTier(
        rate_card_id=card_id,
        min_oz=0.0,
        max_oz=data.max_weight_oz,
        label=data.label,
        active=True if hasattr(ShippingRateTier, "active") else None
    )
    db.add(tier)
    db.commit()
    db.refresh(tier)
    return tier


@router.put("/shipping/tiers/{tier_id}", response_model=ShippingRateTierResponse)
def update_tier(tier_id: int, data: ShippingRateTierUpdate, db: Session = Depends(get_db)):
    tier = db.query(ShippingRateTier).filter(ShippingRateTier.id == tier_id).first()
    if not tier:
        raise HTTPException(status_code=404, detail="Tier not found")

    if data.label is not None:
        tier.label = data.label
    if data.max_weight_oz is not None:
        tier.max_oz = data.max_weight_oz
    if data.active is not None and hasattr(tier, "active"):
        tier.active = data.active

    db.commit()
    db.refresh(tier)
    return tier


@router.delete("/shipping/tiers/{tier_id}")
def delete_tier(tier_id: int, db: Session = Depends(get_db)):
    tier = db.query(ShippingRateTier).filter(ShippingRateTier.id == tier_id).first()
    if not tier:
        raise HTTPException(status_code=404, detail="Tier not found")

    if hasattr(tier, "active"):
        tier.active = False
        db.commit()
        return {"message": "Tier archived"}

    raise HTTPException(status_code=500, detail="Tier model does not support archiving (missing active column)")


# --------------------------
# Zone Rates
# --------------------------

@router.get("/shipping/tiers/{tier_id}/zone-rates", response_model=List[ShippingZoneRateNormalizedResponse])
def list_zone_rates(tier_id: int, db: Session = Depends(get_db)):
    zones = db.query(ShippingZone).order_by(ShippingZone.sort_order).all()

    rates = db.query(ShippingZoneRate).filter(ShippingZoneRate.tier_id == tier_id).all()
    rate_map = {r.zone: r for r in rates}  # zone_id -> rate row

    results = []
    for z in zones:
        r = rate_map.get(z.id)
        results.append({
            "zone_id": z.id,
            "zone_code": z.code,
            "zone_name": z.name,
            "rate_cents": r.rate_cents if r else None,
            "zone_rate_id": r.id if r else None
        })

    return results


@router.put("/shipping/tiers/{tier_id}/zone-rates/{zone_id}", response_model=ShippingZoneRateNormalizedResponse)
def upsert_zone_rate(tier_id: int, zone_id: int, data: ShippingZoneRateUpsertRequest, db: Session = Depends(get_db)):
    tier = db.query(ShippingRateTier).filter(ShippingRateTier.id == tier_id).first()
    if not tier:
        raise HTTPException(status_code=404, detail="Tier not found")

    zone = db.query(ShippingZone).filter(ShippingZone.id == zone_id).first()
    if not zone:
        raise HTTPException(status_code=404, detail="Zone not found")

    existing_rate = db.query(ShippingZoneRate).filter(
        ShippingZoneRate.tier_id == tier_id,
        ShippingZoneRate.zone == zone_id
    ).first()

    returned_rate_id = None
    returned_cents = None

    if data.rate_cents is None:
        if existing_rate:
            db.delete(existing_rate)
            db.commit()
    else:
        if existing_rate:
            existing_rate.rate_cents = data.rate_cents
            db.commit()
            db.refresh(existing_rate)
            returned_rate_id = existing_rate.id
            returned_cents = existing_rate.rate_cents
        else:
            new_rate = ShippingZoneRate(
                rate_card_id=tier.rate_card_id,
                tier_id=tier_id,
                zone=zone_id,
                rate_cents=data.rate_cents
            )
            db.add(new_rate)
            db.commit()
            db.refresh(new_rate)
            returned_rate_id = new_rate.id
            returned_cents = new_rate.rate_cents

    return {
        "zone_id": zone.id,
        "zone_code": zone.code,
        "zone_name": zone.name,
        "rate_cents": returned_cents,
        "zone_rate_id": returned_rate_id
    }


@router.post("/shipping/zone-rates", response_model=ShippingZoneRateResponse)
def create_zone_rate(data: ShippingZoneRateCreate, db: Session = Depends(get_db)):
    existing = db.query(ShippingZoneRate).filter(
        ShippingZoneRate.tier_id == data.tier_id,
        ShippingZoneRate.zone == data.zone
    ).first()
    if existing:
        existing.rate_cents = data.rate_cents
        db.commit()
        db.refresh(existing)
        return existing

    rate = ShippingZoneRate(**data.dict())
    db.add(rate)
    db.commit()
    db.refresh(rate)
    return rate


@router.put("/shipping/zone-rates/{rate_id}", response_model=ShippingZoneRateResponse)
def update_zone_rate(rate_id: int, data: ShippingZoneRateCreate, db: Session = Depends(get_db)):
    rate = db.query(ShippingZoneRate).filter(ShippingZoneRate.id == rate_id).first()
    if not rate:
        raise HTTPException(status_code=404, detail="Rate not found")
    rate.rate_cents = data.rate_cents
    db.commit()
    db.refresh(rate)
    return rate


@router.delete("/shipping/zone-rates/{rate_id}")
def delete_zone_rate(rate_id: int, db: Session = Depends(get_db)):
    rate = db.query(ShippingZoneRate).filter(ShippingZoneRate.id == rate_id).first()
    if not rate:
        raise HTTPException(status_code=404, detail="Rate not found")
    db.delete(rate)
    db.commit()
    return {"message": "Rate deleted"}


# ------------------------------------------------------------------
# 3. Marketplace Profiles
# ------------------------------------------------------------------

@router.get("/shipping/marketplace-profiles", response_model=List[MarketplaceShippingProfileResponse])
def list_marketplace_profiles(include_history: bool = False, db: Session = Depends(get_db)):
    query = db.query(MarketplaceShippingProfile)
    if not include_history:
        now = datetime.utcnow()
        query = query.filter(
            MarketplaceShippingProfile.effective_date <= now,
            (MarketplaceShippingProfile.end_date == None) | (MarketplaceShippingProfile.end_date > now)
        )
    return query.all()


@router.post("/shipping/marketplace-profiles/assign", response_model=MarketplaceShippingProfileResponse)
def assign_marketplace_profile(data: MarketplaceShippingProfileCreate, db: Session = Depends(get_db)):
    now = _coerce_today_midnight_to_now(_normalize_to_utc_naive(data.effective_date))

    # Prevent backdating
    if now < datetime.utcnow() - timedelta(minutes=1):
        raise HTTPException(status_code=400, detail="Effective date cannot be in the past.")

    existing = db.query(MarketplaceShippingProfile).filter(
        MarketplaceShippingProfile.marketplace == data.marketplace,
        MarketplaceShippingProfile.effective_date <= now,
        (MarketplaceShippingProfile.end_date == None) | (MarketplaceShippingProfile.end_date > now)
    ).first()

    if existing:
        if existing.effective_date >= now:
            raise HTTPException(status_code=400, detail="Cannot supersede an assignment with same or future effective date.")
        existing.end_date = now

    new_profile = MarketplaceShippingProfile(
        marketplace=data.marketplace,
        rate_card_id=data.rate_card_id,
        pricing_zone=data.pricing_zone,
        effective_date=now
    )
    db.add(new_profile)
    db.commit()
    db.refresh(new_profile)
    return new_profile


@router.put("/shipping/marketplace-profiles/{profile_id}", response_model=MarketplaceShippingProfileResponse)
def update_marketplace_profile(profile_id: int, data: MarketplaceShippingProfileUpdate, db: Session = Depends(get_db)):
    profile = db.query(MarketplaceShippingProfile).filter(MarketplaceShippingProfile.id == profile_id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Marketplace profile not found")

    if data.marketplace is not None:
        profile.marketplace = data.marketplace
    if data.rate_card_id is not None:
        profile.rate_card_id = data.rate_card_id
    if data.pricing_zone is not None:
        profile.pricing_zone = data.pricing_zone
    if data.effective_date is not None:
        profile.effective_date = _coerce_today_midnight_to_now(_normalize_to_utc_naive(data.effective_date))
    if data.end_date is not None:
        profile.end_date = _normalize_to_utc_naive(data.end_date)

    db.commit()
    db.refresh(profile)
    return profile


@router.delete("/shipping/marketplace-profiles/{profile_id}")
def delete_marketplace_profile(profile_id: int, db: Session = Depends(get_db)):
    profile = db.query(MarketplaceShippingProfile).filter(MarketplaceShippingProfile.id == profile_id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Marketplace profile not found")
    db.delete(profile)
    db.commit()
    return {"message": "Marketplace profile deleted"}


# --------------------------
# Defaults
# --------------------------

@router.get("/shipping/defaults", response_model=ShippingDefaultSettingResponse)
def get_shipping_defaults(db: Session = Depends(get_db)):
    settings = db.query(ShippingDefaultSetting).first()
    if not settings:
        settings = ShippingDefaultSetting()
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return settings


@router.put("/shipping/defaults", response_model=ShippingDefaultSettingResponse)
def update_shipping_defaults(data: ShippingDefaultSettingCreate, db: Session = Depends(get_db)):
    if data.shipping_mode not in ["flat", "calculated", "fixed_cell"]:
        raise HTTPException(status_code=400, detail="Invalid shipping_mode")

    assumed_present = [
        data.assumed_rate_card_id is not None,
        data.assumed_tier_id is not None,
        data.assumed_zone_code is not None
    ]

    if any(assumed_present) and not all(assumed_present):
        raise HTTPException(status_code=400, detail="If any assumed settings are provided (card, tier, zone), all must be provided.")

    if all(assumed_present):
        card = db.query(ShippingRateCard).filter(ShippingRateCard.id == data.assumed_rate_card_id).first()
        if not card:
            raise HTTPException(status_code=400, detail="Assumed rate card not found")

        tier = db.query(ShippingRateTier).filter(ShippingRateTier.id == data.assumed_tier_id).first()
        if not tier:
            raise HTTPException(status_code=400, detail="Assumed tier not found")

        if tier.rate_card_id != data.assumed_rate_card_id:
            raise HTTPException(status_code=400, detail="Assumed tier does not belong to the assumed rate card")

        if data.assumed_zone_code not in [str(i) for i in range(1, 10)]:  # 1..9
            raise HTTPException(status_code=400, detail="Assumed zone code must be '1' through '9'")

    settings = db.query(ShippingDefaultSetting).first()
    if not settings:
        settings = ShippingDefaultSetting()
        db.add(settings)

    settings.shipping_mode = data.shipping_mode
    settings.flat_shipping_cents = data.flat_shipping_cents
    settings.default_rate_card_id = data.default_rate_card_id
    settings.default_zone_code = data.default_zone_code

    settings.assumed_rate_card_id = data.assumed_rate_card_id
    settings.assumed_tier_id = data.assumed_tier_id
    settings.assumed_zone_code = data.assumed_zone_code

    settings.shipping_settings_version += 1

    db.commit()
    db.refresh(settings)
    return settings


# ------------------------------------------------------------------
# 4. Labor Settings
# ------------------------------------------------------------------

@router.get("/labor", response_model=LaborSettingResponse)
def get_labor_settings(db: Session = Depends(get_db)):
    settings = db.query(LaborSetting).first()
    if not settings:
        settings = LaborSetting()
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return settings


@router.put("/labor", response_model=LaborSettingResponse)
def update_labor_settings(data: LaborSettingCreate, db: Session = Depends(get_db)):
    settings = db.query(LaborSetting).first()
    if not settings:
        settings = LaborSetting()
        db.add(settings)

    settings.hourly_rate_cents = data.hourly_rate_cents
    settings.minutes_no_padding = data.minutes_no_padding
    settings.minutes_with_padding = data.minutes_with_padding

    db.commit()
    db.refresh(settings)
    return settings


# ------------------------------------------------------------------
# 5. Marketplace Fees
# ------------------------------------------------------------------

@router.get("/marketplace-fees", response_model=List[MarketplaceFeeRateResponse])
def list_fees(db: Session = Depends(get_db)):
    return db.query(MarketplaceFeeRate).all()


@router.put("/marketplace-fees", response_model=MarketplaceFeeRateResponse)
def update_fee(data: MarketplaceFeeRateCreate, db: Session = Depends(get_db)):
    fee = db.query(MarketplaceFeeRate).filter(MarketplaceFeeRate.marketplace == data.marketplace).first()
    if not fee:
        fee = MarketplaceFeeRate(marketplace=data.marketplace)
        db.add(fee)

    fee.fee_rate = data.fee_rate
    db.commit()
    db.refresh(fee)
    return fee


# ------------------------------------------------------------------
# 6. Profits
# ------------------------------------------------------------------

@router.get("/profits", response_model=List[VariantProfitSettingResponse])
def list_profits(db: Session = Depends(get_db)):
    return db.query(VariantProfitSetting).all()


@router.put("/profits", response_model=VariantProfitSettingResponse)
def update_profit(data: VariantProfitSettingCreate, db: Session = Depends(get_db)):
    profit = db.query(VariantProfitSetting).filter(VariantProfitSetting.variant_key == data.variant_key).first()
    if not profit:
        profit = VariantProfitSetting(variant_key=data.variant_key)
        db.add(profit)

    profit.profit_cents = data.profit_cents
    db.commit()
    db.refresh(profit)
    return profit


# ------------------------------------------------------------------
# 7. Export Settings
# ------------------------------------------------------------------

from app.models.core import ExportSetting
from app.schemas.core import ExportSettingResponse, ExportSettingCreate

@router.get("/export", response_model=ExportSettingResponse)
def get_export_settings(db: Session = Depends(get_db)):
    settings = db.query(ExportSetting).first()
    if not settings:
        settings = ExportSetting(default_save_path_template="")
        db.add(settings)
        db.commit()
        db.refresh(settings)
    elif (
        not getattr(settings, "ebay_store_category_default_level", None)
        or not getattr(settings, "ebay_description_selection_mode", None)
    ):
        if not getattr(settings, "ebay_store_category_default_level", None):
            settings.ebay_store_category_default_level = "series"
        if not getattr(settings, "ebay_description_selection_mode", None):
            settings.ebay_description_selection_mode = "GLOBAL_PRIMARY"
        db.commit()
        db.refresh(settings)
    return settings


@router.put("/export", response_model=ExportSettingResponse)
def update_export_settings(data: ExportSettingCreate, db: Session = Depends(get_db)):
    settings = db.query(ExportSetting).first()
    if not settings:
        settings = ExportSetting()
        db.add(settings)

    if data.default_save_path_template is not None:
        settings.default_save_path_template = data.default_save_path_template
    
    if data.amazon_customization_export_format is not None:
        settings.amazon_customization_export_format = data.amazon_customization_export_format

    if data.ebay_store_category_default_level is not None:
        if data.ebay_store_category_default_level not in ("series", "manufacturer", "equipment_type"):
            raise HTTPException(status_code=400, detail="Invalid ebay_store_category_default_level")
        settings.ebay_store_category_default_level = data.ebay_store_category_default_level

    if data.ebay_description_selection_mode is not None:
        if data.ebay_description_selection_mode not in ("GLOBAL_PRIMARY", "EQUIPMENT_TYPE_PRIMARY"):
            raise HTTPException(status_code=400, detail="Invalid ebay_description_selection_mode")
        settings.ebay_description_selection_mode = data.ebay_description_selection_mode

    def _validate_image_pattern(
        *,
        field_name: str,
        raw_value: str,
        require_color_token: bool,
    ) -> str:
        trimmed = raw_value.strip()
        if "\n" in trimmed or "\r" in trimmed:
            raise HTTPException(status_code=400, detail=f"{field_name} must not contain newlines")
        if "[INDEX]" not in trimmed and "[IMAGE_INDEX]" not in trimmed:
            raise HTTPException(
                status_code=400,
                detail=f"{field_name} must include [INDEX] or [IMAGE_INDEX]",
            )
        if require_color_token and "[COLOR_ABBR]" not in trimmed and "[COLOR_SKU]" not in trimmed:
            raise HTTPException(
                status_code=400,
                detail=f"{field_name} must include [COLOR_ABBR] or [COLOR_SKU]",
            )
        return trimmed

    if "ebay_parent_image_pattern" in data.__fields_set__:
        raw_parent_pattern = data.ebay_parent_image_pattern
        if raw_parent_pattern is None or not raw_parent_pattern.strip():
            settings.ebay_parent_image_pattern = None
        else:
            settings.ebay_parent_image_pattern = _validate_image_pattern(
                field_name="ebay_parent_image_pattern",
                raw_value=raw_parent_pattern,
                require_color_token=False,
            )

    if "ebay_variation_image_pattern" in data.__fields_set__:
        raw_variation_pattern = data.ebay_variation_image_pattern
        if raw_variation_pattern is None or not raw_variation_pattern.strip():
            settings.ebay_variation_image_pattern = None
        else:
            settings.ebay_variation_image_pattern = _validate_image_pattern(
                field_name="ebay_variation_image_pattern",
                raw_value=raw_variation_pattern,
                require_color_token=True,
            )

    # --- eBay fabric template validation ---
    for field_name in ("ebay_fabric_template_no_padding", "ebay_fabric_template_with_padding"):
        raw_value = getattr(data, field_name)
        if raw_value is not None:
            trimmed = raw_value.strip()
            if "\n" in trimmed or "\r" in trimmed:
                raise HTTPException(status_code=400, detail=f"{field_name} must not contain newlines")
            if "{role}" not in trimmed:
                raise HTTPException(status_code=400, detail=f"{field_name} must include {{role}}")
            setattr(settings, field_name, trimmed)
        else:
            setattr(settings, field_name, None)

    db.commit()
    db.refresh(settings)
    return settings


# ------------------------------------------------------------------
# 7.5 eBay Variation Presets
# ------------------------------------------------------------------

def _normalize_ebay_variation_preset_name(value: str | None) -> str:
    text = str(value or "").strip()
    if not text:
        raise HTTPException(status_code=422, detail="name is required")
    return text


def _normalize_equipment_type_ids(values: List[int] | None) -> List[int]:
    return sorted({int(v) for v in (values or [])})


def _validate_equipment_type_ids_exist(equipment_type_ids: List[int], db: Session) -> None:
    if not equipment_type_ids:
        return
    rows = (
        db.query(EquipmentType.id)
        .filter(EquipmentType.id.in_(equipment_type_ids))
        .all()
    )
    found = {int(row[0]) for row in rows}
    missing = [et_id for et_id in equipment_type_ids if et_id not in found]
    if missing:
        raise HTTPException(status_code=422, detail=f"Equipment type(s) not found: {missing}")


@router.get("/ebay-variation-presets", response_model=List[EbayVariationPresetResponse])
def list_ebay_variation_presets(db: Session = Depends(get_db)):
    return (
        db.query(EbayVariationPresetAsset)
        .filter(EbayVariationPresetAsset.marketplace == "EBAY")
        .order_by(EbayVariationPresetAsset.created_at.desc(), EbayVariationPresetAsset.id.desc())
        .all()
    )


@router.post("/ebay-variation-presets", response_model=EbayVariationPresetResponse)
def create_ebay_variation_preset(data: EbayVariationPresetCreate, db: Session = Depends(get_db)):
    name = _normalize_ebay_variation_preset_name(data.name)
    equipment_type_ids = _normalize_equipment_type_ids(data.equipment_type_ids)
    _validate_equipment_type_ids_exist(equipment_type_ids, db)
    row = EbayVariationPresetAsset(
        name=name,
        marketplace="EBAY",
        equipment_type_ids=equipment_type_ids,
        payload=data.payload.model_dump(),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/ebay-variation-presets/{preset_id}", response_model=EbayVariationPresetResponse)
def update_ebay_variation_preset(
    preset_id: int,
    data: EbayVariationPresetUpdate,
    db: Session = Depends(get_db),
):
    row = (
        db.query(EbayVariationPresetAsset)
        .filter(
            EbayVariationPresetAsset.id == preset_id,
            EbayVariationPresetAsset.marketplace == "EBAY",
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Preset not found")

    fields_set = getattr(data, "model_fields_set", getattr(data, "__fields_set__", set()))

    if "name" in fields_set:
        row.name = _normalize_ebay_variation_preset_name(data.name)
    if "equipment_type_ids" in fields_set:
        normalized_ids = _normalize_equipment_type_ids(data.equipment_type_ids)
        _validate_equipment_type_ids_exist(normalized_ids, db)
        row.equipment_type_ids = normalized_ids
    if "payload" in fields_set:
        if data.payload is None:
            raise HTTPException(status_code=422, detail="payload must include all required keys")
        row.payload = data.payload.model_dump()

    db.commit()
    db.refresh(row)
    return row


@router.delete("/ebay-variation-presets/{preset_id}", status_code=204)
def delete_ebay_variation_preset(preset_id: int, db: Session = Depends(get_db)):
    row = (
        db.query(EbayVariationPresetAsset)
        .filter(
            EbayVariationPresetAsset.id == preset_id,
            EbayVariationPresetAsset.marketplace == "EBAY",
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Preset not found")
    db.delete(row)
    db.commit()
    return Response(status_code=204)


# ------------------------------------------------------------------
# 8. Amazon Customization Templates
# ------------------------------------------------------------------

@router.post("/amazon-customization-templates/upload")
def upload_amazon_customization_template(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Upload a new Amazon Customization Template (XLSX).

    Storage policy parity with Product Type Templates:
    - Canonical path in CUSTOMIZATION_DIR (deterministic, collision-safe)
    - Single-backup rotation
    - sha256 verification of persisted bytes
    - Stores canonical file_path in DB (original_filename preserved for display/download name)
    """
    ensure_storage_dirs_exist()

    # Read bytes once (UploadFile stream is consumed when copied)
    file_bytes = file.file.read()
    upload_sha256 = hashlib.sha256(file_bytes).hexdigest()
    mem_size = len(file_bytes)

    # Create DB row first to obtain a stable template_id for canonical naming
    template = AmazonCustomizationTemplate(
        original_filename=file.filename,
        file_path="",  # Will be updated to canonical path after persistence
        file_size=mem_size,
        upload_date=datetime.utcnow(),
    )
    db.add(template)
    db.commit()
    db.refresh(template)

    canonical_path, backup_path = get_customization_template_paths(template.id)

    # Enforce storage policy
    assert_allowed_write_path(canonical_path)
    assert_allowed_write_path(backup_path)

    # Rotate previous canonical -> backup (single backup copy)
    rotate_customization_template_backup(canonical_path, backup_path)

    # Persist to disk
    with open(canonical_path, "wb") as buffer:
        buffer.write(file_bytes)

    disk_size = os.path.getsize(canonical_path)
    with open(canonical_path, "rb") as f:
        persisted_bytes = f.read()
    persisted_sha256 = hashlib.sha256(persisted_bytes).hexdigest()

    if upload_sha256 != persisted_sha256:
        raise HTTPException(status_code=500, detail="Customization template write verification failed (sha256 mismatch).")

    # Update DB with canonical path + final sizes
    template.file_path = canonical_path
    template.file_size = disk_size
    template.upload_date = datetime.utcnow()

    db.commit()
    db.refresh(template)

    print(
        f"[CUSTOMIZATION_UPLOAD] template_id={template.id} "
        f"upload_sha256={upload_sha256} persisted_sha256={persisted_sha256} "
        f"mem_size={mem_size} disk_size={disk_size} path={canonical_path}"
    )

    return {"message": "Upload successful", "id": template.id, "filename": template.original_filename}

@router.get("/amazon-customization-templates")
def list_amazon_customization_templates(db: Session = Depends(get_db)):
    return db.query(AmazonCustomizationTemplate).order_by(desc(AmazonCustomizationTemplate.upload_date)).all()

@router.post("/amazon-customization-templates/{id}/upload")
def update_amazon_customization_template(id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Replace the stored XLSX for an existing Amazon Customization Template.

    Storage policy parity with Product Type Templates:
    - Writes to canonical path for this template_id
    - Single-backup rotation
    - sha256 verification of persisted bytes
    - Updates DB metadata (original_filename is updated for display/download name)
    """
    ensure_storage_dirs_exist()

    template = db.query(AmazonCustomizationTemplate).filter(AmazonCustomizationTemplate.id == id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    file_bytes = file.file.read()
    upload_sha256 = hashlib.sha256(file_bytes).hexdigest()
    mem_size = len(file_bytes)

    canonical_path, backup_path = get_customization_template_paths(template.id)

    # Enforce storage policy
    assert_allowed_write_path(canonical_path)
    assert_allowed_write_path(backup_path)

    # Rotate previous canonical -> backup (single backup copy)
    rotate_customization_template_backup(canonical_path, backup_path)

    # Persist to disk
    with open(canonical_path, "wb") as buffer:
        buffer.write(file_bytes)

    disk_size = os.path.getsize(canonical_path)
    with open(canonical_path, "rb") as f:
        persisted_bytes = f.read()
    persisted_sha256 = hashlib.sha256(persisted_bytes).hexdigest()

    if upload_sha256 != persisted_sha256:
        raise HTTPException(status_code=500, detail="Customization template write verification failed (sha256 mismatch).")

    # Update DB
    template.original_filename = file.filename
    template.file_path = canonical_path
    template.file_size = disk_size
    template.upload_date = datetime.utcnow()

    db.commit()
    db.refresh(template)

    print(
        f"[CUSTOMIZATION_UPDATE] template_id={template.id} "
        f"upload_sha256={upload_sha256} persisted_sha256={persisted_sha256} "
        f"mem_size={mem_size} disk_size={disk_size} path={canonical_path}"
    )

    return {"message": "Update successful", "id": template.id, "filename": template.original_filename}

@router.delete("/amazon-customization-templates/{id}")
def delete_amazon_customization_template(id: int, db: Session = Depends(get_db)):
    """
    Delete an Amazon Customization Template.

    Clean slate deletion:
    - Clears any EquipmentType references (sets amazon_customization_template_id to NULL)
    - Removes any multi-template assignments (join table rows)
    - Deletes ALL on-disk artifacts (legacy file_path, canonical, backup)
    - Deletes DB row
    - Best-effort file deletion (missing files do not fail the request)
    """
    template = db.query(AmazonCustomizationTemplate).filter(AmazonCustomizationTemplate.id == id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Step 1: Clear EquipmentType references (single FK)
    equipment_types = db.query(EquipmentType).filter(
        EquipmentType.amazon_customization_template_id == template.id
    ).all()
    
    for et in equipment_types:
        print(f"[CUSTOMIZATION_DELETE] Clearing assignment from EquipmentType id={et.id} name={et.name}")
        et.amazon_customization_template_id = None
    
    # Step 2: Remove multi-template assignments (join table)
    join_rows = db.query(EquipmentTypeCustomizationTemplate).filter(
        EquipmentTypeCustomizationTemplate.template_id == template.id
    ).all()
    
    for join_row in join_rows:
        print(f"[CUSTOMIZATION_DELETE] Removing multi-template assignment: equipment_type_id={join_row.equipment_type_id} slot={join_row.slot}")
        db.delete(join_row)
    
    db.flush()  # Persist reference clearing before file deletion
    
    # Step 3: Collect all file paths to delete (use set to avoid duplicates)
    paths_to_delete = set()
    
    # Add legacy stored path
    if template.file_path:
        paths_to_delete.add(template.file_path)
    
    # Add canonical and backup paths
    canonical_path, backup_path = get_customization_template_paths(template.id)
    paths_to_delete.add(canonical_path)
    paths_to_delete.add(backup_path)
    
    # Step 4: Delete files (best-effort)
    deleted_count = 0
    for path in paths_to_delete:
        if os.path.exists(path):
            try:
                os.remove(path)
                deleted_count += 1
                print(f"[CUSTOMIZATION_DELETE] Deleted file: {path}")
            except Exception as e:
                print(f"[CUSTOMIZATION_DELETE] Warning: Could not delete {path}: {e}")
        else:
            print(f"[CUSTOMIZATION_DELETE] File not found (skipping): {path}")
    
    # Step 5: Delete DB row
    db.delete(template)
    db.commit()
    
    print(f"[CUSTOMIZATION_DELETE] template_id={id} deleted_files={deleted_count}")
    
    return {"message": "Template deleted", "deleted_files": deleted_count}

@router.get("/amazon-customization-templates/{id}/download")
def download_amazon_customization_template(id: int, db: Session = Depends(get_db)):
    template = db.query(AmazonCustomizationTemplate).filter(AmazonCustomizationTemplate.id == id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
        
    if not os.path.exists(template.file_path):
        raise HTTPException(status_code=404, detail="File missing on disk")
        
    from fastapi.responses import FileResponse
    return FileResponse(template.file_path, filename=template.original_filename)

@router.get("/amazon-customization-templates/{id}/preview", response_model=AmazonCustomizationTemplatePreviewResponse)
def preview_amazon_customization_template(id: int, db: Session = Depends(get_db)):
    template = db.query(AmazonCustomizationTemplate).filter(AmazonCustomizationTemplate.id == id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    if not os.path.exists(template.file_path):
        raise HTTPException(status_code=404, detail="File missing on disk")

    MAX_PREVIEW_ROWS = 50
    MAX_PREVIEW_COLS = 50

    try:
        from openpyxl import load_workbook

        wb = load_workbook(template.file_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        sheet_name = ws.title

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        row_limit = min(max_row, MAX_PREVIEW_ROWS)
        col_limit = min(max_col, MAX_PREVIEW_COLS)

        grid: List[List[str]] = []

        # Use explicit bounds so we don't accidentally iterate huge sheets
        for r in range(1, row_limit + 1):
            row_values: List[str] = []
            for c in range(1, col_limit + 1):
                v = ws.cell(row=r, column=c).value
                row_values.append("" if v is None else str(v))
            grid.append(row_values)

        wb.close()

        return AmazonCustomizationTemplatePreviewResponse(
            template_id=template.id,
            original_filename=template.original_filename,
            sheet_name=sheet_name,
            max_row=max_row,
            max_column=max_col,
            preview_row_count=row_limit,
            preview_column_count=col_limit,
            grid=grid,
        )
    except HTTPException:
        raise
    except Exception as e:
        # Keep it consistent with existing patterns: return 500 with message
        raise HTTPException(status_code=500, detail=f"Failed to preview customization template: {str(e)}")

@router.post("/equipment-types/{id}/amazon-customization-template/assign", response_model=EquipmentTypeResponse)
def assign_amazon_customization_template(id: int, data: AmazonCustomizationTemplateAssignmentRequest, db: Session = Depends(get_db)):
    """
    Legacy single-template assignment endpoint (BACKWARD COMPATIBLE).
    
    Sets EquipmentType.amazon_customization_template_id (the "primary" template).
    Also maintains coherence with multi-template system by upserting slot 1.
    """
    equipment_type = db.query(EquipmentType).filter(EquipmentType.id == id).first()
    if not equipment_type:
        raise HTTPException(status_code=404, detail="Equipment type not found")
    
    if data.template_id is not None:
        template = db.query(AmazonCustomizationTemplate).filter(AmazonCustomizationTemplate.id == data.template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Amazon Customization Template not found")
    
    # Set single FK (existing behavior - UNCHANGED)
    equipment_type.amazon_customization_template_id = data.template_id
    
    # BACKWARD COMPATIBILITY: Also upsert join table slot 1 to keep systems coherent
    if data.template_id is not None:
        # Check if slot 1 already has a different template
        existing_slot_1 = db.query(EquipmentTypeCustomizationTemplate).filter(
            EquipmentTypeCustomizationTemplate.equipment_type_id == equipment_type.id,
            EquipmentTypeCustomizationTemplate.slot == 1
        ).first()
        
        if existing_slot_1:
            if existing_slot_1.template_id != data.template_id:
                # Update slot 1 to new template
                existing_slot_1.template_id = data.template_id
                existing_slot_1.created_at = datetime.utcnow()
                print(f"[LEGACY_ASSIGN] Updated slot 1 for equipment_type_id={equipment_type.id} to template_id={data.template_id}")
        else:
            # Create new slot 1 assignment
            new_assignment = EquipmentTypeCustomizationTemplate(
                equipment_type_id=equipment_type.id,
                template_id=data.template_id,
                slot=1,
                created_at=datetime.utcnow()
            )
            db.add(new_assignment)
            print(f"[LEGACY_ASSIGN] Created slot 1 for equipment_type_id={equipment_type.id} with template_id={data.template_id}")
    else:
        # If unsetting the primary template, also remove slot 1
        existing_slot_1 = db.query(EquipmentTypeCustomizationTemplate).filter(
            EquipmentTypeCustomizationTemplate.equipment_type_id == equipment_type.id,
            EquipmentTypeCustomizationTemplate.slot == 1
        ).first()
        
        if existing_slot_1:
            db.delete(existing_slot_1)
            print(f"[LEGACY_ASSIGN] Removed slot 1 for equipment_type_id={equipment_type.id}")
            
    db.commit()
    db.refresh(equipment_type)
    return equipment_type

# ------------------------------------------------------------------
# 9. Multi-Template Assignment (Slot-based, up to 3 templates)
# ------------------------------------------------------------------

@router.get("/equipment-types/{equipment_type_id}/amazon-customization-templates", response_model=EquipmentTypeCustomizationTemplatesResponse)
def list_equipment_type_customization_templates(equipment_type_id: int, db: Session = Depends(get_db)):
    """
    List all customization templates assigned to an equipment type (up to 3 slots).
    Includes default_template_id to indicate which template is the default.
    """
    equipment_type = db.query(EquipmentType).filter(EquipmentType.id == equipment_type_id).first()
    if not equipment_type:
        raise HTTPException(status_code=404, detail="Equipment type not found")
    
    # Query join table for all assignments
    assignments = db.query(EquipmentTypeCustomizationTemplate).filter(
        EquipmentTypeCustomizationTemplate.equipment_type_id == equipment_type_id
    ).order_by(EquipmentTypeCustomizationTemplate.slot).all()
    
    # Build response with template details
    template_items = []
    for assignment in assignments:
        template = db.query(AmazonCustomizationTemplate).filter(
            AmazonCustomizationTemplate.id == assignment.template_id
        ).first()
        
        if template:  # Safety check
            template_items.append(EquipmentTypeCustomizationTemplateItem(
                template_id=template.id,
                slot=assignment.slot,
                original_filename=template.original_filename,
                upload_date=template.upload_date
            ))
    
    return EquipmentTypeCustomizationTemplatesResponse(
        equipment_type_id=equipment_type_id,
        templates=template_items,
        default_template_id=equipment_type.amazon_customization_template_id
    )

@router.post("/equipment-types/{equipment_type_id}/amazon-customization-templates/assign", response_model=EquipmentTypeCustomizationTemplatesResponse)
def assign_equipment_type_customization_template(
    equipment_type_id: int,
    data: EquipmentTypeCustomizationTemplateAssignRequest,
    db: Session = Depends(get_db)
):
    """
    Assign a customization template to a specific slot (1-3) for an equipment type.
    
    - Validates slot is 1, 2, or 3
    - Validates both equipment type and template exist
    - If slot already occupied, replaces it (upsert behavior)
    - Prevents duplicate template assignments across slots
    """
    # Validate equipment type exists
    equipment_type = db.query(EquipmentType).filter(EquipmentType.id == equipment_type_id).first()
    if not equipment_type:
        raise HTTPException(status_code=404, detail="Equipment type not found")
    
    # Validate template exists
    template = db.query(AmazonCustomizationTemplate).filter(AmazonCustomizationTemplate.id == data.template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Validate slot is 1, 2, or 3
    if data.slot not in [1, 2, 3]:
        raise HTTPException(status_code=400, detail="Slot must be 1, 2, or 3")
    
    # Check if this template is already assigned to a different slot for this equipment type
    existing_different_slot = db.query(EquipmentTypeCustomizationTemplate).filter(
        EquipmentTypeCustomizationTemplate.equipment_type_id == equipment_type_id,
        EquipmentTypeCustomizationTemplate.template_id == data.template_id,
        EquipmentTypeCustomizationTemplate.slot != data.slot
    ).first()
    
    if existing_different_slot:
        raise HTTPException(
            status_code=400,
            detail=f"Template {data.template_id} is already assigned to slot {existing_different_slot.slot} for this equipment type"
        )
    
    # Check if slot already has an assignment (upsert behavior)
    existing_slot = db.query(EquipmentTypeCustomizationTemplate).filter(
        EquipmentTypeCustomizationTemplate.equipment_type_id == equipment_type_id,
        EquipmentTypeCustomizationTemplate.slot == data.slot
    ).first()
    
    if existing_slot:
        # Update existing slot
        existing_slot.template_id = data.template_id
        existing_slot.created_at = datetime.utcnow()
        print(f"[MULTI_ASSIGN] Updated equipment_type_id={equipment_type_id} slot={data.slot} to template_id={data.template_id}")
    else:
        # Create new assignment
        new_assignment = EquipmentTypeCustomizationTemplate(
            equipment_type_id=equipment_type_id,
            template_id=data.template_id,
            slot=data.slot,
            created_at=datetime.utcnow()
        )
        db.add(new_assignment)
        print(f"[MULTI_ASSIGN] Created equipment_type_id={equipment_type_id} slot={data.slot} with template_id={data.template_id}")
    
    db.commit()
    
    # Return updated list of all assignments
    return list_equipment_type_customization_templates(equipment_type_id, db)

@router.post("/equipment-types/{equipment_type_id}/amazon-customization-templates/default")
def set_equipment_type_customization_template_default(
    equipment_type_id: int,
    data: EquipmentTypeCustomizationTemplateSetDefaultRequest,
    db: Session = Depends(get_db)
):
    """
    Set one of the assigned templates as the default for this equipment type.
    
    - Validates equipment type exists
    - Validates template exists
    - Validates template is currently assigned to this equipment type (in join table)
    - Sets equipment_type.amazon_customization_template_id to the specified template
    - Export behavior continues to use amazon_customization_template_id
    """
    # Validate equipment type exists
    equipment_type = db.query(EquipmentType).filter(EquipmentType.id == equipment_type_id).first()
    if not equipment_type:
        raise HTTPException(status_code=404, detail="Equipment type not found")
    
    # Validate template exists
    template = db.query(AmazonCustomizationTemplate).filter(AmazonCustomizationTemplate.id == data.template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Validate template is assigned to this equipment type (must be in join table)
    assignment = db.query(EquipmentTypeCustomizationTemplate).filter(
        EquipmentTypeCustomizationTemplate.equipment_type_id == equipment_type_id,
        EquipmentTypeCustomizationTemplate.template_id == data.template_id
    ).first()
    
    if not assignment:
        raise HTTPException(status_code=400, detail="Template is not assigned to this equipment type")
    
    # Set the default (single FK)
    equipment_type.amazon_customization_template_id = data.template_id
    db.commit()
    
    print(f"[SET_DEFAULT] equipment_type_id={equipment_type_id} default_template_id={data.template_id} slot={assignment.slot}")
    
    return {
        "message": "Default template set",
        "equipment_type_id": equipment_type_id,
        "template_id": data.template_id
    }

@router.delete("/equipment-types/{equipment_type_id}/amazon-customization-templates/{template_id}")
def unassign_equipment_type_customization_template(equipment_type_id: int, template_id: int, db: Session = Depends(get_db)):
    """
    Remove a specific template assignment from an equipment type.
    
    Deletes the join table row(s) for this equipment type + template combination.
    If the template being unassigned is the current default, clears the default.
    """
    # Validate equipment type exists
    equipment_type = db.query(EquipmentType).filter(EquipmentType.id == equipment_type_id).first()
    if not equipment_type:
        raise HTTPException(status_code=404, detail="Equipment type not found")
    
    # Find and delete the assignment(s)
    assignments = db.query(EquipmentTypeCustomizationTemplate).filter(
        EquipmentTypeCustomizationTemplate.equipment_type_id == equipment_type_id,
        EquipmentTypeCustomizationTemplate.template_id == template_id
    ).all()
    
    if not assignments:
        raise HTTPException(status_code=404, detail="Template assignment not found for this equipment type")
    
    # If this template is the current default, clear it
    if equipment_type.amazon_customization_template_id == template_id:
        equipment_type.amazon_customization_template_id = None
        print(f"[MULTI_UNASSIGN] Cleared default template for equipment_type_id={equipment_type_id}")
    
    for assignment in assignments:
        print(f"[MULTI_UNASSIGN] Removing equipment_type_id={equipment_type_id} template_id={template_id} slot={assignment.slot}")
        db.delete(assignment)
    
    db.commit()
    
    return {"message": "Template unassigned", "removed_slots": [a.slot for a in assignments]}


@router.post("/equipment-types/{id}/reverb-template/assign", response_model=EquipmentTypeResponse)
def assign_reverb_template_to_equipment_type(id: int, data: ReverbTemplateAssignmentRequest, db: Session = Depends(get_db)):
    """
    Assign a Reverb Template to an Equipment Type.
    """
    equipment_type = db.query(EquipmentType).filter(EquipmentType.id == id).first()
    if not equipment_type:
        raise HTTPException(status_code=404, detail="Equipment type not found")
    
    if data.template_id is not None:
        template = db.query(ReverbTemplate).filter(ReverbTemplate.id == data.template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Reverb Template not found")
    
    equipment_type.reverb_template_id = data.template_id
    db.commit()
    db.refresh(equipment_type)
    return equipment_type
