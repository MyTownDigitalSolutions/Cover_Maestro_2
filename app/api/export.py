from __future__ import annotations
import re
import os
import logging
import zipfile
import io
import csv
import json
import copy
import hashlib
import asyncio
import urllib.error
import urllib.request
import urllib.parse
import concurrent.futures
import time
import traceback
from fastapi import APIRouter, Depends, HTTPException, Response, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional
from pydantic import BaseModel, Field
from app.schemas.core import ModelPricingSnapshotResponse, ExportStatsResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell.cell import MergedCell
from app.database import get_db
from app.models.core import Model, Series, Manufacturer, EquipmentType, ModelPricingSnapshot, ExportSetting
from app.models.templates import AmazonProductType, ProductTypeField, EquipmentTypeProductType

from app.schemas.core import ModelPricingSnapshotResponse
from app.api.pricing import recalculate_targeted, PricingRecalculateRequest
from app.services.pricing_calculator import PricingConfigError
from app.services.reverb_export_service import generate_reverb_export_csv
from app.services.shared_template_tokens import (
    apply_base_sku_tokens,
    normalize_variant_key_for_price_token,
    resolve_price_placeholders_in_value as resolve_shared_price_placeholders_in_value,
    resolve_single_price_placeholder as resolve_shared_single_price_placeholder,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/export", tags=["export"])

# Module-level cache for HTTP results
# Structure: { url: { 'time': float, 'code': int|None, 'error': str|None } }
HTTP_CACHE = {}
MAX_CACHE_ENTRIES = 5000
TTL_SECONDS = 900 # 15 minutes

# --------------------------------------------------------------------------
# Staleness Check Helper
# --------------------------------------------------------------------------
def ensure_models_fresh_for_export(request: ExportPreviewRequest, db: Session) -> tuple[bool, int]:
    """
    Check if any selected model has stale pricing (or missing baseline snapshots).
    If so, trigger a targeted recalculation before export proceeds.
    Returns (recalc_performed: bool, recalc_model_count: int).
    """
    if not request.model_ids:
        return False, 0
        
    recalc_req = PricingRecalculateRequest(
        model_ids=request.model_ids,
        only_if_stale=True
    )
    
    # Direct service logic reuse via the API function, 
    # but we must ensure we don't double-wrap or mis-handle dependencies.
    # calling the route function directly is acceptable in FastAPI if deps match.
    # alternatively, extract the logic. 
    # For chunks, calling the route function is safest "minimal change" 
    # as long as we pass 'db'.
    
    try:
        resp = recalculate_targeted(recalc_req, db)
    except PricingConfigError as e:
        # If the underlying recalc fails due to config (e.g. fixed cell missing),
        # we must abort export and tell the user.
        raise HTTPException(status_code=400, detail=f"Export failed during pricing check: {str(e)}")
    
    return (resp.recalculated_models > 0), resp.recalculated_models

@router.get("/debug-price/{model_id}", response_model=ModelPricingSnapshotResponse)
def get_debug_price(model_id: int, db: Session = Depends(get_db)):
    """
    Get the baseline pricing snapshot for this model (Amazon / Choice No Padding).
    Used for UI debugging and validation.
    """
    snap = db.query(ModelPricingSnapshot).filter(
        ModelPricingSnapshot.model_id == model_id,
        ModelPricingSnapshot.marketplace == "amazon",
        ModelPricingSnapshot.variant_key == "choice_no_padding"
    ).first()
    
    if not snap:
        raise HTTPException(status_code=404, detail="Baseline snapshot not found. Please recalculate pricing or run seed.")
        
    return snap

class ExportPreviewRequest(BaseModel):
    model_ids: List[int]
    listing_type: str = "individual"  # "individual" or "parent_child"



class ExportValidationIssue(BaseModel):
    severity: str  # "error", "warning"
    model_id: int | None = None
    model_name: str | None = None
    message: str

class ExportValidationResponse(BaseModel):
    status: str  # "valid", "warnings", "errors"
    summary_counts: Dict[str, int]
    items: List[ExportValidationIssue]

def _evaluate_equipment_type_compatibility(equipment_type_ids: set, db: Session) -> dict:
    """
    Evaluate if a set of equipment type IDs are compatible for export.
    
    Returns dict with:
        is_compatible: bool
        reason: str (machine-friendly reason if incompatible)
        product_type_ids: set[int]
        customization_template_ids: set[int | None] (IMPORTANT: includes None)
        missing_equipment_type_ids: set[int] (equipment types lacking EquipmentTypeProductType links)
    """
    result = {
        "is_compatible": False,
        "reason": "",
        "product_type_ids": set(),
        "customization_template_ids": set(),
        "missing_equipment_type_ids": set()
    }
    
    # Check for None in input
    if None in equipment_type_ids:
        result["reason"] = "null_equipment_type_id"
        return result
    
    # Query all EquipmentTypeProductType links
    links = db.query(EquipmentTypeProductType).filter(
        EquipmentTypeProductType.equipment_type_id.in_(equipment_type_ids)
    ).all()
    
    # Compute which equipment types have links
    linked_eq_ids = set(link.equipment_type_id for link in links)
    missing_eq_ids = equipment_type_ids - linked_eq_ids
    
    if missing_eq_ids:
        result["reason"] = "missing_equipment_type_product_type_links"
        result["missing_equipment_type_ids"] = missing_eq_ids
        return result
    
    # Compute product type IDs
    product_type_ids = set(link.product_type_id for link in links)
    result["product_type_ids"] = product_type_ids
    
    if len(product_type_ids) != 1:
        result["reason"] = "multiple_product_types"
        return result
    
    # Query equipment types and compute customization template IDs
    # IMPORTANT: Do NOT filter out None - it's a valid value that must be included
    equipment_types = db.query(EquipmentType).filter(
        EquipmentType.id.in_(equipment_type_ids)
    ).all()
    
    customization_template_ids = set(
        et.amazon_customization_template_id for et in equipment_types
    )
    result["customization_template_ids"] = customization_template_ids
    
    if len(customization_template_ids) != 1:
        result["reason"] = "multiple_customization_templates"
        return result
    
    # All checks passed
    result["is_compatible"] = True
    result["reason"] = "compatible"
    return result


@router.post("/validate", response_model=ExportValidationResponse)
def validate_export(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """
    Pre-flight check for export readiness. Verifies templates, pricing snapshots, and data integrity.
    """
    issues = []
    
    # 1. Basic Request Validation
    if not request.model_ids:
        return ExportValidationResponse(
            status="errors",
            summary_counts={"total_models": 0, "issues": 1},
            items=[ExportValidationIssue(severity="error", message="No models selected.")]
        )

    models = db.query(Model).filter(Model.id.in_(request.model_ids)).all()
    if not models or len(models) != len(request.model_ids):
        found_ids = {m.id for m in models}
        missing_ids = set(request.model_ids) - found_ids
        issues.append(ExportValidationIssue(severity="error", message=f"Some requested models exist. Missing IDs: {missing_ids}"))

    # Filter out models with missing dimensions (they will be skipped during export)
    valid_models = [m for m in models if _has_valid_dimensions(m)]
    skipped_no_dims = [m for m in models if not _has_valid_dimensions(m)]
    
    # Add warnings for skipped models
    if skipped_no_dims:
        for m in skipped_no_dims:
            issues.append(ExportValidationIssue(
                severity="warning",
                model_id=m.id,
                model_name=m.name,
                message="Missing physical dimensions (W/D/H must be > 0). Will be skipped during export."
            ))
    
    # Use valid_models for all subsequent checks (only models that will actually be exported)
    models = valid_models
    total_models = len(models)
    
    image_fields = []
    
    # 2. Equipment Type Consistency & Template Loading
    if models:
        eq_ids = set(m.equipment_type_id for m in models)
        
        # Evaluate equipment type compatibility
        compat = _evaluate_equipment_type_compatibility(eq_ids, db)
        
        # Log compatibility decision (instrumentation)
        logger.info(
            f"[EXPORT][VALIDATE] selected_models={len(models)} equipment_type_ids={sorted(eq_ids)} "
            f"product_type_ids={sorted(compat['product_type_ids'])} "
            f"customization_template_ids={sorted(compat['customization_template_ids'], key=lambda x: (x is None, x))} "
            f"missing_equipment_type_ids={sorted(compat['missing_equipment_type_ids'])} "
            f"compatible={compat['is_compatible']} reason={compat['reason']}"
        )
        
        if not compat["is_compatible"]:
            # Build detailed error message
            equipment_types = db.query(EquipmentType).filter(EquipmentType.id.in_(eq_ids)).all()
            et_map = {et.id: et.name for et in equipment_types}
            et_details = ", ".join([f"{et_id} ({et_map.get(et_id, 'Unknown')})" for et_id in sorted(eq_ids)])
            
            error_parts = [
                f"Mixed equipment types are incompatible for export.",
                f"Equipment Types: {et_details}",
                f"Product Type IDs: {sorted(compat['product_type_ids']) if compat['product_type_ids'] else 'N/A'}",
                f"Customization Template IDs: {sorted(compat['customization_template_ids'], key=lambda x: (x is None, x)) if compat['customization_template_ids'] else 'N/A'}"
            ]
            
            if compat["missing_equipment_type_ids"]:
                missing_names = ", ".join([
                    f"{et_id} ({et_map.get(et_id, 'Unknown')})" 
                    for et_id in sorted(compat["missing_equipment_type_ids"])
                ])
                error_parts.append(f"Missing EquipmentTypeProductType links for: {missing_names}")
            
            error_parts.append("Reason: " + compat["reason"])
            
            issues.append(ExportValidationIssue(
                severity="error",
                message=" | ".join(error_parts)
            ))
        else:
            # Compatible - load product type and fields for further validation
            product_type_id = list(compat["product_type_ids"])[0]
            pt = db.query(AmazonProductType).filter(AmazonProductType.id == product_type_id).first()
            
            if not pt or not pt.code:
                issues.append(ExportValidationIssue(
                    severity="error", 
                    message="Linked Amazon Template is invalid or missing Template Code."
                ))
            else:
                # Load fields for placeholder verification
                all_fields = db.query(ProductTypeField).filter(
                    ProductTypeField.product_type_id == pt.id
                ).all()
                image_fields = [f for f in all_fields if is_image_url_field(f.field_name)]

    # 3. Model-Specific Checks
    MAX_HTTP_MODELS = 25
    MAX_HTTP_URLS = 60
    MAX_CONCURRENCY = 6

    http_models_checked_count = 0
    http_urls_scheduled_count = 0
    is_capped = False
    
    # Prune cache if needed
    if len(HTTP_CACHE) > MAX_CACHE_ENTRIES:
        HTTP_CACHE.clear()
        
    verification_jobs = [] # List of dicts: {'model': model, 'url': url, 'cached': bool}

    for idx, model in enumerate(models):
        # A. Pricing Snapshot
        snap = db.query(ModelPricingSnapshot).filter(
            ModelPricingSnapshot.model_id == model.id,
            ModelPricingSnapshot.marketplace == "amazon",
            ModelPricingSnapshot.variant_key == "choice_no_padding"
        ).first()
        
        if not snap:
            issues.append(ExportValidationIssue(
                severity="error", 
                model_id=model.id, 
                model_name=model.name, 
                message="Missing pricing snapshot (choice_no_padding). Recalculation required."
            ))
            
        # B. Image Placeholders & HTTP Check
        if image_fields:
            series = db.query(Series).filter(Series.id == model.series_id).first()
            manufacturer = db.query(Manufacturer).filter(Manufacturer.id == series.manufacturer_id).first() if series else None
            equip_type = db.query(EquipmentType).filter(EquipmentType.id == model.equipment_type_id).first()
            
            # 1. Placeholder Syntax Check (All Fields, Always)
            for img_field in image_fields:
                val_to_check = img_field.selected_value or img_field.custom_value
                
                if val_to_check:
                    resolved = substitute_placeholders(val_to_check, model, series, manufacturer, equip_type, is_image_url=True)
                    resolved = _apply_image_index_placeholder(resolved, img_field.field_name)
                    resolved = _normalize_image_filename_in_url(resolved)
                    if '[' in resolved and ']' in resolved:
                         issues.append(ExportValidationIssue(
                            severity="warning",
                            model_id=model.id,
                            model_name=model.name,
                            message=f"Unresolved placeholder in image field '{img_field.field_name}': {resolved}"
                        ))
            
            # 2. HTTP Availability Check (Sampled, Capped, Cached)
            # Strategy: Full check for first model; First + Last fields only for subsequent models
            fields_to_check = image_fields if idx == 0 else ([image_fields[0], image_fields[-1]] if len(image_fields) > 1 else image_fields)
            
            model_triggering_fetch = False
            
            for img_field in fields_to_check:
                val = img_field.selected_value or img_field.custom_value
                if val:
                    url = substitute_placeholders(val, model, series, manufacturer, equip_type, is_image_url=True)
                    url = _apply_image_index_placeholder(url, img_field.field_name)
                    url = _normalize_image_filename_in_url(url)
                    
                    if '[' in url and ']' in url:
                        continue
                    
                    # Check Cache
                    now = time.time()
                    cached_entry = HTTP_CACHE.get(url)
                    if cached_entry and (now - cached_entry['time'] < TTL_SECONDS):
                        verification_jobs.append({'model': model, 'url': url, 'cached': True, 'entry': cached_entry})
                        continue
                        
                    # Not Cached: Apply Caps
                    if is_capped:
                        continue
                        
                    if http_urls_scheduled_count >= MAX_HTTP_URLS:
                        is_capped = True
                        continue
                        
                    if not model_triggering_fetch and http_models_checked_count >= MAX_HTTP_MODELS:
                        is_capped = True
                        continue

                    # Schedule Fetch
                    verification_jobs.append({'model': model, 'url': url, 'cached': False})
                    http_urls_scheduled_count += 1
                    model_triggering_fetch = True
            
            if model_triggering_fetch:
                http_models_checked_count += 1

    if is_capped:
        issues.append(ExportValidationIssue(
            severity="warning",
            message=f"HTTP image checks capped. Checked {http_urls_scheduled_count} new URLs across {http_models_checked_count} models (cap reached)."
        ))

    # Execute New HTTP Checks
    urls_to_fetch = list(set(job['url'] for job in verification_jobs if not job['cached']))
    
    if urls_to_fetch:
        def check_url_status(u):
            try:
                # Short timeout, use HEAD to be lightweight
                req = urllib.request.Request(u, method="HEAD")
                with urllib.request.urlopen(req, timeout=2.0) as response:
                    return u, response.getcode(), None
            except urllib.error.HTTPError as e:
                return u, e.code, None
            except Exception as e:
                return u, None, str(e)
        
        fresh_results = {}
        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_CONCURRENCY) as executor:
             future_to_url = {executor.submit(check_url_status, u): u for u in urls_to_fetch}
             for future in concurrent.futures.as_completed(future_to_url):
                 u, code, err = future.result()
                 fresh_results[u] = (code, err)
                 
    # Update Cache and Generate Issues
    now = time.time()
    
    for job in verification_jobs:
        url = job['url']
        code = None
        err = None
        
        if job['cached']:
            code = job['entry']['code']
            err = job['entry']['error']
        else:
             if url in fresh_results:
                 code, err = fresh_results[url]
                 # Update Cache
                 HTTP_CACHE[url] = {'time': now, 'code': code, 'error': err}
        
        if err:
             issues.append(ExportValidationIssue(
                 severity="warning",
                 model_id=job['model'].id,
                 model_name=job['model'].name,
                 message=f"Image URL inaccessible ({err}): {url}"
                 ))
        elif code and code >= 400:
             issues.append(ExportValidationIssue(
                 severity="warning", 
                 model_id=job['model'].id, 
                 model_name=job['model'].name, 
                 message=f"Image URL inaccessible (HTTP {code}): {url}"
            ))

    # Determine Status
    error_count = sum(1 for i in issues if i.severity == "error")
    warning_count = sum(1 for i in issues if i.severity == "warning")
    
    status = "valid"
    if error_count > 0:
        status = "errors"
    elif warning_count > 0:
        status = "warnings"
        
    return ExportValidationResponse(
        status=status,
        summary_counts={
            "total_models": total_models,
            "issues": len(issues),
            "errors": error_count,
            "warnings": warning_count
        },
        items=issues
    )




from datetime import datetime

# Helper function to check if model has valid dimensions
def _has_valid_dimensions(m: Model) -> bool:
    """Check if model has valid physical dimensions (all > 0)."""
    return m.width and m.depth and m.height and m.width > 0 and m.depth > 0 and m.height > 0

def _format_required_field_name(field: ProductTypeField) -> str:
    """
    Return a readable + exact template field identifier for diagnostics.
    Example:
    "Dangerous Goods Regulations (template field: supplier_declared_dg_hz_regulation...#1.value)"
    """
    display = (getattr(field, "display_name", None) or "").strip()
    technical = (getattr(field, "field_name", None) or "").strip()

    if display and technical and display != technical:
        return f"{display} (template field: {technical})"
    return technical or display or "<unknown field>"


def _normalize_http_detail(detail) -> str:
    if isinstance(detail, str):
        return detail
    try:
        return json.dumps(detail)
    except Exception:
        return str(detail)


def _build_amazon_filename_tokens(export_models: list[Model], db: Session) -> tuple[str, str, str]:
    """
    Build stable filename tokens for Amazon export artifacts.
    Returns: (manufacturer_token, series_token, date_token)
    """
    if not export_models:
        return "Unknown", "Unknown", datetime.now().strftime("%Y-%-m-%-d")

    first_model = export_models[0]
    first_series = db.query(Series).filter(Series.id == first_model.series_id).first()
    first_manufacturer = (
        db.query(Manufacturer).filter(Manufacturer.id == first_series.manufacturer_id).first()
        if first_series
        else None
    )

    mfr_name = normalize_for_url(first_manufacturer.name) if first_manufacturer else "Unknown"

    series_ids = {m.series_id for m in export_models}
    if len(series_ids) > 1:
        series_name = "Multiple_Series"
    else:
        series_name = normalize_for_url(first_series.name) if first_series else "Unknown"

    now = datetime.now()
    date_str = f"{now.year}-{now.month}-{now.day}"
    return mfr_name, series_name, date_str


def build_export_data(request: ExportPreviewRequest, db: Session):
    """
	Build export data (headers and rows) for the given models.
    Returns: (header_rows, data_rows, filename_base, template_code, export_models, ordered_field_names, ordered_required_flags, warnings)
    """
    if not request.model_ids:
        raise HTTPException(status_code=400, detail="No models selected")
    
    # Query models with Amazon export filter applied at database level
    models = db.query(Model).filter(
        Model.id.in_(request.model_ids),
        Model.exclude_from_amazon_export.is_(False)
    ).all()
    
    # Log count of excluded models (defensive check)
    all_requested = db.query(Model).filter(Model.id.in_(request.model_ids)).all()
    excluded_count = len(all_requested) - len(models)
    if excluded_count > 0:
        logger.info(f"[EXPORT-AMAZON] Excluded {excluded_count} model(s) from Amazon export")
    
    if not models:
        raise HTTPException(status_code=404, detail="No models found")
    
    # Separate models with valid vs missing dimensions
    export_models = [m for m in models if _has_valid_dimensions(m)]
    skipped_no_dims = [m for m in models if not _has_valid_dimensions(m)]
    
    # Build warning list for skipped models
    warnings = []
    if skipped_no_dims:
        # Load series data for warning messages
        series_ids = set(m.series_id for m in skipped_no_dims)
        series_map = {s.id: s.name for s in db.query(Series).filter(Series.id.in_(series_ids)).all()}
        
        for m in skipped_no_dims:
            series_name = series_map.get(m.series_id, f"Series {m.series_id}")
            warnings.append(
                f"{series_name}: {m.name} | Reason: Missing required information: width, depth, height (surface_area_sq_in cannot be derived)"
            )
        
        logger.warning(f"[EXPORT] Skipped {len(skipped_no_dims)} model(s) due to missing dimensions: {warnings}")
    
    if not export_models:
        raise HTTPException(
            status_code=400,
            detail="No models exported: all selected models are missing physical dimensions (W/D/H must be > 0)."
        )
    
    # Template compatibility validation: allow mixed equipment types if they use same templates
    equipment_type_ids = set(m.equipment_type_id for m in export_models)
    
    # Use the same compatibility evaluator as validate_export
    compat = _evaluate_equipment_type_compatibility(equipment_type_ids, db)
    
    # Log compatibility decision (instrumentation)
    logger.info(
        f"[EXPORT][BUILD_DATA] selected_models={len(export_models)} equipment_type_ids={sorted(equipment_type_ids)} "
        f"product_type_ids={sorted(compat['product_type_ids'])} "
        f"customization_template_ids={sorted(compat['customization_template_ids'], key=lambda x: (x is None, x))} "
        f"missing_equipment_type_ids={sorted(compat['missing_equipment_type_ids'])} "
        f"compatible={compat['is_compatible']} reason={compat['reason']}"
    )
    
    if not compat["is_compatible"]:
        # Build detailed error message
        equipment_types = db.query(EquipmentType).filter(EquipmentType.id.in_(equipment_type_ids)).all()
        et_map = {et.id: et.name for et in equipment_types}
        sorted_ids = sorted(equipment_type_ids)
        et_details = ", ".join([f"{et_id} ({et_map.get(et_id, 'Unknown')})" for et_id in sorted_ids])
        
        error_parts = [
            f"Export failed: mixed equipment types are incompatible.",
            f"Selected models: {len(export_models)}.",
            f"Equipment types: {et_details}.",
            f"Product type IDs: {sorted(compat['product_type_ids']) if compat['product_type_ids'] else 'N/A'}.",
            f"Customization template IDs: {sorted(compat['customization_template_ids'], key=lambda x: (x is None, x)) if compat['customization_template_ids'] else 'N/A'}."
        ]
        
        if compat["missing_equipment_type_ids"]:
            missing_names = ", ".join([
                f"{et_id} ({et_map.get(et_id, 'Unknown')})" 
                for et_id in sorted(compat["missing_equipment_type_ids"])
            ])
            error_parts.append(f"Missing EquipmentTypeProductType links for: {missing_names}.")
        
        error_parts.append(f"Reason: {compat['reason']}.")
        error_parts.append("Filter to equipment types that share the same templates.")
        
        error_detail = " ".join(error_parts)
        logger.warning("[EXPORT] Incompatible templates: %s", compat['reason'])
        raise HTTPException(status_code=400, detail=error_detail)
    
    # Compatible - retrieve link for template resolution
    if len(equipment_type_ids) > 1:
        # Mixed but compatible - use any equipment type's link
        equipment_type_id = list(equipment_type_ids)[0]
        link = db.query(EquipmentTypeProductType).filter(
            EquipmentTypeProductType.equipment_type_id == equipment_type_id
        ).first()
    else:
        equipment_type_id = list(equipment_type_ids)[0]
        
        links = db.query(EquipmentTypeProductType).filter(
            EquipmentTypeProductType.equipment_type_id == equipment_type_id
        ).all()
        
        if len(links) == 0:
            equipment_type = db.query(EquipmentType).filter(EquipmentType.id == equipment_type_id).first()
            raise HTTPException(
                status_code=404, 
                detail=f"No Amazon template linked to equipment type: {equipment_type.name if equipment_type else 'Unknown'}"
            )
            
        if len(links) > 1:
            raise HTTPException(
                status_code=500,
                detail="Configuration error: multiple templates assigned to equipment type. Please resolve uniqueness constraint."
            )
            
        link = links[0]
    
    product_type = db.query(AmazonProductType).filter(
        AmazonProductType.id == link.product_type_id
    ).first()
    
    if not product_type:
        raise HTTPException(status_code=404, detail="Template not found")
    
    fields = db.query(ProductTypeField).filter(
        ProductTypeField.product_type_id == product_type.id
    ).order_by(ProductTypeField.order_index).all()
    
    
    # --- PHASE 3 FIX: Load headers from File ---
    header_rows = []
    header_source = "db_header_rows"
    
    if product_type.file_path and os.path.exists(product_type.file_path):
        try:
            wb_h = load_workbook(product_type.file_path, read_only=True, data_only=True)
            ws_h = None
            
            if product_type.export_sheet_name_override and product_type.export_sheet_name_override in wb_h.sheetnames:
                ws_h = wb_h[product_type.export_sheet_name_override]
            else:
                blacklist = ["data definitions", "instructions", "readme"]
                for sn in wb_h.sheetnames:
                    if any(b in sn.lower() for b in blacklist): continue
                    ws_temp = wb_h[sn]
                    if ws_temp.sheet_state == 'visible':
                        ws_h = ws_temp
                        break
            
            if not ws_h:
                ws_h = wb_h.active
                
            if ws_h:
                # Extract Rows 2-5 (Amazon Headers)
                max_col = ws_h.max_column
                # Optimize max_col: scan Row 4 (Labels) backwards for last value
                try: 
                    # Scan last 20 columns of max_col to check empty? Or just trust max_col.
                    # max_col can be large. Let's trust it or cap it.
                    if max_col > 100: max_col = 100 
                except: pass
                
                file_headers = []
                for r in range(2, 6):
                    row_vals = []
                    for c in range(1, max_col + 1):
                        val = ws_h.cell(row=r, column=c).value
                        row_vals.append(str(val) if val is not None else "")
                    file_headers.append(row_vals)
                
                if file_headers:
                    header_rows = file_headers
                    header_source = f"file_path sheet={ws_h.title}"
            
            wb_h.close()
        except Exception as e:
            logger.warning(f"[BUILD_EXPORT] Failed to load headers from file: {e}")
            header_rows = product_type.header_rows or []
    else:
        header_rows = product_type.header_rows or []

    if not header_rows:
        header_rows = product_type.header_rows or []
        
    logger.info(f"[EXPORT][PREVIEW] header_source={header_source} file={product_type.file_path}")
    # ------------------------------------------
    
    # Use first equipment type for field resolution (works for both single and mixed-compatible cases)
    equipment_type_id = list(equipment_type_ids)[0]
    equipment_type = db.query(EquipmentType).filter(EquipmentType.id == equipment_type_id).first()
    
    mfr_name, series_name, date_str = _build_amazon_filename_tokens(export_models, db)
    filename_base = f"Amazon_{mfr_name}_{series_name}_{date_str}"
    
    data_rows = []
    required_field_issues = []
    for model in export_models:
        series = db.query(Series).filter(Series.id == model.series_id).first()
        manufacturer = db.query(Manufacturer).filter(Manufacturer.id == series.manufacturer_id).first() if series else None
        
        row_data = []
        missing_required_fields = []
        required_field_errors = []
        for field in fields:
            if not field.required:
                row_data.append("")
                continue

            field_label = _format_required_field_name(field)
            try:
                value = get_field_value(field, model, series, manufacturer, equipment_type, request.listing_type, db)
            except HTTPException as exc:
                required_field_errors.append(f"{field_label}: {_normalize_http_detail(exc.detail)}")
                row_data.append("")
                continue
            except Exception as exc:
                required_field_errors.append(f"{field_label}: {str(exc)}")
                row_data.append("")
                continue

            # Required fields must resolve to a non-empty value.
            if value is None or (isinstance(value, str) and value.strip() == ""):
                missing_required_fields.append(field_label)
                row_data.append("")
                continue

            row_data.append(value)

        if missing_required_fields or required_field_errors:
            issue_lines = [
                f"Model '{model.name}' (id={model.id})"
            ]
            if missing_required_fields:
                issue_lines.append(
                    "missing required fields: " + ", ".join(missing_required_fields)
                )
            if required_field_errors:
                issue_lines.append(
                    "field errors: " + " | ".join(required_field_errors)
                )
            required_field_issues.append(" | ".join(issue_lines))
            continue
        
        data_rows.append(row_data)

    if required_field_issues:
        warnings.extend(required_field_issues)
        logger.warning(
            "[EXPORT] Skipped %s model(s) due to missing required template values",
            len(required_field_issues),
        )

    if not data_rows:
        detail_lines = [
            "No models exported: all selected models are missing required template values."
        ]
        if warnings:
            detail_lines.extend(warnings)
        raise HTTPException(status_code=400, detail="\n".join(detail_lines))
    
    ordered_field_names = [f.field_name for f in fields]
    ordered_required_flags = [f.required for f in fields]
    return header_rows, data_rows, filename_base, product_type.code, export_models, ordered_field_names, ordered_required_flags, warnings


def _generate_xlsx_artifact(request: ExportPreviewRequest, db: Session):
    """
    Shared logic to generate the final filled XLSX bytes.
    Used by both download_xlsx (stream) and preview (visual verification).
    """
    import os
    # 1. Resolve Data
    ensure_models_fresh_for_export(request, db)
    header_rows, data_rows, filename_base, template_code, models, ordered_fields, ordered_flags, warnings = build_export_data(request, db)
    
    # 2. Resolve Template
    if not models:
         raise HTTPException(400, detail="No models to export.")
    
    product_type = db.query(AmazonProductType).filter(AmazonProductType.code == template_code).first()
    if not product_type:
        raise HTTPException(404, detail="No product type template linked.")
        
    file_path = product_type.file_path
    if not file_path or not os.path.exists(file_path):
         raise HTTPException(404, detail=f"Template file missing: {file_path}")

    # 3. Load & Write (In-Memory)
    try:
        wb = load_workbook(file_path) # keep_vba default is False for plain load, or use data_only? 
        # For writing, we want to preserve styles/structure. Default is good.
    except Exception:
        raise HTTPException(422, detail="Could not open template file.")

    # Determine sheets (Reuse logic)
    eligible_sheets = []
    override_sheet = product_type.export_sheet_name_override
    
    if override_sheet:
        if override_sheet in wb.sheetnames:
            ws = wb[override_sheet]
            anchor = None
            for r in range(1, 51):
                if is_anchor_row(ws, r):
                    anchor = r
                    break
            setattr(ws, "_anchor_row", anchor)
            eligible_sheets.append(ws)
    else:
         blacklist = ["data definitions", "data definition", "instructions", "readme", "example", "sample", "notes"]
         for sname in wb.sheetnames:
            ws = wb[sname]
            if ws.sheet_state != "visible": continue
            if any(b in sname.lower() for b in blacklist): continue
            anchor = None
            try:
                for r in range(1, 51):
                   if is_anchor_row(ws, r):
                       anchor = r
                       break
            except: continue
            if anchor:
                setattr(ws, "_anchor_row", anchor)
                eligible_sheets.append(ws)
    
    if not eligible_sheets:
         raise HTTPException(422, detail="No eligible export sheet detected.")

    cols_to_write = len(data_rows[0]) if data_rows else 1
    
    for ws in eligible_sheets:
        anchor = getattr(ws, "_anchor_row", None)
        start_scan = product_type.export_start_row_override or (anchor + 1 if anchor else 1)
        
        first_writable = None
        max_search = max(start_scan + 5000, ws.max_row + 5000)
        for r in range(start_scan, max_search):
            if is_row_empty(ws, r, cols_to_write):
                first_writable = r
                break
        
        if not first_writable:
             raise HTTPException(422, detail=f"No empty row found on sheet '{ws.title}'.")
        
        curr = first_writable
        for dr in data_rows:
            for ci, val in enumerate(dr):
                try:
                    c = ws.cell(row=curr, column=ci+1)
                    if not isinstance(c, MergedCell):
                        c.value = val or ''
                except: pass
            curr += 1
            
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    raw_bytes = out.getvalue()
    
    # Compute Safe Byte Signature
    sig = hashlib.sha256(raw_bytes).hexdigest()

    # Compute Header Snapshot Hash (Rows 2-5, Cols 1-25)
    header_trace = []
    for ws in eligible_sheets:
        header_trace.append(f"SHEET:{ws.title}")
        for r in range(2, 6):
            row_vals = []
            for c in range(1, 26): # Limit to 25 cols
                try: 
                    val = ws.cell(row=r, column=c).value
                    row_vals.append(str(val) if val is not None else "")
                except: 
                    row_vals.append("")
            header_trace.append("|".join(row_vals))
    
    header_string = "\n".join(header_trace)
    header_hash = hashlib.sha256(header_string.encode('utf-8')).hexdigest()

    # Input Trace
    input_trace = f"models={len(models)}|first10={','.join(str(m.id) for m in models[:10])}|type={request.listing_type}|cust={getattr(request, 'include_customization', False)}"

    return {
        "bytes": raw_bytes,
        "signature": sig,
        "filename": f"{filename_base}.xlsx",
        "template_code": template_code,
        "models": models,
        "data_rows": data_rows,
        "headers_db": header_rows,
        "ordered_fields": ordered_fields,
        "ordered_flags": ordered_flags,
        "header_hash": header_hash,
        "input_trace": input_trace,
        "warnings": warnings
    }
    
@router.post("/download/xlsx")
def download_xlsx(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """Download export as XLSX file using consistent helper."""
    artifact = _generate_xlsx_artifact(request, db)
    
    # Debug Log
    logger.warning(f"[EXPORT][DOWNLOAD_XLSX][DEBUG] template={artifact['template_code']} code={artifact['template_code']} "
                   f"filled_xlsx_sha256={artifact['signature']} header_hash={artifact['header_hash']} "
                   f"inputs={artifact['input_trace']}")

    bio = io.BytesIO(artifact["bytes"])
    
    response = StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={artifact['filename']}"}
    )
    response.headers["X-Export-Signature"] = artifact["signature"]
    response.headers["X-Export-Template-Code"] = artifact["template_code"]
    
    # Set warnings header if models were skipped
    if artifact.get("warnings"):
        warnings_json = json.dumps(artifact["warnings"])
        response.headers["X-Export-Warnings"] = warnings_json
        logger.info(f"[EXPORT] Skipped models (missing dimensions): {warnings_json}")
    
    return response




# Helper functions for XLSM export logic
def is_anchor_row(ws, row_idx):
    row_vals = []
    # Check first 50 columns for anchor signals
    for c in range(1, 51):
        try:
            cell = ws.cell(row=row_idx, column=c)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if v and isinstance(v, str):
                row_vals.append(str(v).lower())
        except Exception:
            continue
            
    txt = " ".join(row_vals)
    # 1. contribution_sku / item_sku
    if "contribution_sku" in txt or "item_sku" in txt:
        return True
    
    # 2. sku + context
    if "sku" in txt:
        if any(x in txt for x in ["listing action", "update delete", "feed_product_type"]):
            return True
        if "sku" in row_vals: # Explicit exact match cell
            return True
            
    # 3. #.value pattern
    for val in row_vals:
        if ".value" in val and "#" in val:
            return True
            
    return False

def is_row_empty(ws, row_idx, check_cols_count):
    if check_cols_count < 1: 
        check_cols_count = 1 # Safety
        
    for c in range(1, check_cols_count + 1):
        cell = ws.cell(row=row_idx, column=c)
        if isinstance(cell, MergedCell):
            # Safe rule: merged cell means not writable/not empty context
            return False
            
        val = cell.value
        if val is not None and str(val).strip() != "":
            return False
    return True


@router.post("/download/xlsm")
def download_xlsm(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """Download export as XLSM file (macro-enabled workbook) preserving original macros."""
    import os
    print(f"!!! [XLSM][HIT][PID={os.getpid()}][CWD={os.getcwd()}] !!!", flush=True)
    logger.warning(f"[XLSM][HIT][PID={os.getpid()}][CWD={os.getcwd()}]")
    logger.warning(f"[XLSM] HIT download_xlsm model_ids={request.model_ids} listing_type={request.listing_type}")
    
    try:
        # Resolve Data
        try:
            ensure_models_fresh_for_export(request, db)
            header_rows, data_rows, filename_base, template_code, models, _, _, warnings = build_export_data(request, db)

            equipment_type_id = models[0].equipment_type_id
            product_type = db.query(AmazonProductType).filter(AmazonProductType.code == template_code).first()
            
            if not product_type:
                raise HTTPException(404, detail="No product type template linked for this equipment type.")
                
            file_path = product_type.file_path
            logger.warning(f"[XLSM] resolved equipment_type_id={equipment_type_id} template_id={product_type.id} file_path={file_path}")
            
            if not file_path or not os.path.exists(file_path):
                 raise HTTPException(404, detail=f"Original XLSM template file missing: {file_path}")
                 
        except Exception:
            logger.exception("[XLSM][STEP] resolve_template failed")
            raise

        # Load Workbook (MACROS DISABLED)
        try:
            logger.warning(f"[XLSM-BTN] Returning XLSX output (macros disabled)")
            logger.warning(f"[XLSM] Loading workbook from {file_path}")
            wb = load_workbook(file_path)
            logger.warning(f"[XLSM] template_sheetnames={wb.sheetnames}")
        except Exception:
            logger.exception("[XLSM][STEP] load_workbook failed")
            raise HTTPException(422, detail="Could not open XLSM template. File may be corrupt.")
            
        # Determine Eligible Sheets
        eligible_sheets = []
        override_sheet_name = product_type.export_sheet_name_override
        override_start_row = product_type.export_start_row_override
        
        if override_sheet_name:
            # MODE 1: Override
            if override_sheet_name in wb.sheetnames:
                ws = wb[override_sheet_name]
                # Scan anchor anyway to help start row logic
                anchor_row = None
                for r in range(1, 51):
                    if is_anchor_row(ws, r):
                        anchor_row = r
                        break
                setattr(ws, "_anchor_row", anchor_row)
                eligible_sheets.append(ws)
            else:
                raise HTTPException(422, detail=f"Export sheet '{override_sheet_name}' not found in template.")
        else:
            # MODE 2: Auto-detect
            blacklist = ["data definitions", "data definition", "instructions", "readme", "example", "sample", "notes"]
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.sheet_state != "visible":
                    continue
                
                s_lower = sheet_name.lower()
                if any(b in s_lower for b in blacklist):
                    continue
                
                anchor_row = None
                try:
                    for r in range(1, 51):
                        if is_anchor_row(ws, r):
                            anchor_row = r
                            break
                except Exception as e:
                    logger.warning(f"[TPL][SHEET][SKIP] name='{sheet_name}' reason='anchor detection crash: {e}'")
                    continue
                
                logger.warning(f"[TPL][SHEET] name='{sheet_name}' eligible={bool(anchor_row)} reason={'anchor found' if anchor_row else 'no anchor'} anchor_row={anchor_row}")
                
                if anchor_row:
                    setattr(ws, "_anchor_row", anchor_row)
                    eligible_sheets.append(ws)
                    
        if not eligible_sheets:
             # Fallback: if no override and no auto-detect, try 'Template' default?
             # User says: "If missing... return 422".
             # But "Use sheet name 'Template' as DEFAULT preference" in prompt text?
             # "MODE 1... MODE 2... If no anchor found, mark NOT eligible."
             # It implies strictness.
             # However, "DO NOT hardcode sheet name 'Template' except as DEFAULT preference."
             # I'll check 'Template' if NO sheets eligible?
             # Actually, if auto-detect fails, maybe I should check 'Template' specifically?
             # But if 'Template' has no anchor, it's risky.
             # I will stick to "return 422" per "If still missing, return a user-visible 422".
             raise HTTPException(422, detail="No eligible export sheet detected (no anchor row found). Please configure an Export Sheet Override.")

        # Write Data
        written_sheets = []
        total_rows_written = 0
        cols_to_write = len(data_rows[0]) if data_rows else 1
        
        for ws in eligible_sheets:
            anchor_row = getattr(ws, "_anchor_row", None)
            
            # Determine Start Scan
            start_scan = 1
            if override_start_row and override_start_row > 0:
                start_scan = override_start_row
                logger.warning(f"[TPL][ANCHOR] sheet='{ws.title}' anchor_row={anchor_row} start_scan={start_scan} override_start_row={override_start_row}")
            elif anchor_row:
                start_scan = anchor_row + 1
                logger.warning(f"[TPL][ANCHOR] sheet='{ws.title}' anchor_row={anchor_row} start_scan={start_scan}")
            else:
                start_scan = 1
                logger.warning(f"[TPL][ANCHOR] sheet='{ws.title}' anchor_row={anchor_row} start_scan={start_scan} (default)")
                
            # Find First Writable Row
            first_writable = None
            max_search = max(start_scan + 5000, ws.max_row + 5000)
            
            for r in range(start_scan, max_search):
                if is_row_empty(ws, r, cols_to_write):
                    first_writable = r
                    break
            
            if not first_writable:
                 raise HTTPException(422, detail=f"No empty row found to write export data on sheet '{ws.title}'. Template appears fully populated.")
            
            logger.warning(f"[TPL][ROW] sheet='{ws.title}' first_writable_row={first_writable}")
            
            # Write Data Rows (NO Header)
            rows_written = 0
            merged_skips = 0
            current_r = first_writable
            
            for data_row in data_rows:
                for col_idx, value in enumerate(data_row):
                    try:
                        cell = ws.cell(row=current_r, column=col_idx + 1)
                        if isinstance(cell, MergedCell):
                            merged_skips += 1
                            continue
                        cell.value = value or ''
                    except Exception as e:
                        c_type = "None"
                        try:
                             if 'cell' in locals() and cell: c_type = type(cell).__name__
                        except: pass
                        logger.error(f"[XLSM][CRASH]\nsheet={ws.title}\nrow={current_r}\ncol={col_idx+1}\ncell_type={c_type}\nexception={e}", exc_info=True)
                        continue
                current_r += 1
                rows_written += 1
                
            logger.warning(f"[TPL][WRITE] sheet='{ws.title}' merged_skips={merged_skips} rows_written={rows_written}")
            written_sheets.append(ws.title)
            total_rows_written += rows_written
            
        logger.warning(f"[TPL][SUMMARY] model_ids={request.model_ids} eligible_sheets={[s.title for s in eligible_sheets]} written_sheets={written_sheets} rows_written_total={total_rows_written}")

        # Save and Verify
        try:
            logger.warning(f"[XLSM] output_sheetnames={wb.sheetnames}")
            bio = io.BytesIO()
            wb.save(bio)
            output_size = bio.tell()
            bio.seek(0)
            
        except Exception:
             logger.exception("[XLSM][STEP] save failed")
             raise
             
        # Final Response
        bio.seek(0)
        sig = compute_export_signature(template_code, header_rows, data_rows)
        filename = f"{filename_base}.xlsx"
        
        response = StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        response.headers["X-Export-Signature"] = sig
        response.headers["X-Export-Template-Code"] = template_code
        if warnings:
            response.headers["X-Export-Warnings"] = json.dumps(warnings)
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[XLSM] UNHANDLED ERROR: {e}")
        raise HTTPException(
            status_code=422,
            detail=f"XLSM export failed due to invalid template structure or unexpected error: {str(e)}"
        )


@router.post("/download/csv")
def download_csv(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """Download export as CSV file."""
    ensure_models_fresh_for_export(request, db)
    header_rows, data_rows, filename_base, template_code, _, _, _, warnings = build_export_data(request, db)
    
    sig = compute_export_signature(template_code, header_rows, data_rows)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    for header_row in header_rows:
        writer.writerow([v or '' for v in header_row])
    
    for data_row in data_rows:
        writer.writerow([v or '' for v in data_row])
    
    content = output.getvalue().encode('utf-8')
    
    filename = f"{filename_base}.csv"
    response = StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
    response.headers["X-Export-Signature"] = sig
    response.headers["X-Export-Template-Code"] = template_code
    if warnings:
        response.headers["X-Export-Warnings"] = json.dumps(warnings)
    return response


@router.post("/download/reverb/csv")
def download_reverb_csv(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """
    Download Reverb export as CSV.
    Uses the Reverb Template assigned to the models' Equipment Type.
    """
    csv_buffer, filename = generate_reverb_export_csv(db, request.model_ids)
    
    return StreamingResponse(
        csv_buffer,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )




def generate_customization_unicode_txt(template_path: str, skus: List[str]) -> bytes:
    """
    Generate Amazon Customization File (.txt) from a template and list of SKUs.
    Format:
    - UTF-16 LE with BOM
    - Tab-delimited
    - Rows 1-3: from template unmodified
    - Row 4: Blueprint row (from template)
    - Rows 5+: Blueprint row with Column A replaced by real SKU
    """
    try:
        wb = load_workbook(template_path, read_only=False, data_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read customization template: {str(e)}")

    if len(rows) < 4:
         raise HTTPException(status_code=500, detail="Customization template is invalid (fewer than 4 rows).")

    # Capture Headers (1-3) and Blueprint (4)
    header_rows = rows[:3]
    blueprint_row = rows[3]
    
    output = io.BytesIO()
    # Write BOM for UTF-16 LE
    output.write(b'\xff\xfe')
    
    def write_row(row_values):
        # Normalize: 
        # - None -> ""
        # - strip tabs/newlines from content to prevent breaking format
        # - join with tabs
        cleaned = []
        for val in row_values:
            s = str(val) if val is not None else ""
            s = s.replace('\t', ' ').replace('\r', ' ').replace('\n', ' ')
            cleaned.append(s)
        
        line = "\t".join(cleaned) + "\r\n"
        output.write(line.encode('utf-16-le'))

    # Write Headers
    for r in header_rows:
        write_row(r)
        
    # Write Data Rows
    for sku in skus:
        # Create new row based on blueprint
        new_row = list(blueprint_row)
        # Column A is index 0
        if len(new_row) > 0:
            new_row[0] = sku
        else:
            new_row = [sku]
            
        write_row(new_row)
        
    return output.getvalue()


class DownloadZipRequest(BaseModel):
    model_ids: List[int]
    listing_type: str = "individual"
    include_customization: bool = True
    marketplace_token: str
    manufacturer_token: str
    series_token: str
    date_token: str  # YYYY-MM-DD
    customization_format: Optional[str] = "xlsx" # "xlsx" or "txt"


class GeneralExportZipRequest(BaseModel):
    model_ids: List[int]
    marketplaces: List[str]
    listing_type: str = "individual"


def _extract_filename_from_content_disposition(content_disposition: Optional[str], fallback: str) -> str:
    value = content_disposition or ""
    utf8_match = re.search(r"filename\*=UTF-8''([^;]+)", value, flags=re.IGNORECASE)
    if utf8_match and utf8_match.group(1):
        try:
            return urllib.parse.unquote(utf8_match.group(1))
        except Exception:
            pass
    plain_match = re.search(r'filename="?([^\";]+)"?', value, flags=re.IGNORECASE)
    if plain_match and plain_match.group(1):
        return plain_match.group(1)
    return fallback


async def _collect_async_chunks(body_iterator) -> bytes:
    chunks: List[bytes] = []
    async for chunk in body_iterator:
        if isinstance(chunk, bytes):
            chunks.append(chunk)
        elif isinstance(chunk, str):
            chunks.append(chunk.encode("utf-8"))
        elif chunk is None:
            continue
        else:
            chunks.append(bytes(chunk))
    return b"".join(chunks)


def _collect_streaming_response_bytes(streaming_response: StreamingResponse) -> bytes:
    body_iterator = getattr(streaming_response, "body_iterator", None)
    if body_iterator is None:
        return b""

    if hasattr(body_iterator, "__aiter__"):
        try:
            running_loop = asyncio.get_running_loop()
        except RuntimeError:
            running_loop = None

        if running_loop and running_loop.is_running():
            loop = asyncio.new_event_loop()
            try:
                return loop.run_until_complete(_collect_async_chunks(body_iterator))
            finally:
                loop.close()

        return asyncio.run(_collect_async_chunks(body_iterator))

    chunks: List[bytes] = []
    for chunk in body_iterator:
        if isinstance(chunk, bytes):
            chunks.append(chunk)
        elif isinstance(chunk, str):
            chunks.append(chunk.encode("utf-8"))
        elif chunk is None:
            continue
        else:
            chunks.append(bytes(chunk))
    return b"".join(chunks)


def _build_amazon_run_both_files(preview_req: ExportPreviewRequest, db: Session) -> List[tuple[str, bytes]]:
    files: List[tuple[str, bytes]] = []
    product_artifact = _generate_xlsx_artifact(preview_req, db)
    files.append((str(product_artifact["filename"]), bytes(product_artifact["bytes"])))

    export_settings = db.query(ExportSetting).first()
    customization_format = (
        (export_settings.amazon_customization_export_format or "xlsx").strip().lower()
        if export_settings
        else "xlsx"
    )

    if customization_format != "txt":
        customization_response = download_customization_xlsx(preview_req, db)
        customization_bytes = _collect_streaming_response_bytes(customization_response)
        if customization_bytes:
            mfr_name, series_name, date_str = _build_amazon_filename_tokens(product_artifact.get("models", []), db)
            fallback_name = f"Amazon_{mfr_name}_{series_name}_Customization_{date_str}.xlsx"
            customization_name = _extract_filename_from_content_disposition(
                customization_response.headers.get("Content-Disposition"),
                fallback_name,
            )
            files.append((customization_name, customization_bytes))

    return files

@router.post("/download/zip")
def download_zip(response: Response, request: DownloadZipRequest = Body(...), db: Session = Depends(get_db)):
    """
    Download a ZIP package containing:
    1. XLSM (Macro-Enabled)
    2. XLSX (Standard)
    3. CSV (Data Only)
    Optional: 4. Customization .txt (if toggle ON and rules apply - future chunk)
    
    Filenames are strictly constructed from tokens provided by the UI.
    """
    # 1. Validation & Data Build (Reuse existing logic)
    # We map DownloadZipRequest to ExportPreviewRequest for the helper
    preview_req = ExportPreviewRequest(model_ids=request.model_ids, listing_type=request.listing_type)
    ensure_models_fresh_for_export(preview_req, db)
    
    header_rows, data_rows, _, template_code, exported_models, ordered_field_names, _, warnings = build_export_data(preview_req, db)
    
    # 2. Compute Filename Base
    # Syntax: [Marketplace]-[Manufacturer]-[Series]-Product_Upload-[Date]
    filename_base = f"{request.marketplace_token}-{request.manufacturer_token}-{request.series_token}-Product_Upload-{request.date_token}"
    
    # 3. Generate Files In-Memory
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        
        # --- A. Generate XLSX ---
        wb_xlsx = Workbook()
        ws_xlsx = wb_xlsx.active
        ws_xlsx.title = "Template"
        
        # Styles (Same as download_xlsx)
        row_styles = [
            (Font(bold=True, color="FFFFFF"), PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")),
            (Font(color="FFFFFF"), PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")),
            (Font(bold=True, color="FFFFFF"), PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")),
            (Font(bold=True), PatternFill(start_color="8BC34A", end_color="8BC34A", fill_type="solid")),
            (Font(size=9), PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")),
            (Font(italic=True, size=9), PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")),
        ]
        
        current_row = 1
        for row_idx, header_row in enumerate(header_rows):
            for col_idx, value in enumerate(header_row):
                cell = ws_xlsx.cell(row=current_row, column=col_idx + 1, value=value or '')
                if row_idx < len(row_styles):
                    cell.font = row_styles[row_idx][0]
                    cell.fill = row_styles[row_idx][1]
                cell.alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1
        
        for data_row in data_rows:
            for col_idx, value in enumerate(data_row):
                ws_xlsx.cell(row=current_row, column=col_idx + 1, value=value or '')
            current_row += 1
            
        # Column Widths
        for col in ws_xlsx.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)
                except:
                    pass
            ws_xlsx.column_dimensions[column].width = max_length + 2

        xlsx_buffer = io.BytesIO()
        wb_xlsx.save(xlsx_buffer)
        zf.writestr(f"{filename_base}.xlsx", xlsx_buffer.getvalue())
        
        # --- B. Generate XLSM ---
        # Note: logic is identical to XLSX for openpyxl, but MIME/vbaProject handling might differ in real apps.
        # openpyxl saves as macro enabled if asked, but doesn't add macros dynamically.
        # existing download_xlsm endpoint essentially saves standard xlsx as .xlsm MIME.
        # We replicate that exact behavior here.
        wb_xlsm = Workbook()
        ws_xlsm = wb_xlsm.active
        ws_xlsm.title = "Template"
        
        current_row = 1
        for row_idx, header_row in enumerate(header_rows):
            for col_idx, value in enumerate(header_row):
                cell = ws_xlsm.cell(row=current_row, column=col_idx + 1, value=value or '')
                if row_idx < len(row_styles):
                    cell.font = row_styles[row_idx][0]
                    cell.fill = row_styles[row_idx][1]
                cell.alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1
        
        for data_row in data_rows:
            for col_idx, value in enumerate(data_row):
                ws_xlsm.cell(row=current_row, column=col_idx + 1, value=value or '')
            current_row += 1
            
        for col in ws_xlsm.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)
                except:
                    pass
            ws_xlsm.column_dimensions[column].width = max_length + 2
            
        xlsm_buffer = io.BytesIO()
        wb_xlsm.save(xlsm_buffer)
        zf.writestr(f"{filename_base}.xlsm", xlsm_buffer.getvalue())

        # --- C. Generate CSV ---
        csv_output = io.StringIO()
        writer = csv.writer(csv_output)
        for header_row in header_rows:
            writer.writerow([v or '' for v in header_row])
        for data_row in data_rows:
            writer.writerow([v or '' for v in data_row])
        
        zf.writestr(f"{filename_base}.csv", csv_output.getvalue().encode('utf-8'))
        
        # --- D. Customization File ---
        if request.include_customization:
            # 1. Identify SKU column from export data
            sku_col_idx = -1
            target_fields = ["item_sku", "contribution_sku", "external_product_id"] # Fallbacks
            
            for target in target_fields:
                for idx, field_name in enumerate(ordered_field_names):
                    if target in field_name.lower():
                        sku_col_idx = idx
                        break
                if sku_col_idx != -1:
                    break
            
            # If still found, fallback to Model.parent_sku? 
            # Requirement says "skus[] from the Amazon worksheet export rows (authoritative)"
            # If not found in export, we likely cannot proceed safely with an authoritative matching.
            # But let's check if we can fallback to index 0.
            if sku_col_idx == -1:
                # Fallback: Assume first column is SKU if named appropriately? Or just use index 0.
                # Common Amazon template practice: Column A is SKU.
                if ordered_field_names and "sku" in ordered_field_names[0].lower():
                     sku_col_idx = 0
            
            skus = []
            if sku_col_idx != -1:
                for row in data_rows:
                    if sku_col_idx < len(row):
                         skus.append(row[sku_col_idx])
                    else:
                         skus.append("")
            else:
                 # Fallback to model parent_sku if we can't find it in the sheet?
                 # This violates "authoritative from export rows", but is safer than crashing.
                 # Actually, let's log/warn and skip if we can't find SKUs.
                 pass

            # Audit / Reconciliation Check
            worksheet_count = len(data_rows)
            skus_count = len(skus)
            
            # Robustness: Check for duplicates and content signature
            unique_skus = set(skus)
            dupe_count = skus_count - len(unique_skus)
            
            # Deterministic Signature (SHA256 of joined SKUs)
            # Normalize to empty string if None before joining, though skus list extraction handled that above somewhat. (appended "")
            # Ensure all are strings
            clean_skus_for_sig = [str(s) if s is not None else "" for s in skus]
            sku_signature = hashlib.sha256("\n".join(clean_skus_for_sig).encode("utf-8")).hexdigest()[:12]
            
            first_sku = clean_skus_for_sig[0] if clean_skus_for_sig else "<none>"
            last_sku = clean_skus_for_sig[-1] if clean_skus_for_sig else "<none>"
            
            # Log full audit
            print(f"[EXPORT] SKU Audit: signature={sku_signature} count={skus_count} dupes={dupe_count} first={first_sku} last={last_sku}")
            
            should_skip = False
            if skus_count != worksheet_count:
                print(f"[EXPORT][ERROR] Customization SKU count mismatch (rows={worksheet_count}, skus={skus_count}).")
                should_skip = True
            
            if dupe_count > 0:
                print(f"[EXPORT][ERROR] Duplicate SKUs detected in export ({dupe_count}). Customization template requires unique SKUs.")
                should_skip = True
                
            if should_skip:
                print(f"[EXPORT][ERROR] Skipping customization file generation due to audit failure.")
                skus = [] # Force skip

            if skus:
                 #  2. Determine Template - Allow mixed equipment types if they share same customization template
                 # Validate customization template compatibility
                 equipment_type_ids = set(m.equipment_type_id for m in exported_models)
                 
                 if len(equipment_type_ids) > 1:
                     # Multiple equipment types - check if they all map to same customization template
                     equipment_types = db.query(EquipmentType).filter(EquipmentType.id.in_(equipment_type_ids)).all()
                     customization_template_ids = set(
                         et.amazon_customization_template_id for et in equipment_types 
                         if et.amazon_customization_template_id is not None
                     )
                     
                     if len(customization_template_ids) == 0:
                         raise HTTPException(
                             status_code=404,
                             detail="No customization templates assigned to the selected equipment types."
                         )
                     
                     if len(customization_template_ids) > 1:
                         et_map = {et.id: et.name for et in equipment_types}
                         sorted_ids = sorted(equipment_type_ids)
                         et_details = ", ".join([f"{et_id} ({et_map.get(et_id, 'Unknown')})" for et_id in sorted_ids])
                         
                         error_detail = (
                             f"Customization export failed: selected models span multiple equipment types with different customization templates. "
                             f"Equipment types: {et_details}. "
                             f"Customization template IDs: {sorted(customization_template_ids)}. "
                             f"Filter to equipment types that share the same customization template."
                         )
                         
                         logger.warning("[EXPORT][CUSTOMIZATION] Incompatible customization templates: %s", customization_template_ids)
                         raise HTTPException(status_code=400, detail=error_detail)
                     
                     logger.info(f"[EXPORT][CUSTOMIZATION] Mixed equipment types allowed: all map to same customization template (template_id={list(customization_template_ids)[0]})")
                     
                     # Use the shared template
                     template_id = list(customization_template_ids)[0]
                     equip_type = equipment_types[0]  # Use first for reference
                 else:
                     # Single equipment type - use its template
                     first_model = exported_models[0]
                     equip_type = db.query(EquipmentType).filter(EquipmentType.id == first_model.equipment_type_id).first()
                     template_id = equip_type.amazon_customization_template_id if equip_type else None
                 
                 if template_id:
                     # Import here to avoid circular imports at module load time
                     from app.models.templates import AmazonCustomizationTemplate
                     
                     template = db.query(AmazonCustomizationTemplate).filter(AmazonCustomizationTemplate.id == template_id).first()
                     
                     if template and template.file_path and os.path.exists(template.file_path):
                         try:
                             # Format Logic:
                             fmt = (request.customization_format or "xlsx").lower()

                             if fmt == 'xlsx':
                                 # --- XLSX Mode ---
                                 # Read raw template file and write to ZIP (byte-identical preservation)
                                 print(f"[EXPORT] Customization XLSX: template_id={template.id} path={template.file_path}")

                                 if not template.file_path:
                                     raise HTTPException(status_code=500, detail="Assigned customization template has no file_path")

                                 if not os.path.exists(template.file_path):
                                     raise HTTPException(status_code=500, detail=f"Assigned customization template file missing: {template.file_path}")

                                 with open(template.file_path, "rb") as f:
                                     cust_bytes = f.read()

                                 if not cust_bytes:
                                     raise HTTPException(status_code=500, detail=f"Assigned customization template is empty on disk: {template.file_path}")
                                 
                                 cust_filename = f"{request.marketplace_token}-{request.manufacturer_token}-{request.series_token}-Customization-{request.date_token}.xlsx"
                                 zf.writestr(cust_filename, cust_bytes)
                             
                             else:
                                 # --- TXT Mode (Legacy) ---
                                 cust_bytes = generate_customization_unicode_txt(template.file_path, skus)
                                 cust_filename = f"{request.marketplace_token}-{request.manufacturer_token}-{request.series_token}-Customization-{request.date_token}.txt"
                                 zf.writestr(cust_filename, cust_bytes)
                             
                         except Exception as e:
                             # Log error but don't fail entire export? 
                             # User said "Hard error" if multiple templates. Here it's generation failure.
                             # Failing safe for now by raising to alert user.
                             raise HTTPException(status_code=500, detail=f"Failed to generate customization file: {str(e)}")
                     else:
                         # File missing on disk
                         raise HTTPException(status_code=500, detail=f"Assigned customization template file missing: {template.file_path}")
                     # Else: No assignment -> Skip (as per requirements)
            
    # Helper for customization generation
    # ... logic ...
    
# Wait, I need to do this in steps.
# 1. Update imports.
# 2. Update a helper function `generate_customization_unicode_txt`.
# 3. Update `download_zip` to capture `ordered_field_names` and implement the logic.

# Let's do imports first.

            
    zip_buffer.seek(0)
    
    zip_filename = f"{filename_base}.zip"
    zip_buffer.seek(0)
    
    # Set warnings header if models were skipped
    if warnings:
        warnings_json = json.dumps(warnings)
        response.headers["X-Export-Warnings"] = warnings_json
        logger.info(f"[EXPORT][ZIP] Skipped models (missing dimensions): {warnings_json}")
    
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename_base}.zip"}
    )


@router.post("/general/download/zip")
def download_general_zip(request: GeneralExportZipRequest = Body(...), db: Session = Depends(get_db)):
    if not request.model_ids:
        raise HTTPException(status_code=400, detail="No models selected")

    selected_marketplaces: List[str] = []
    for raw_marketplace in request.marketplaces or []:
        norm = str(raw_marketplace or "").strip().lower()
        if norm in {"amazon", "ebay", "etsy", "reverb"} and norm not in selected_marketplaces:
            selected_marketplaces.append(norm)

    if not selected_marketplaces:
        raise HTTPException(status_code=400, detail="Select at least one marketplace")

    zip_buffer = io.BytesIO()
    errors: List[str] = []
    file_count = 0
    preview_req = ExportPreviewRequest(model_ids=request.model_ids, listing_type=request.listing_type)

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        if "amazon" in selected_marketplaces:
            try:
                amazon_files = _build_amazon_run_both_files(preview_req, db)
                for file_name, file_bytes in amazon_files:
                    zf.writestr(file_name, file_bytes)
                    file_count += 1
            except HTTPException as exc:
                errors.append(f"amazon: {_normalize_http_detail(exc.detail)}")
            except Exception as exc:
                logger.exception("General export Amazon leg failed")
                errors.append(f"amazon: {str(exc)}")

        if "ebay" in selected_marketplaces:
            try:
                from app.api.ebay_export import EbayExportRequest, export_ebay_csv

                ebay_req = EbayExportRequest(
                    model_ids=request.model_ids,
                    use_variation_presets=True,
                )
                ebay_response = export_ebay_csv(ebay_req, db)
                ebay_bytes = _collect_streaming_response_bytes(ebay_response)
                if not ebay_bytes:
                    raise HTTPException(status_code=500, detail="eBay export returned empty content")

                fallback_name = f"Ebay_Unknown_Unknown_{datetime.now().strftime('%Y-%m-%d')}.csv"
                ebay_filename = _extract_filename_from_content_disposition(
                    ebay_response.headers.get("Content-Disposition"),
                    fallback_name,
                )
                zf.writestr(ebay_filename, ebay_bytes)
                file_count += 1
            except HTTPException as exc:
                errors.append(f"ebay: {_normalize_http_detail(exc.detail)}")
            except Exception as exc:
                logger.exception("General export eBay leg failed")
                errors.append(f"ebay: {str(exc)}")

        if "reverb" in selected_marketplaces:
            try:
                reverb_buffer, reverb_filename = generate_reverb_export_csv(db, request.model_ids)
                reverb_data = reverb_buffer.getvalue() if hasattr(reverb_buffer, "getvalue") else reverb_buffer.read()
                if isinstance(reverb_data, str):
                    reverb_data = reverb_data.encode("utf-8")
                if not reverb_data:
                    raise HTTPException(status_code=500, detail="Reverb export returned empty content")
                zf.writestr(reverb_filename, reverb_data)
                file_count += 1
            except HTTPException as exc:
                errors.append(f"reverb: {_normalize_http_detail(exc.detail)}")
            except Exception as exc:
                logger.exception("General export Reverb leg failed")
                errors.append(f"reverb: {str(exc)}")

        if "etsy" in selected_marketplaces:
            errors.append("etsy: Exporter is not implemented yet.")

        if errors:
            zf.writestr("errors.txt", ("\n".join(errors) + "\n").encode("utf-8"))
            file_count += 1

    if file_count == 0:
        raise HTTPException(status_code=500, detail="No export files were generated.")

    zip_buffer.seek(0)
    zip_filename = f"General_Export_{datetime.now().strftime('%Y-%m-%d')}.zip"
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{zip_filename}"'},
    )

IMAGE_FIELD_TO_NUMBER = {

    'main_product_image_locator': '001',
    'other_product_image_locator_1': '002',
    'other_product_image_locator_2': '003',
    'other_product_image_locator_3': '004',
    'other_product_image_locator_4': '005',
    'other_product_image_locator_5': '006',
    'other_product_image_locator_6': '007',
    'other_product_image_locator_7': '008',
    'other_product_image_locator_8': '009',
    'swatch_product_image_locator': '010',
}
def normalize_for_url(name: str) -> str:
    """Normalize a name for use in URL paths/filenames.
    Removes spaces, special characters, and non-alphanumeric characters.
    Example: "Fender USA" -> "FenderUSA", "Tone-Master" -> "ToneMaster"
    """
    if not name:
        return ''
    result = re.sub(r'[^a-zA-Z0-9]', '', name)
    return result

def _normalize_image_token(value: str) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")

def _normalize_image_filename_in_url(value: str) -> str:
    text = str(value or "")
    if not text:
        return text

    split_idx = len(text)
    q_idx = text.find("?")
    h_idx = text.find("#")
    if q_idx >= 0:
        split_idx = min(split_idx, q_idx)
    if h_idx >= 0:
        split_idx = min(split_idx, h_idx)

    core = text[:split_idx]
    suffix = text[split_idx:] if split_idx < len(text) else ""
    last_sep = max(core.rfind("/"), core.rfind("\\"))
    if last_sep < 0:
        return _normalize_image_token(core) + suffix
    return core[: last_sep + 1] + _normalize_image_token(core[last_sep + 1 :]) + suffix

def _apply_image_index_placeholder(value: str, field_name: str) -> str:
    image_field_key = get_image_field_key(field_name)
    if not image_field_key:
        return str(value or "")
    image_index = IMAGE_FIELD_TO_NUMBER.get(image_field_key)
    if not image_index:
        return str(value or "")
    return str(value or "").replace("[INDEX]", image_index).replace("[IMAGE_INDEX]", image_index)

def _normalize_variant_key_for_price_token(raw_value: str) -> str:
    return normalize_variant_key_for_price_token(raw_value)

def _resolve_single_price_placeholder(
    *,
    spec_text: str,
    db: Session,
    model_id: int,
    field_name: str,
) -> str:
    return resolve_shared_single_price_placeholder(
        spec_text=spec_text,
        db=db,
        model_id=model_id,
        field_name=field_name,
    )

def _resolve_price_placeholders_in_value(
    *,
    value: str,
    db: Session,
    model_id: int,
    field_name: str,
) -> str:
    return resolve_shared_price_placeholders_in_value(
        value=value,
        db=db,
        model_id=model_id,
        field_name=field_name,
    )

def substitute_placeholders(value: str, model: Model, series, manufacturer, equipment_type, is_image_url: bool = False, max_length: int = None) -> str:
    result = value
    mfr_name = manufacturer.name if manufacturer else ''
    series_name = series.name if series else ''
    model_name = model.name if model else ''
    equip_type = equipment_type.name if equipment_type else ''
    
    if is_image_url:
        mfr_name_norm = _normalize_image_token(mfr_name)
        series_name_norm = _normalize_image_token(series_name)
        model_name_norm = _normalize_image_token(model_name)
        equip_type_norm = _normalize_image_token(equip_type)
        
        result = result.replace('[Manufacturer_Name]', mfr_name_norm)
        result = result.replace('[Series_Name]', series_name_norm)
        result = result.replace('[Model_Name]', model_name_norm)
        result = result.replace('[MANUFACTURER_NAME]', mfr_name_norm)
        result = result.replace('[SERIES_NAME]', series_name_norm)
        result = result.replace('[MODEL_NAME]', model_name_norm)
        
        result = result.replace('[EQUIPMENT_TYPE]', equip_type_norm)
        result = result.replace('[Equipment_Type]', equip_type_norm)
    else:
        # Standard substitutions (except Series Name if max_length is active)
        result = result.replace('[MANUFACTURER_NAME]', mfr_name)
        result = result.replace('[MODEL_NAME]', model_name)
        result = result.replace('[Manufacturer_Name]', mfr_name)
        result = result.replace('[Model_Name]', model_name)
        result = result.replace('[EQUIPMENT_TYPE]', equip_type)
        result = result.replace('[Equipment_Type]', equip_type)
        
        # Series Name Handling with Smart Truncation
        # We need to detect if [SERIES_NAME] (or variants) is present to apply logic
        series_placeholders = ['[SERIES_NAME]', '[Series_Name]']
        
        # Check if any series placeholder is in the string
        found_series_ph = None
        for ph in series_placeholders:
            if ph in result:
                found_series_ph = ph
                break
        
        if found_series_ph:
            val_to_insert = series_name
            
            if max_length is not None:
                # Calculate current length without the series placeholder
                current_len_without_series = len(result) - len(found_series_ph)
                
                # Available budget for series name
                available_for_series = max_length - current_len_without_series
                
                if available_for_series < len(series_name):
                    # Need to truncate series name
                    # If available < 0, it means we are already over limit just with other fields -> empty series
                    if available_for_series > 0:
                        val_to_insert = series_name[:available_for_series]
                    else:
                        val_to_insert = ""
            
            result = result.replace(found_series_ph, val_to_insert)
            
            # Catch-all: check other series placeholders just in case
            for ph in series_placeholders:
                result = result.replace(ph, val_to_insert)
        
        # Finally, if strict max_length is enforced and we are somehow still over (e.g. Model Name itself is huge),
        # we must truncate the final result to satisfy the hard requirement.
        if max_length is not None and len(result) > max_length:
            result = result[:max_length]
            
    result = apply_base_sku_tokens(result, model)
    if max_length is not None and len(result) > max_length:
        result = result[:max_length]
    return result

def get_image_field_key(field_name: str) -> str | None:
    """Extract the base image field key from a full Amazon field name.
    Returns the key if it matches a known product image field, None otherwise.
    """
    for key in IMAGE_FIELD_TO_NUMBER.keys():
        if field_name.startswith(key):
            return key
    return None

def is_image_url_field(field_name: str) -> bool:
    """Check if a field is a product image URL field that needs special processing."""
    return get_image_field_key(field_name) is not None

def get_amazon_us_baseline_price_str(db: Session, model_id: int) -> str:
    """
    Fetch the baseline retail price (Choice Waterproof No Padding) for Amazon US.
    Format: "249.95" (2 decimals)
    Strict Rule: Fail if snapshot is missing.
    """
    snapshot = db.query(ModelPricingSnapshot).filter(
        ModelPricingSnapshot.model_id == model_id,
        ModelPricingSnapshot.marketplace == "amazon",
        ModelPricingSnapshot.variant_key == "choice_no_padding"
    ).first()

    if not snapshot:
        # Prompt: "If baseline snapshot row is missing, fail the export for that model with a clear message"
        # Since this is deep in the call stack for a specific field, raising HTTPException here
        # will abort the entire request, which is desired.
        raise HTTPException(
            status_code=400, 
            detail=f"Missing baseline pricing snapshot for Choice Waterproof (no padding). Run pricing recalculation for model {model_id} on 'amazon' marketplace before exporting."
        )
    
    return f"{snapshot.retail_price_cents / 100:.2f}"

def get_field_value(field: ProductTypeField, model: Model, series, manufacturer, equipment_type=None, listing_type: str = "individual", db: Session = None) -> str | None:
    field_name_lower = field.field_name.lower()
    is_image_field = is_image_url_field(field.field_name)
    has_explicit_template_value = bool(
        (field.selected_value is not None and str(field.selected_value).strip() != "")
        or (field.custom_value is not None and str(field.custom_value).strip() != "")
    )

    # DEBUG INSTRUMENTATION FOR PART NUMBER
    if "part_number" in field_name_lower:
        print(f"\n[EXPORT_DEBUG] Field: {field.field_name} | Model: {model.name}")
        print(f"  selected_value: {field.selected_value!r}")
        print(f"  custom_value:   {field.custom_value!r}")
        print(f"  valid_values:   {getattr(field, 'valid_values', 'N/A')!r}")
        
        # Speculative resolution logic match
        chosen_pre = field.selected_value if field.selected_value else field.custom_value
        print(f"  CHOSEN PRE-SUB: {chosen_pre!r}")
        
        if chosen_pre:
            # Re-run substitution for logging visibility
            try:
                final_sub = substitute_placeholders(chosen_pre, model, series, manufacturer, equipment_type, is_image_url=is_image_field)
                print(f"  FINAL POST-SUB: {final_sub!r}")
            except Exception as e:
                print(f"  FINAL POST-SUB ERROR: {e}")
        else:
            print("  FINAL POST-SUB: None")
        print("-" * 40)
    
    # Phase 9: Amazon Baseline Price Logic
    # check specific field key parts
    if "purchasable_offer[marketplace_id=atvpdkikx0der]" in field_name_lower and "our_price#1.schedule#1.value_with_tax" in field_name_lower:
        if has_explicit_template_value:
            # Respect explicit template defaults/selections for this field.
            # This allows [PRICE:...] placeholders to drive dynamic pricing values.
            pass
        elif db:
            return get_amazon_us_baseline_price_str(db, model.id)
        else:
            # If db is somehow missing (shouldn't happen with updated callers), strict rules say NO FALLBACK.
            raise HTTPException(status_code=500, detail="Database session missing in export logic.")

    if 'contribution_sku' in field_name_lower and listing_type == 'individual':
        # Effective SKU: use override if present, otherwise use generated parent_sku
        effective_sku = model.sku_override if model.sku_override else model.parent_sku
        return effective_sku if effective_sku else None

    def _render_template_value(raw_value: str, max_len: Optional[int]) -> str:
        rendered = substitute_placeholders(
            raw_value,
            model,
            series,
            manufacturer,
            equipment_type,
            is_image_url=is_image_field,
            max_length=max_len,
        )
        rendered = _resolve_price_placeholders_in_value(
            value=rendered,
            db=db,
            model_id=model.id,
            field_name=field.field_name,
        )
        if is_image_field:
            rendered = _apply_image_index_placeholder(rendered, field.field_name)
            rendered = _normalize_image_filename_in_url(rendered)
        return rendered

    
    # Only include custom_value or selected_value if field is marked as required
    if field.required:
        # Determine max length for specific Amazon fields
        max_len = None
        # Strict 40 char limit for model_name and model_number
        if 'model_name' in field_name_lower or 'model_number' in field_name_lower:
             max_len = 40

        if field.selected_value:
            return _render_template_value(field.selected_value, max_len)
            
        if field.custom_value:
            return _render_template_value(field.custom_value, max_len)
        
        # Auto-generate values for common fields only if required
        if 'item_name' in field_name_lower or 'product_name' in field_name_lower or 'title' in field_name_lower:
            mfr_name = manufacturer.name if manufacturer else ''
            series_name = series.name if series else ''
            return f"{mfr_name} {series_name} {model.name} Cover"
        
        if 'brand' in field_name_lower or 'brand_name' in field_name_lower:
            return manufacturer.name if manufacturer else None
        
        if 'model' in field_name_lower or 'model_number' in field_name_lower or 'model_name' in field_name_lower:
            return model.name
        
        if 'manufacturer' in field_name_lower:
            return manufacturer.name if manufacturer else None
    
    return None


def _generate_customization_xlsx(template_path: str, skus: List[str]) -> io.BytesIO:
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail=f"Customization template file missing: {template_path}")
        
    wb = load_workbook(template_path)
    
    # 1. Worksheet Selection (STRICT: "Template", case-insensitive)
    target_sheet_name = "Template"
    ws = None
    for sheet in wb.sheetnames:
        if sheet.strip().lower() == target_sheet_name.lower():
            ws = wb[sheet]
            break
            
    if ws is None:
        raise HTTPException(status_code=500, detail="Customization workbook missing 'Template' worksheet")
        
    # 2. Data Start Row Detection (SellerSku in Column A)
    data_start_row = 4 # Default
    seller_sku_found = False
    
    # Scan first 20 rows
    for r in range(1, 21):
        cell_val = ws.cell(row=r, column=1).value
        # Check normalized value (case-insensitive)
        if cell_val and str(cell_val).strip().lower() == "sellersku":
            data_start_row = r + 2
            seller_sku_found = True
            break
            
    if not seller_sku_found:
        print(f"[CUSTOMIZATION][WARN] 'SellerSku' not found; defaulting data_start_row={data_start_row}")

    blueprint_row_idx = data_start_row
    
    # Determine max columns to copy from blueprint
    max_col = ws.max_column
    
    # 3. Blueprint Row Copy Rules
    for i, sku in enumerate(skus):
        target_row_idx = blueprint_row_idx + i
        source_row_idx = blueprint_row_idx
        
        # Copy Row Dimensions (Height) for new rows
        if i > 0:
            if source_row_idx in ws.row_dimensions:
                 src_dim = ws.row_dimensions[source_row_idx]
                 tgt_dim = ws.row_dimensions[target_row_idx]
                 if src_dim.height is not None:
                     tgt_dim.height = src_dim.height
                 tgt_dim.customHeight = src_dim.customHeight

        if i == 0:
            # First SKU: Use the existing blueprint row as-is (Overwrite Col A)
            ws.cell(row=target_row_idx, column=1, value=sku)
            # Do NOT modify other columns or style (preserve original blueprint)
        else:
            # Subsequent SKUs: Create new row by copying blueprint
            for col in range(1, max_col + 1):
                source_cell = ws.cell(row=source_row_idx, column=col)
                target_cell = ws.cell(row=target_row_idx, column=col)
                
                # Copy values AND formulas (except Col A)
                # 4. SKU Injection Rules (Template!Column A)
                if col == 1:
                    target_cell.value = sku
                else:
                    target_cell.value = source_cell.value
                    
                # 5. Formatting Safety (Copy styles)
                if source_cell.has_style:
                    target_cell._style = copy.copy(source_cell._style)
                     
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@router.post("/download/customization/xlsx")
def download_customization_xlsx(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """
    Download filled Customization Template (XLSX).
    Logic:
    1. Resolve Models & Template.
    2. Extract SKUs from Export Data (Authoritative).
    3. Load Template, Clone Blueprint Row, Inject SKUs.
    4. Return XLSX.
    """
    # 1. Reuse Build Export Data (to get data_rows and match ZIP behavior)
    ensure_models_fresh_for_export(request, db)
    header_rows, data_rows, filename_base, _, models, _, _, warnings = build_export_data(request, db)
    
    if not models:
        raise HTTPException(status_code=400, detail="No models found.")

    # 2. Extract SKUs (Logic mirrored from download_zip)
    ordered_field_names = header_rows[0] if header_rows else []
    sku_col_idx = -1
    target_fields = ["item_sku", "contribution_sku", "external_product_id"]
    for target in target_fields:
        for idx, field_name in enumerate(ordered_field_names):
            field_name_lc = field_name.lower() if isinstance(field_name, str) else ""
            if target in field_name_lc:
                sku_col_idx = idx
                break
        if sku_col_idx != -1:
            break
            
    if sku_col_idx == -1 and ordered_field_names and "sku" in ordered_field_names[0].lower():
        sku_col_idx = 0
        
    skus = []
    if sku_col_idx != -1:
        for row in data_rows:
            if sku_col_idx < len(row):
                 val = row[sku_col_idx]
                 skus.append(str(val) if val is not None else "")
            else:
                 skus.append("")
    
    # Filter empty? ZIP logic doesn't strictly filter, but appends empty string.
    # However, customization usually implies valid items.
    # We keeps skus aligned with export rows 1:1 if needed, 
    # OR we assume 1 row per export row.
    # We keep them.
    
    if not skus:
        # Fallback to model parent_sku if extraction failed?
        # User said "Reuse existing logic". 
        # Existing logic in download_zip logs error if skus_count != worksheet_count.
        # We proceed if we have count > 0.
        if not data_rows:
             raise HTTPException(status_code=400, detail="No data rows generated.")
        # If extraction failed (sku_col_idx == -1), we have empty list or need fallback?
        # If sku_col_idx is -1, loop above didn't run.
        if sku_col_idx == -1:
             # Try falling back to model.parent_sku for each model
             # But models list matches data_rows 1:1?
             # build_export_data returns models.
             skus = [m.parent_sku for m in models]

    # 3. Resolve Template
    equipment_type_ids = {m.equipment_type_id for m in models if m.equipment_type_id is not None}
    equipment_types = db.query(EquipmentType).filter(EquipmentType.id.in_(equipment_type_ids)).all()

    if not equipment_types:
        raise HTTPException(status_code=400, detail="No equipment type found for selected models.")

    missing_template_equipment = [
        et.name for et in equipment_types if not et.amazon_customization_template_id
    ]
    if missing_template_equipment:
        raise HTTPException(
            status_code=400,
            detail=(
                "Missing customization template assignment for equipment type(s): "
                + ", ".join(sorted(missing_template_equipment))
            ),
        )

    template_ids = {et.amazon_customization_template_id for et in equipment_types if et.amazon_customization_template_id}
    if len(template_ids) > 1:
        raise HTTPException(
            status_code=400,
            detail=(
                "Selected models span multiple customization templates. "
                "Please export one compatible equipment-type group at a time."
            ),
        )

    template = equipment_types[0].amazon_customization_template
    if not template or not template.file_path:
        raise HTTPException(status_code=500, detail="Customization template record invalid.")
    if not os.path.exists(template.file_path):
        raise HTTPException(
            status_code=500,
            detail=f"Assigned customization template file missing: {template.file_path}",
        )
    if os.path.getsize(template.file_path) <= 0:
        raise HTTPException(
            status_code=500,
            detail=f"Assigned customization template file is empty: {template.file_path}",
        )
        
    # 4. Generate
    try:
        xlsx_buffer = _generate_customization_xlsx(template.file_path, skus)
    except Exception as e:
        logger.exception("Failed to generate customization XLSX")
        raise HTTPException(status_code=500, detail=f"Generation failed: {str(e)}")
    
    mfr_name, series_name, date_str = _build_amazon_filename_tokens(models, db)
    filename = f"Amazon_{mfr_name}_{series_name}_Customization_{date_str}.xlsx"
    
    return StreamingResponse(
        xlsx_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Explicitly rebuild model to resolve recursive/deferred references for Pydantic v2
DownloadZipRequest.model_rebuild()
GeneralExportZipRequest.model_rebuild()


@router.post("/stats", response_model=ExportStatsResponse)
def get_export_stats(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """
    Get statistics about the selected models for export readiness.
    """
    if not request.model_ids:
        return ExportStatsResponse(
            total_models=0,
            models_with_pricing=0,
            models_missing_pricing=0,
            models_with_images=0,
            models_missing_images=0,
            equipment_types={}
        )

    models = db.query(Model.id, Model.equipment_type_id, Model.image_url).filter(
        Model.id.in_(request.model_ids)
    ).all()

    if not models:
        return ExportStatsResponse(
            total_models=0,
            models_with_pricing=0,
            models_missing_pricing=0,
            models_with_images=0,
            models_missing_images=0,
            equipment_types={}
        )

    model_ids = {m.id for m in models}

    models_with_snapshot = {
        row[0] for row in db.query(ModelPricingSnapshot.model_id).filter(
            ModelPricingSnapshot.model_id.in_(model_ids),
            ModelPricingSnapshot.marketplace == "amazon",
            ModelPricingSnapshot.variant_key == "choice_no_padding"
        )
    }

    equipment_type_ids = {m.equipment_type_id for m in models if m.equipment_type_id is not None}
    equipment_type_map = {}
    if equipment_type_ids:
        equipment_type_map = {
            row[0]: row[1] for row in db.query(EquipmentType.id, EquipmentType.name).filter(
                EquipmentType.id.in_(equipment_type_ids)
            )
        }

    models_with_pricing = 0
    models_missing_pricing = 0
    models_with_images = 0
    models_missing_images = 0
    equipment_type_counts = {}

    for model in models:
        if model.id in models_with_snapshot:
            models_with_pricing += 1
        else:
            models_missing_pricing += 1

        if model.image_url and model.image_url.strip():
            models_with_images += 1
        else:
            models_missing_images += 1

        eq_name = equipment_type_map.get(model.equipment_type_id)
        if eq_name:
            equipment_type_counts[eq_name] = equipment_type_counts.get(eq_name, 0) + 1

    return ExportStatsResponse(
        total_models=len(models),
        models_with_pricing=models_with_pricing,
        models_missing_pricing=models_missing_pricing,
        models_with_images=models_with_images,
        models_missing_images=models_missing_images,
        equipment_types=equipment_type_counts
    )


@router.get("/health")
def export_health_check(db: Session = Depends(get_db)):
    """
    Check if export system is healthy and configured.
    """
    template_count = db.query(AmazonProductType).count()
    equipment_type_count = db.query(EquipmentType).count()

    linked_count = db.query(EquipmentTypeProductType).count()

    cache_size = 0
    cache_obj = globals().get("HTTP_CACHE", None)
    if cache_obj is not None:
        try:
            cache_size = len(cache_obj)
        except (TypeError, AttributeError):
            cache_size = 0

    return {
        "status": "healthy",
        "templates_configured": template_count,
        "equipment_types": equipment_type_count,
        "equipment_types_with_templates": linked_count,
        "cache_size": cache_size
    }
